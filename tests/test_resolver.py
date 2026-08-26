"""The workbook resolver, the contract checks and the continuity state.

Config names a folder and a contract, never a file, a row count or a column letter.
These tests pin the behaviours that replace those constants: the highest version in
the folder is read (date with or without dashes), everything else is explained,
the header must match exactly, required sheets are a subset, every data row carries
its formula, and a file that goes backwards relative to the last run is refused.
"""

from __future__ import annotations

import os
import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from efe.cli import main
from efe.models import (
    ContinuityError,
    DriveSyncError,
    SchemaMismatchError,
    VersionConflictError,
)
from efe.workbook.reader import header_diff, load_workbook_view
from efe.workbook.resolve import highest_version_present, parse_version, resolve_workbook
from efe.workbook.state import WorkbookState, header_hash, load_state, save_state, state_path
from efe.workbook.writer import CHANGELOG_DETAIL_HEADERS, next_version_path, write_enriched

BASE = "EFE_Alpine_Partner_Database"


def _stub(path: Path, size: int = 5000) -> Path:
    path.write_bytes(b"x" * size)
    return path


# ---------------------------------------------------------------------------
# Resolver
# ---------------------------------------------------------------------------


def test_parse_version_accepts_both_date_spellings():
    assert parse_version(Path(f"20260824_{BASE}_v05.xlsx"), BASE) == 5
    assert parse_version(Path(f"2026-08-21_{BASE}_v04.xlsx"), BASE) == 4
    assert parse_version(Path(f"2026-08-21_{BASE}_v04_SUPERSEDED.xlsx"), BASE) is None
    assert parse_version(Path("notes.xlsx"), BASE) is None


def test_highest_version_wins_and_everything_else_is_explained(
    synthetic_config, synthetic_workbook
):
    folder = synthetic_config.workbook_directory
    v05 = shutil.copy(synthetic_workbook, folder / f"20260824_{BASE}_v05.xlsx")  # no dashes
    _stub(folder / f"2026-08-21_{BASE}_v04.xlsx")
    _stub(folder / f"2026-08-21_{BASE}_v09_SUPERSEDED.xlsx")
    _stub(folder / f"~$2026-08-21_{BASE}_v09.xlsx")
    _stub(folder / f"2026-08-21_{BASE}_v09.xlsx", size=10)  # Drive placeholder
    _stub(folder / "notes.xlsx")

    res = resolve_workbook(folder, basename=BASE, min_bytes=4000)

    assert res.chosen is not None
    assert res.chosen.path == Path(v05)
    assert res.chosen.version == 5
    assert res.chosen.date == "2026-08-24"
    reasons = dict(res.rejected)
    assert "SUPERSEDED" in reasons[f"2026-08-21_{BASE}_v09_SUPERSEDED.xlsx"]
    assert "lock" in reasons[f"~$2026-08-21_{BASE}_v09.xlsx"]
    assert "min_plausible_bytes" in reasons[f"2026-08-21_{BASE}_v09.xlsx"]
    assert "does not match" in reasons["notes.xlsx"]
    assert "below the highest" in reasons[f"2026-08-21_{BASE}_v04.xlsx"]
    assert "below the highest" in reasons[synthetic_workbook.name]
    assert "chosen" in res.describe() and "skipped" in res.describe()

    view = load_workbook_view(synthetic_config)
    assert view.path == Path(v05)
    assert view.version == 5
    assert view.resolution is not None and not view.resolution.override


def test_same_version_tie_goes_to_the_newest_file(synthetic_config, synthetic_workbook):
    folder = synthetic_config.workbook_directory
    # The OLDER file sorts last by name, so only mtime can pick the newer one.
    older = Path(shutil.copy(synthetic_workbook, folder / f"20260821_{BASE}_v06.xlsx"))
    newer = Path(shutil.copy(synthetic_workbook, folder / f"2026-08-20_{BASE}_v06.xlsx"))
    os.utime(older, (1_700_000_000, 1_700_000_000))
    os.utime(newer, (1_700_100_000, 1_700_100_000))

    res = resolve_workbook(folder, basename=BASE, min_bytes=4000)

    assert res.chosen is not None and res.chosen.path == newer
    assert "older" in dict(res.rejected)[older.name]


def test_truly_empty_folder_says_so(tmp_path):
    empty = tmp_path / "empty"
    empty.mkdir()
    with pytest.raises(DriveSyncError) as exc:
        resolve_workbook(empty, basename=BASE, min_bytes=4000)
    assert str(empty) in str(exc.value)
    assert "(empty)" in str(exc.value)


def test_folder_without_a_workbook_lists_what_it_holds(tmp_path):
    empty = tmp_path / "empty"
    empty.mkdir()
    _stub(empty / "unrelated.xlsx")
    _stub(empty / f"2026-08-25_{BASE}_v06.xlsm")  # wrong extension is named, not hidden
    with pytest.raises(DriveSyncError) as exc:
        resolve_workbook(empty, basename=BASE, min_bytes=4000)
    text = str(exc.value)
    assert str(empty) in text
    assert "unrelated.xlsx" in text
    assert f"2026-08-25_{BASE}_v06.xlsm  -- not an .xlsx file" in text
    assert "Nothing has been changed" in text


def test_missing_folder_is_a_drive_error(tmp_path):
    with pytest.raises(DriveSyncError, match="Workbook folder not found"):
        resolve_workbook(tmp_path / "nowhere", basename=BASE, min_bytes=4000)


def test_explicit_workbook_override_skips_the_scan(synthetic_config, synthetic_workbook):
    folder = synthetic_config.workbook_directory
    shutil.copy(synthetic_workbook, folder / f"2026-08-22_{BASE}_v03.xlsx")

    view = load_workbook_view(synthetic_config, synthetic_workbook)

    assert view.path == synthetic_workbook
    assert view.version == 1
    assert view.resolution is not None and view.resolution.override


def test_highest_version_present_counts_every_file(synthetic_config):
    folder = synthetic_config.workbook_directory
    out = synthetic_config.output_directory
    _stub(folder / f"2026-08-21_{BASE}_v07.xlsx", size=10)  # even a placeholder counts
    _stub(out / f"2026-08-21_{BASE}_v03.xlsx")
    highest, where = highest_version_present([folder, out], BASE)
    assert highest == 7 and where is not None and where.name.endswith("_v07.xlsx")


# ---------------------------------------------------------------------------
# Contract
# ---------------------------------------------------------------------------


def test_header_diff_is_positional_and_names_the_strays():
    assert header_diff(["A", "B"], ["A", "C", "D"]) == [
        "B: expected 'B', found 'C'",
        "C: unexpected extra column 'D'",
        "names missing entirely: ['B']",
        "names not in the contract: ['C', 'D']",
    ]
    assert header_diff(["A", "B"], ["A", "B"]) == []


def test_header_drift_aborts_with_a_diff(synthetic_config, synthetic_workbook):
    """Someone inserted a column in Sheets: every letter after it would be wrong."""
    wb = load_workbook(synthetic_workbook)
    ws = wb["PARTNERS"]
    ws.insert_cols(5)
    ws.cell(1, 5).value = "Notes"
    wb.save(synthetic_workbook)
    wb.close()

    with pytest.raises(SchemaMismatchError) as exc:
        load_workbook_view(synthetic_config)
    text = str(exc.value)
    assert "E: expected 'Resort_Base', found 'Notes'" in text
    assert "names not in the contract: ['Notes']" in text
    assert "Nothing has been changed" in text


def test_required_sheets_are_a_subset(synthetic_config, synthetic_workbook):
    wb = load_workbook(synthetic_workbook)
    wb.create_sheet("NOTES")
    wb.save(synthetic_workbook)
    wb.close()
    assert "NOTES" in load_workbook_view(synthetic_config).sheets  # extra sheet tolerated

    wb = load_workbook(synthetic_workbook)
    del wb["CHANGELOG_DETAIL"]
    wb.save(synthetic_workbook)
    wb.close()
    with pytest.raises(SchemaMismatchError, match=r"required sheets missing: \['CHANGELOG_DETAIL'"):
        load_workbook_view(synthetic_config)


def test_formula_column_must_hold_a_formula_on_every_row(synthetic_config, synthetic_workbook):
    wb = load_workbook(synthetic_workbook)
    wb["PARTNERS"]["AC5"] = "14"
    wb.save(synthetic_workbook)
    wb.close()
    expected = (
        r"Next_Follow_Up must hold a formula on every data row; 1 row\(s\) do not, e.g. \[5\]"
    )
    with pytest.raises(SchemaMismatchError, match=expected):
        load_workbook_view(synthetic_config)


def test_letters_are_derived_from_the_header_not_config(synthetic_config):
    view = load_workbook_view(synthetic_config)  # binds letters from the sheet's header
    spec = synthetic_config.workbook
    for logical, name in spec.columns.items():
        assert spec.column_for(logical) == view.header_letters[name]
    assert spec.column_for("id") == "A"
    assert spec.column_for("general_email") == "I"
    assert spec.column_for("round") == "AM"
    assert spec.writable_letters == ["I", "J", "K", "L", "M", "N", "O", "P", "U"]
    assert spec.provenance_letters == ["AK", "AL", "AM"]
    assert spec.crm_letters == [
        "Z",
        "AA",
        "AB",
        "AC",
        "AD",
        "AE",
        "AF",
        "AG",
        "AH",
        "AI",
        "AJ",
        "AN",
    ]
    assert spec.formula_letters == ["AC"]
    assert spec.header_of("AC") == "Next_Follow_Up"
    with pytest.raises(ValueError, match="not in the bound header"):
        spec.letter_of("Nope")


# ---------------------------------------------------------------------------
# Continuity
# ---------------------------------------------------------------------------


def test_state_is_recorded_after_a_load(synthetic_config):
    view = load_workbook_view(synthetic_config, command="test")
    saved = load_state(state_path(synthetic_config.state_directory))
    assert saved is not None
    assert saved.file == view.path.name
    assert saved.version == 1
    assert saved.data_rows == 8
    assert saved.header_sha256 == header_hash(view.header)
    assert saved.command == "test"
    assert saved.file_sha256 == view.fingerprint  # the bytes read are on record
    assert view.previous_state is None  # first run: nothing to compare with


def test_version_regression_is_refused_unless_reset(synthetic_config):
    save_state(
        state_path(synthetic_config.state_directory),
        WorkbookState.now(
            file=f"2026-08-24_{BASE}_v05.xlsx",
            version=5,
            data_rows=8,
            header=list(synthetic_config.workbook.header),
            command="test",
        ),
    )
    with pytest.raises(ContinuityError, match="version went backwards"):
        load_workbook_view(synthetic_config)

    view = load_workbook_view(synthetic_config, reset_state=True)
    assert view.version == 1
    assert load_state(state_path(synthetic_config.state_directory)).version == 1


def test_header_change_between_runs_is_refused(synthetic_config):
    save_state(
        state_path(synthetic_config.state_directory),
        WorkbookState.now(
            file=synthetic_config.workbook.header[0],
            version=1,
            data_rows=8,
            header=["Something", "Else"],
            command="test",
        ),
    )
    with pytest.raises(ContinuityError, match="header differs from the baseline"):
        load_workbook_view(synthetic_config)


# ---------------------------------------------------------------------------
# Writer: symmetric output
# ---------------------------------------------------------------------------


def test_output_version_is_detected_plus_one_and_resolves_next_time(synthetic_config):
    from tests.test_workbook import change

    view = load_workbook_view(synthetic_config)
    assert view.version == 1
    out = write_enriched(
        synthetic_config,
        view,
        [change(2, "P", "instagram_handle", "@summitlodgeverbier")],
        run_id="test-sym",
    )
    assert out.name.endswith(f"_{BASE}_v02.xlsx")
    assert out.parent == synthetic_config.output_directory

    # Same folder for input and output (the production layout): the next run
    # reads the file just written, with no config edit.
    synthetic_config.output_dir = None
    moved = shutil.move(str(out), synthetic_config.workbook_directory / out.name)
    nxt = load_workbook_view(synthetic_config)
    assert nxt.path == Path(moved)
    assert nxt.version == 2
    with pytest.raises(VersionConflictError, match="v02 already exists"):
        next_version_path(synthetic_config, 1)
    assert next_version_path(synthetic_config, nxt.version).name.endswith("_v03.xlsx")


def test_writer_appends_to_the_existing_audit_sheet(synthetic_config):
    from tests.test_workbook import change

    view = load_workbook_view(synthetic_config)
    out = write_enriched(
        synthetic_config,
        view,
        [change(2, "P", "instagram_handle", "@summitlodgeverbier")],
        run_id="test-append",
    )
    wb = load_workbook(out)
    try:
        assert wb.sheetnames.count("CHANGELOG_DETAIL") == 1
        ws = wb["CHANGELOG_DETAIL"]
        assert [ws.cell(1, i).value for i in range(1, 16)] == CHANGELOG_DETAIL_HEADERS
        assert ws.cell(2, 2).value == "20260821-100000"  # the prior run's row, untouched
        assert ws.cell(3, 2).value == "test-append"  # new rows follow it directly
        log = wb["CHANGELOG"]
        assert log.cell(4, 1).value == "v01"
        assert log.cell(5, 1).value == "v02"  # right after the last real entry
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# `efe check`
# ---------------------------------------------------------------------------


def test_check_prints_the_full_resolution_report(synthetic_config, capsys, monkeypatch):
    monkeypatch.setenv("COLUMNS", "200")
    rc = main(["--config", str(synthetic_config.config_path), "check"])
    out = capsys.readouterr().out
    assert rc == 0
    assert "chosen" in out and f"_{BASE}_v01.xlsx" in out
    assert "exact and in order" in out
    assert "holds a formula on all 8 data rows" in out
    assert "first run" in out
    assert "rows selected for enrichment" in out

    # Second run: the baseline is now known and reported.
    rc = main(["--config", str(synthetic_config.config_path), "check"])
    out = capsys.readouterr().out
    assert rc == 0
    assert "baseline" in out and "not behind the baseline" in out


def test_check_explains_a_regression_instead_of_a_traceback(synthetic_config, capsys, monkeypatch):
    monkeypatch.setenv("COLUMNS", "200")
    save_state(
        state_path(synthetic_config.state_directory),
        WorkbookState.now(
            file=f"2026-08-24_{BASE}_v05.xlsx",
            version=5,
            data_rows=277,
            header=list(synthetic_config.workbook.header),
            command="test",
        ),
    )
    rc = main(["--config", str(synthetic_config.config_path), "check"])
    out = capsys.readouterr().out
    assert rc == 5
    assert "REFUSED" in out
    assert "version went backwards" in out
    assert "fewer data rows than last run: 277" in out
    assert "--reset-state" in out


def test_writer_refuses_an_input_that_changed_since_it_was_read(
    synthetic_config, synthetic_workbook
):
    """A Drive re-sync between load and write must stop the run, never be applied."""
    from efe.models import VerificationError
    from tests.test_workbook import change

    view = load_workbook_view(synthetic_config)
    wb = load_workbook(synthetic_workbook)
    wb["PARTNERS"]["D2"] = "edited meanwhile"
    wb.save(synthetic_workbook)
    wb.close()

    with pytest.raises(VerificationError, match="changed since it was read"):
        write_enriched(
            synthetic_config,
            view,
            [change(2, "P", "instagram_handle", "@summitlodgeverbier")],
            run_id="test-drift",
        )
    assert not list(synthetic_config.output_directory.glob("*.xlsx"))


# ---------------------------------------------------------------------------
# Review findings, pinned
# ---------------------------------------------------------------------------


def test_single_digit_date_is_still_a_version(synthetic_config, synthetic_workbook):
    folder = synthetic_config.workbook_directory
    v03 = Path(shutil.copy(synthetic_workbook, folder / f"2026-8-4_{BASE}_v03.xlsx"))
    res = resolve_workbook(folder, basename=BASE, min_bytes=4000)
    assert res.chosen is not None and res.chosen.path == v03 and res.chosen.date == "2026-08-04"
    assert parse_version(Path(f"bogus-prefix_{BASE}_v09.xlsx"), BASE) == 9  # any prefix counts


def test_same_version_landing_mid_run_is_refused_at_write_time(
    synthetic_config, synthetic_workbook
):
    """The destination is derived before the (long) fetch; the conflict scan must run
    again right before the copy, or a human's same-version export gets shadowed."""
    from tests.test_workbook import change

    view = load_workbook_view(synthetic_config)
    destination = next_version_path(synthetic_config, view.version)  # v02, fine now
    shutil.copy(
        synthetic_workbook, synthetic_config.workbook_directory / f"20260826_{BASE}_v02.xlsx"
    )

    with pytest.raises(VersionConflictError, match="v02 already exists"):
        write_enriched(
            synthetic_config,
            view,
            [change(2, "P", "instagram_handle", "@summitlodgeverbier")],
            run_id="test-shadow",
            output_path=destination,
        )
    assert not destination.exists()


def test_unreadable_baseline_is_an_error_not_a_first_run(synthetic_config):
    path = state_path(synthetic_config.state_directory)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text('{"file": "x", "version": 5, "data_rows": 277, ', encoding="utf-8")

    with pytest.raises(ContinuityError, match="unreadable"):
        load_workbook_view(synthetic_config)
    view = load_workbook_view(synthetic_config, reset_state=True)
    assert view.version == 1
    assert load_state(path).data_rows == 8  # rewritten, atomically, as the new baseline


def test_unversioned_override_never_lowers_the_baseline(
    synthetic_config, synthetic_workbook, tmp_path
):
    save_state(
        state_path(synthetic_config.state_directory),
        WorkbookState.now(
            file=f"2026-08-24_{BASE}_v05.xlsx",
            version=5,
            data_rows=8,
            header=list(synthetic_config.workbook.header),
            command="test",
        ),
    )
    copy = Path(shutil.copy(synthetic_workbook, tmp_path / "EFE_copy.xlsx"))
    with pytest.raises(ContinuityError, match="carries no version"):
        load_workbook_view(synthetic_config, copy)

    view = load_workbook_view(synthetic_config, copy, reset_state=True)
    assert view.version == 0
    saved = load_state(state_path(synthetic_config.state_directory))
    assert saved is not None and saved.version == 5  # untouched: v00 would disarm the guard


def test_values_that_look_like_formulas_are_refused(synthetic_config):
    from efe.models import VerificationError
    from tests.test_workbook import change

    view = load_workbook_view(synthetic_config)
    with pytest.raises(VerificationError, match="live formula"):
        write_enriched(
            synthetic_config,
            view,
            [change(2, "U", "commission_terms", "=10% commission on net rates")],
            run_id="test-formula",
        )


def test_array_formula_covers_its_rows(synthetic_config, synthetic_workbook):
    from openpyxl.worksheet.formula import ArrayFormula

    wb = load_workbook(synthetic_workbook)
    ws = wb["PARTNERS"]
    for row in range(3, 10):
        ws.cell(row, 29).value = None
    ws["AC2"] = ArrayFormula("AC2:AC9", '=IFERROR(AA2:AA9+AB2:AB9,"")')
    wb.save(synthetic_workbook)
    wb.close()
    assert load_workbook_view(synthetic_config).data_rows == 8


def test_efe_verify_accepts_a_legitimate_output(synthetic_config, capsys, monkeypatch):
    from tests.test_workbook import change

    monkeypatch.setenv("COLUMNS", "200")
    view = load_workbook_view(synthetic_config)
    out = write_enriched(
        synthetic_config,
        view,
        [change(2, "P", "instagram_handle", "@summitlodgeverbier")],
        run_id="test-verify",
    )
    rc = main(
        [
            "--config",
            str(synthetic_config.config_path),
            "verify",
            str(out),
            "--against",
            str(view.path),
            "--allow-partners-changes",
        ]
    )
    text = capsys.readouterr().out
    assert rc == 0, text
    assert "fidelity gate: PASS" in text

    # ... but a real change to a pre-existing changelog cell is still caught.
    wb = load_workbook(out)
    wb["CHANGELOG"]["C4"] = "history rewritten"
    wb.save(out)
    wb.close()
    rc = main(
        [
            "--config",
            str(synthetic_config.config_path),
            "verify",
            str(out),
            "--against",
            str(view.path),
            "--allow-partners-changes",
        ]
    )
    assert rc == 1


def test_enrich_records_the_emitted_version_as_baseline(synthetic_config, capsys, monkeypatch):
    """The only writer of the post-run baseline is cmd_enrich; pin it end to end."""
    import efe.cli as cli
    from efe.pipeline import RunOutcome
    from tests.test_workbook import change

    monkeypatch.setenv("COLUMNS", "200")

    async def fake_run(cfg, view, candidates, **kwargs):
        outcome = RunOutcome()
        outcome.changes = [change(2, "P", "instagram_handle", "@summitlodgeverbier")]
        return outcome

    monkeypatch.setattr(cli, "run_enrichment", fake_run)
    rc = main(["--config", str(synthetic_config.config_path), "enrich", "--limit", "1"])
    text = capsys.readouterr().out
    assert rc == 0, text
    saved = load_state(state_path(synthetic_config.state_directory))
    assert saved is not None
    assert saved.version == 2
    assert saved.file.endswith(f"_{BASE}_v02.xlsx")
    assert saved.data_rows == 8
    assert saved.command == "efe enrich (output)"

    synthetic_config.output_dir = None  # production layout: same folder
    rc = main(["--config", str(synthetic_config.config_path), "check"])
    text = capsys.readouterr().out
    assert rc == 0, text
    assert "not behind the baseline" in text


def test_dry_run_keeps_its_own_resume_ledger(synthetic_config, capsys, monkeypatch):
    from efe.cli import ledger_for, report_resumed

    monkeypatch.setenv("COLUMNS", "200")
    dry = ledger_for(synthetic_config, "R3-hotels", "run-1", dry_run=True)
    real = ledger_for(synthetic_config, "R3-hotels", "run-2", dry_run=False)
    assert dry.progress_path.name == "progress-R3-hotels-dryrun.json"
    assert real.progress_path.name == "progress-R3-hotels.json"
    dry.mark_complete({"EFE-0001"})
    assert real.completed_entities() == set()  # the real run starts untouched

    assert report_resumed(27, 0, real, "R3-hotels") == 27
    text = capsys.readouterr().out
    assert "27 of 27 rows skipped" in text and "--fresh" in text
    assert report_resumed(27, 27, real, "R3-hotels") == 0
    assert capsys.readouterr().out.strip() == ""


def test_input_without_cached_results_is_written_not_refused(synthetic_config, synthetic_workbook):
    """A file last saved by openpyxl carries no cached formula results. Nothing can be
    reinjected, nothing is lost, and the tripwire must not fire."""
    from tests.test_workbook import change

    stripped = synthetic_config.workbook_directory / f"2026-08-22_{BASE}_v03.xlsx"
    wb = load_workbook(synthetic_workbook)
    wb.save(stripped)  # openpyxl drops every cached <v>
    wb.close()

    view = load_workbook_view(synthetic_config)
    assert view.path == stripped and view.version == 3
    assert view.formula_cells == 11 and view.cached_results == 0
    assert "penpyxl" in view.last_writer
    out = write_enriched(
        synthetic_config,
        view,
        [change(2, "P", "instagram_handle", "@summitlodgeverbier")],
        run_id="test-nocache",
    )
    assert out.name.endswith("_v04.xlsx")

    # ... while an input WITH cached results still trips the wire if none come back.
    cached_view = load_workbook_view(synthetic_config, synthetic_workbook, reset_state=True)
    assert cached_view.cached_results == 11


def test_typed_value_on_a_human_verified_row_is_reported_not_refused(
    synthetic_config, synthetic_workbook, capsys, monkeypatch
):
    """Row 7 is Contacted = YES: its Next_Follow_Up may be a typed date. Row 5 is not."""
    monkeypatch.setenv("COLUMNS", "200")
    wb = load_workbook(synthetic_workbook)
    wb["PARTNERS"]["AC7"] = "2026-09-08"
    wb.save(synthetic_workbook)
    wb.close()
    view = load_workbook_view(synthetic_config)
    assert view.schema is not None
    assert view.schema.formula_overrides == {"Next_Follow_Up": [7]}
    assert view.schema.formula_gaps == {}

    rc = main(["--config", str(synthetic_config.config_path), "check"])
    text = capsys.readouterr().out
    assert rc == 0
    assert "holds a formula on 7 data rows" in text and "1 human-verified row(s)" in text

    wb = load_workbook(synthetic_workbook)
    wb["PARTNERS"]["AC5"] = "2026-09-08"
    wb.save(synthetic_workbook)
    wb.close()
    with pytest.raises(SchemaMismatchError, match=r"1 row\(s\) do not, e.g. \[5\]"):
        load_workbook_view(synthetic_config)


def test_rows_marked_duplicate_of_are_not_reported_again(synthetic_config, synthetic_workbook):
    from efe.dedupe import find_duplicates

    wb = load_workbook(synthetic_workbook)
    ws = wb["PARTNERS"]
    ws["A10"], ws["B10"], ws["C10"] = "EFE-0009", "Summit Lodge Verbier", "1. Hotels"
    ws["H10"], ws["Z10"], ws["AI10"] = "https://summitlodge.example", "NO", "Duplicate of EFE-0001"
    ws["AC10"] = '=IFERROR(AA10+AB10,"")'
    wb.save(synthetic_workbook)
    wb.close()
    assert find_duplicates(load_workbook_view(synthetic_config), synthetic_config) == []

    wb = load_workbook(synthetic_workbook)
    wb["PARTNERS"]["AI10"] = "Not started"
    wb.save(synthetic_workbook)
    wb.close()
    assert len(find_duplicates(load_workbook_view(synthetic_config), synthetic_config)) == 1
