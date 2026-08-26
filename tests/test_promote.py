"""Promotion: candidate CSV rows become PARTNERS rows, through the fidelity gate."""

from __future__ import annotations

import csv
from pathlib import Path

import pytest
from openpyxl import load_workbook

from efe.cli import main
from efe.models import VerificationError
from efe.workbook.promote import (
    inspect_ranges,
    name_keys,
    plan_promotion,
    read_candidates,
    registrable,
    write_promoted,
)
from efe.workbook.reader import load_workbook_view
from efe.workbook.state import load_state, state_path
from efe.workbook.verify import compare, snapshot
from tests.conftest import PARTNERS_HEADERS


def _csv(path: Path, rows: list[dict[str, str]], columns: list[str] | None = None) -> Path:
    columns = columns or PARTNERS_HEADERS[:39]  # the discovery CSV predates Material_Sent
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({c: row.get(c, "") for c in columns})
    return path


def _candidate(entity_id: str, name: str, website: str, **extra) -> dict[str, str]:
    row = {
        "ID": entity_id,
        "Entity_Name": name,
        "Category": "1. Hotels",
        "Resort_Base": "Seefeld",
        "Region_Valley": "Tirol",
        "Country": "AT",
        "Website_URL": website,
        "General_Email": "TBD",
        "Contacted": "NO",
        "Follow_Up_Days": "14",
        "Status": "Not started",
        "Next_Action": "TBD",
        "Source_URL": "https://www.seefeld.com/de/hotels.html",
        "Date_Verified": "2026-08-25",
        "Round": "R3-discovery",
        "Strategic_Fit_Note": "[CANDIDATO] seefeld.com - 4*S",
    }
    row.update(extra)
    return row


def _plan(cfg, view, source):
    return plan_promotion(view, cfg, read_candidates(source, cfg), source)


# ---------------------------------------------------------------------------
# Matching helpers
# ---------------------------------------------------------------------------


def test_name_keys_see_through_umlauts_suffixes_and_parentheses():
    a = name_keys("Schlosshotel Kitzbühel – THE SPA MOMENT")
    b = name_keys("Schlosshotel Kitzbuehel (ex A-ROSA)")
    assert "schlosshotel kitzbuehel" in a and "schlosshotel kitzbuehel" in b
    assert name_keys("Hotel Post Lech") != name_keys("Hotel Post Seefeld")


def test_registrable_domain():
    assert registrable("de.kristiania.at") == "kristiania.at"
    assert registrable("kristiania.at") == "kristiania.at"
    assert registrable("booking.example.co.uk") == "example.co.uk"


# ---------------------------------------------------------------------------
# Planning
# ---------------------------------------------------------------------------


def test_plan_keeps_new_rows_and_explains_every_rejection(synthetic_config, tmp_path):
    view = load_workbook_view(synthetic_config)
    source = _csv(
        tmp_path / "cands.csv",
        [
            _candidate("EFE-0359", "Hotel Seespitz", "https://seespitz.at"),
            _candidate(
                "EFE-0360", "Summit Lodge Verbier", "https://other.example", Resort_Base="Verbier"
            ),
            _candidate(
                "EFE-0361", "Grandclass Twin", "https://grandclass.example"
            ),  # domain exists
            _candidate("EFE-0001", "Fresh Name", "https://fresh.example"),  # ID exists
            _candidate("EFE-0362", "Hotel Kristall", "https://hotel-kristall.at"),
            _candidate("EFE-0362", "Hotel Kristall Bis", "https://kristall-bis.at"),  # ID repeated
            _candidate("EFE-0363", "Hotel Seespitz", "https://seespitz-two.at"),  # name repeated
            _candidate("EFE-0364", "Chalet Belle Étoile – Résidence", "TBD"),  # accent + suffix
            _candidate("EFE-0365", "Web Only", "https://de.summitlodge.example"),  # subdomain
            _candidate("EFE-0005", "Low Id", "https://low.example"),  # not above highest ID
            _candidate("efe-0366", "Bad Id", "https://bad.example"),  # malformed
        ],
    )
    plan = _plan(synthetic_config, view, source)

    assert [r for r, _ in plan.accepted] == [10, 11]  # data ends at row 9
    assert plan.ids == ["EFE-0359", "EFE-0362"]
    reasons = {entity_id: reason for entity_id, _, reason in plan.rejected}
    assert reasons["EFE-0360"] == "name already in PARTNERS (row 2)"
    assert "domain grandclass.example already in PARTNERS" in reasons["EFE-0361"]
    assert "not above the sheet's highest ID EFE-0008" in reasons["EFE-0001"]
    assert "repeated in the CSV" in reasons["EFE-0363"]
    assert reasons["EFE-0364"] == "name already in PARTNERS (row 4)"
    assert "is under summitlodge.example, already in PARTNERS (row 2)" in reasons["EFE-0365"]
    assert "not above" in reasons["EFE-0005"]
    assert "not of the form EFE-dddd" in reasons["efe-0366"]
    assert sum(1 for i, _, r in plan.rejected if i == "EFE-0362") == 1
    assert "skipped : EFE-0001" in plan.describe()


def test_same_name_with_own_site_in_another_resort_is_kept_with_a_notice(
    synthetic_config, tmp_path
):
    view = load_workbook_view(synthetic_config)
    source = _csv(
        tmp_path / "cands.csv",
        [
            _candidate(
                "EFE-0359",
                "Summit Lodge Verbier",
                "https://summit-seefeld.at",
                Resort_Base="Seefeld",
            )
        ],
    )
    plan = _plan(synthetic_config, view, source)
    assert plan.ids == ["EFE-0359"]
    assert any("name matches row 2" in n and "kept" in n for n in plan.notices)


def test_ranges_are_reported_not_extended(synthetic_config, tmp_path):
    view = load_workbook_view(synthetic_config)
    ranges = inspect_ranges(view.path, synthetic_config)
    assert ranges == {
        "autofilter_last_row": 9,
        "validation_last_row": 9,
        "formula_reach_last_row": 9,
    }
    source = _csv(
        tmp_path / "c.csv", [_candidate("EFE-0359", "Hotel Seespitz", "https://seespitz.at")]
    )
    plan = plan_promotion(
        view, synthetic_config, read_candidates(source, synthetic_config), source, ranges
    )
    notes = "\n".join(plan.notices)
    assert "read PARTNERS only down to row 9; rows 10..10 would NOT be counted" in notes
    assert "cached totals" in notes
    assert (
        "autofilter ends at row 9" in notes and "data validations (dropdowns) end at row 9" in notes
    )


def test_csv_shape_is_strict(synthetic_config, tmp_path):
    bad = tmp_path / "bad.csv"
    bad.write_text(
        "ID,Entity_Name,Website_URL\nEFE-0359,Hotel A, B,https://a.at\n", encoding="utf-8"
    )
    with pytest.raises(VerificationError, match="line 2: 4 fields, the header has 3"):
        read_candidates(bad, synthetic_config)
    twice = tmp_path / "twice.csv"
    twice.write_text("ID,Entity_Name,ID\nEFE-0359,Hotel A,EFE-0359\n", encoding="utf-8")
    with pytest.raises(VerificationError, match="names a column twice"):
        read_candidates(twice, synthetic_config)
    unknown = _csv(
        tmp_path / "unk.csv",
        [_candidate("EFE-0359", "X", "https://x.at")],
        ["ID", "Entity_Name", "Bogus"],
    )
    with pytest.raises(VerificationError, match="not PARTNERS columns"):
        read_candidates(unknown, synthetic_config)


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------


def test_promoted_rows_land_after_the_data_with_the_live_formula(synthetic_config, tmp_path):
    view = load_workbook_view(synthetic_config)
    source = _csv(
        tmp_path / "cands.csv",
        [
            _candidate(
                "EFE-0359",
                "Hotel Seespitz",
                "https://seespitz.at",
                Phone="0043512123",
                Priority_Score="3",
            ),
            _candidate(
                "EFE-0360",
                "Hotel Kristall",
                "https://hotel-kristall.at",
                Capacity_Keys_or_Beds="120",
            ),
        ],
    )
    plan = _plan(synthetic_config, view, source)
    out = write_promoted(synthetic_config, view, plan, run_id="test-promote")
    assert out.name.endswith("_v02.xlsx")

    wb = load_workbook(out)
    try:
        ws = wb["PARTNERS"]
        assert ws["A10"].value == "EFE-0359" and ws["B10"].value == "Hotel Seespitz"
        assert ws["K10"].value == "0043512123"  # leading zero kept: text, not a number
        assert ws["Y10"].value == 3 and ws["AB10"].value == 14  # numeric columns become numbers
        assert ws["S11"].value == "120"  # capacity is text in the sheet
        assert ws["AC10"].value == '=IFERROR(AA10+AB10,"")'
        assert ws["AC11"].value == '=IFERROR(AA11+AB11,"")'
        assert ws["Z10"].value == "NO" and ws["AM10"].value == "R3-discovery"
        assert ws["AN10"].value is None  # Material_Sent: not in the CSV, left blank
        for row in range(2, 10):  # existing rows untouched
            assert ws.cell(row, 1).value == f"EFE-{row - 1:04d}"
        log = wb["CHANGELOG"]
        assert log["A5"].value == "v02" and "2 candidate rows appended" in log["C5"].value
        assert "dominant pattern (8 rows)" in log["C5"].value and "cached totals" in log["C5"].value
        assert log["D5"].value == "efe promote"
        detail = wb["CHANGELOG_DETAIL"]
        assert detail["B3"].value == "test-promote" and detail["G3"].value == "row"
        assert detail["D4"].value == "EFE-0360"
    finally:
        wb.close()

    before, after = snapshot(view.path), snapshot(out)
    partners = {k for k in set(before.values) | set(after.values) if k[0] == "PARTNERS"}
    logs = {
        k
        for k in after.values
        if k[0] in ("CHANGELOG", "CHANGELOG_DETAIL") and k not in before.values
    }
    assert (
        compare(
            before,
            after,
            allowed_value_changes=partners | logs,
            allowed_autofilter_changes={"CHANGELOG_DETAIL"},
        )
        == []
    )

    synthetic_config.output_dir = None
    moved = Path(out).replace(synthetic_config.workbook_directory / out.name)
    nxt = load_workbook_view(synthetic_config)
    assert nxt.path == moved and nxt.version == 2 and nxt.data_rows == 10


def test_a_value_on_a_padded_row_stops_the_run(synthetic_config, synthetic_workbook, tmp_path):
    """Sheets pads the used range; a note typed on a 'blank' row is data."""
    wb = load_workbook(synthetic_workbook)
    wb["PARTNERS"]["AI11"] = "Waiting for reply"
    wb.save(synthetic_workbook)
    wb.close()
    view = load_workbook_view(synthetic_config)
    assert view.data_rows == 8  # the stray value is invisible to the row count...
    source = _csv(
        tmp_path / "c.csv",
        [
            _candidate("EFE-0359", "Hotel Seespitz", "https://seespitz.at"),
            _candidate("EFE-0360", "Hotel Kristall", "https://hotel-kristall.at"),
        ],
    )
    plan = _plan(synthetic_config, view, source)
    with pytest.raises(
        VerificationError, match=r"rows 10..11 are not empty.*AI11='Waiting for reply'"
    ):
        write_promoted(
            synthetic_config, view, plan, run_id="test-stray"
        )  # ...but not to the writer
    assert not list(synthetic_config.output_directory.glob("*.xlsx"))


def test_formula_comes_from_the_dominant_pattern_not_the_nearest_row(
    synthetic_config, synthetic_workbook, tmp_path
):
    wb = load_workbook(synthetic_workbook)
    wb["PARTNERS"]["AC9"] = "=AA9+AB9"  # a hand-edited last row
    wb.save(synthetic_workbook)
    wb.close()
    view = load_workbook_view(synthetic_config)
    source = _csv(
        tmp_path / "c.csv", [_candidate("EFE-0359", "Hotel Seespitz", "https://seespitz.at")]
    )
    out = write_promoted(
        synthetic_config, view, _plan(synthetic_config, view, source), run_id="test-dom"
    )
    wb = load_workbook(out)
    try:
        assert wb["PARTNERS"]["AC10"].value == '=IFERROR(AA10+AB10,"")'
        assert "7 rows, 1 other form(s) ignored" in wb["CHANGELOG"]["C5"].value
    finally:
        wb.close()


def test_nothing_to_promote_is_refused(synthetic_config, tmp_path):
    view = load_workbook_view(synthetic_config)
    source = _csv(
        tmp_path / "dupes.csv",
        [_candidate("EFE-0001", "Summit Lodge Verbier", "https://summitlodge.example")],
    )
    plan = _plan(synthetic_config, view, source)
    assert "no rows to append" in plan.describe()
    with pytest.raises(VerificationError, match="nothing to promote"):
        write_promoted(synthetic_config, view, plan, run_id="test-none")
    assert not list(synthetic_config.output_directory.glob("*.xlsx"))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def test_cli_promote_dry_run_then_write(synthetic_config, tmp_path, capsys, monkeypatch):
    monkeypatch.setenv("COLUMNS", "200")
    source = _csv(
        tmp_path / "cands.csv",
        [
            _candidate("EFE-0359", "Hotel Seespitz", "https://seespitz.at"),
            _candidate("EFE-0001", "Fresh Name", "https://fresh.example"),
        ],
    )
    rc = main(["--config", str(synthetic_config.config_path), "promote", str(source), "--dry-run"])
    text = capsys.readouterr().out
    assert rc == 0, text
    assert "1 row(s) to append" in text and "EFE-0001" in text and "Dry run" in text
    assert not list(synthetic_config.output_directory.glob("*.xlsx"))
    dry_report = next(synthetic_config.artifacts_directory.glob("DRYRUN_promotion_*.md"))
    assert "DRY RUN" in dry_report.read_text(encoding="utf-8").splitlines()[2]

    rc = main(["--config", str(synthetic_config.config_path), "promote", str(source)])
    text = capsys.readouterr().out
    assert rc == 0, text
    written = list(synthetic_config.output_directory.glob("*_v02.xlsx"))
    assert len(written) == 1
    report = next(p for p in synthetic_config.artifacts_directory.glob("promotion_*.md"))
    assert f"WRITTEN: {written[0].name}" in report.read_text(encoding="utf-8")
    saved = load_state(state_path(synthetic_config.state_directory))
    assert saved is not None and saved.version == 2 and saved.data_rows == 9
    assert saved.command == "efe promote (output)"
    assert saved.file_sha256  # re-fingerprinted in the ledger


def test_cli_promote_with_nothing_to_do_writes_nothing_and_says_so(
    synthetic_config, tmp_path, capsys, monkeypatch
):
    monkeypatch.setenv("COLUMNS", "200")
    source = _csv(
        tmp_path / "dupes.csv",
        [_candidate("EFE-0001", "Summit Lodge Verbier", "https://summitlodge.example")],
    )
    rc = main(["--config", str(synthetic_config.config_path), "promote", str(source)])
    text = capsys.readouterr().out
    assert rc == 0, text
    assert "Nothing to promote" in text
    assert not list(synthetic_config.output_directory.glob("*.xlsx"))
    assert list(synthetic_config.artifacts_directory.glob("NOWRITE_promotion_*.md"))
