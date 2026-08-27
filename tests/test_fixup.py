"""Corrective releases: renumbering IDs and normalising cells, through the gate."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from efe.cli import main
from efe.models import VerificationError
from efe.workbook.fixup import (
    FixupPlan,
    FixupRecord,
    assert_fixups_legal,
    assert_ids_unique_after,
    fixable_columns,
    propose_fixups,
    read_fixup_plan,
    write_fixup_plan,
    write_fixups,
)
from efe.workbook.reader import load_workbook_view
from efe.workbook.state import load_state, state_path
from efe.workbook.verify import compare, snapshot


def _duplicate_id(workbook: Path, *, row: int = 9, value: str = "EFE-0003") -> None:
    """Row `row` becomes a second holder of `value`, in a block of its own."""
    wb = load_workbook(workbook)
    ws = wb["PARTNERS"]
    ws.cell(row, 1).value = value
    ws.cell(row, 39).value = "R9-new"
    wb.save(workbook)
    wb.close()


def _plan_of(view, cfg, *records: FixupRecord) -> FixupPlan:
    plan = FixupPlan(
        source=Path("test-plan.csv"),
        input_name=view.path.name,
        input_sha256=view.fingerprint,
    )
    for record in records:
        (plan.renumbers if record.kind == "renumber" else plan.cells).append(record)
    return plan


def _renumber(row: int, old: str, new: str, name: str) -> FixupRecord:
    return FixupRecord(
        kind="renumber",
        row=row,
        column="ID",
        old=old,
        new=new,
        guard_column="Entity_Name",
        guard_value=name,
        why="test",
    )


# ---------------------------------------------------------------------------
# Guards
# ---------------------------------------------------------------------------


def test_fixable_columns_exclude_everything_human_owned(synthetic_config):
    allowed = fixable_columns(synthetic_config)
    assert "ID" in allowed and "Phone" in allowed and "Source_URL" in allowed
    for forbidden in ("Contacted", "Status", "Next_Follow_Up", "Category", "Material_Sent"):
        assert forbidden not in allowed


def test_plan_refuses_a_stale_old_value(synthetic_config, synthetic_workbook):
    view = load_workbook_view(synthetic_config)
    plan = _plan_of(
        view,
        synthetic_config,
        FixupRecord(
            kind="cell",
            row=2,
            column="Phone",
            old="not what the cell holds",
            new="+41270000000",
            guard_column="ID",
            guard_value="EFE-0001",
        ),
    )
    with pytest.raises(VerificationError, match="plan expected 'not what the cell holds'"):
        write_fixups(synthetic_config, view, plan, run_id="test-stale")
    assert not list(synthetic_config.output_directory.glob("*.xlsx"))


def test_plan_meta_row_pins_the_input_fingerprint(synthetic_config, synthetic_workbook):
    view = load_workbook_view(synthetic_config)
    plan = _plan_of(view, synthetic_config, _renumber(9, "EFE-0008", "EFE-0009", "No Website Row"))
    plan.input_sha256 = "0" * 64
    with pytest.raises(VerificationError, match="built against 000000000000"):
        assert_fixups_legal(plan, synthetic_config, view)


def test_a_crm_or_precedent_column_in_the_plan_is_refused(synthetic_config):
    view = load_workbook_view(synthetic_config)
    for column, value in (("Contacted", "YES"), ("Status", "Contacted"), ("Category", "9. X")):
        plan = _plan_of(
            view,
            synthetic_config,
            FixupRecord(kind="cell", row=2, column=column, old="NO", new=value),
        )
        with pytest.raises(VerificationError, match="human-owned CRM, formula or precedent"):
            assert_fixups_legal(plan, synthetic_config, view)


def test_renumber_refuses_a_new_id_already_in_use(synthetic_config):
    view = load_workbook_view(synthetic_config)
    plan = _plan_of(view, synthetic_config, _renumber(9, "EFE-0008", "EFE-0002", "No Website Row"))
    with pytest.raises(VerificationError, match=r"EFE-0002 would still be on rows \[3, 9\]"):
        assert_ids_unique_after(plan, synthetic_config, view, allow_backfill=True)


def test_renumber_refuses_a_reserved_id(synthetic_config):
    view = load_workbook_view(synthetic_config)
    plan = _plan_of(view, synthetic_config, _renumber(9, "EFE-0008", "EFE-0009", "No Website Row"))
    with pytest.raises(VerificationError, match="EFE-0009 is reserved"):
        assert_ids_unique_after(plan, synthetic_config, view, reserved_ids={"EFE-0009"})
    assert_ids_unique_after(plan, synthetic_config, view, reserved_ids={"EFE-0010"})


def test_renumber_post_image_must_be_unique(synthetic_config, synthetic_workbook):
    """A -> B while an unmapped row still holds B is the case a naive check misses."""
    _duplicate_id(synthetic_workbook)  # row 9 becomes a second EFE-0003
    view = load_workbook_view(synthetic_config)
    both = _plan_of(
        view,
        synthetic_config,
        _renumber(9, "EFE-0003", "EFE-0008", "No Website Row"),
    )
    assert_ids_unique_after(both, synthetic_config, view, allow_backfill=True)

    collide = _plan_of(
        view,
        synthetic_config,
        _renumber(9, "EFE-0003", "EFE-0005", "No Website Row"),
    )
    with pytest.raises(VerificationError, match=r"EFE-0005 would still be on rows \[6, 9\]"):
        assert_ids_unique_after(collide, synthetic_config, view, allow_backfill=True)


def test_renumber_below_the_highest_id_needs_allow_backfill(synthetic_config, synthetic_workbook):
    """A freed number is only reused when a human says so: gaps usually mean history."""
    wb = load_workbook(synthetic_workbook)
    wb["PARTNERS"]["A9"] = "EFE-0020"  # leaves EFE-0009..0019 free
    wb.save(synthetic_workbook)
    wb.close()
    view = load_workbook_view(synthetic_config)
    plan = _plan_of(view, synthetic_config, _renumber(9, "EFE-0020", "EFE-0009", "No Website Row"))
    with pytest.raises(VerificationError, match="not above the sheet's highest ID EFE-0020"):
        assert_ids_unique_after(plan, synthetic_config, view)
    assert_ids_unique_after(plan, synthetic_config, view, allow_backfill=True)


def test_max_cells_stops_a_runaway_plan(synthetic_config, monkeypatch):
    import efe.workbook.fixup as fixup_mod

    monkeypatch.setattr(fixup_mod, "MAX_CELLS", 1)
    view = load_workbook_view(synthetic_config)
    plan = _plan_of(
        view,
        synthetic_config,
        FixupRecord(kind="cell", row=2, column="Phone", old="TBD", new="+41270000000"),
        FixupRecord(kind="cell", row=3, column="Phone", old="TBD", new="+41270000001"),
    )
    with pytest.raises(VerificationError, match="exceed the 1 ceiling"):
        assert_fixups_legal(plan, synthetic_config, view)


# ---------------------------------------------------------------------------
# The write
# ---------------------------------------------------------------------------


def test_renumber_writes_only_the_id_cells_and_passes_the_gate(
    synthetic_config, synthetic_workbook
):
    """The core claim: one cell per renumbered row, and the file is otherwise identical."""
    _duplicate_id(synthetic_workbook)
    view = load_workbook_view(synthetic_config)
    assert view.schema is not None
    assert view.schema.duplicate_ids == {"EFE-0003": [4, 9]}

    plan = _plan_of(view, synthetic_config, _renumber(9, "EFE-0003", "EFE-0008", "No Website Row"))
    out = write_fixups(synthetic_config, view, plan, run_id="test-renumber")
    assert out.name.endswith("_v02.xlsx")

    wb = load_workbook(out)
    try:
        ws = wb["PARTNERS"]
        assert ws["A9"].value == "EFE-0008"
        assert [ws.cell(r, 1).value for r in range(2, 9)] == [f"EFE-{n:04d}" for n in range(1, 8)]
        assert ws["B9"].value == "No Website Row"  # nothing else on the row moved
        log = wb["CHANGELOG"]
        assert log["A5"].value == "v02" and log["D5"].value == "efe fixup"
        assert "renumbered 1 IDs" in log["C5"].value
        detail = wb["CHANGELOG_DETAIL"]
        assert detail["H3"].value == "EFE-0003" and detail["I3"].value == "EFE-0008"
        assert detail["N3"].value == "fixup.renumber"
    finally:
        wb.close()

    before, after = snapshot(view.path), snapshot(out)
    logs = {
        k
        for k in after.values
        if k[0] in ("CHANGELOG", "CHANGELOG_DETAIL") and k not in before.values
    }
    assert (
        compare(
            before,
            after,
            allowed_value_changes={("PARTNERS", "A9")} | logs,
            allowed_autofilter_changes={"CHANGELOG_DETAIL"},
        )
        == []
    )
    assert after.cached_values == before.cached_values
    assert after.formulas == before.formulas


def test_sentinel_lookalike_becomes_the_sentinel(synthetic_config, synthetic_workbook):
    wb = load_workbook(synthetic_workbook)
    wb["PARTNERS"]["J2"] = "TBD (sin email trade dedicado)"
    wb.save(synthetic_workbook)
    wb.close()
    view = load_workbook_view(synthetic_config)
    assert not synthetic_config.workbook.is_empty("TBD (sin email trade dedicado)")

    plan = propose_fixups(view, synthetic_config)
    proposed = [r for r in plan.cells if r.column == "Sales_B2B_Email"]
    assert len(proposed) == 1 and proposed[0].new == "TBD"
    # The point of the fix: the enricher can see the cell again.
    assert synthetic_config.workbook.is_empty(proposed[0].new)


def test_two_addresses_in_one_cell_are_listed_not_applied(synthetic_config, synthetic_workbook):
    wb = load_workbook(synthetic_workbook)
    ws = wb["PARTNERS"]
    ws["J2"] = "office@one.example (attn. Director)"
    ws["J3"] = "a.person@two.example · events@two.example"
    ws["J4"] = "Kai Schweigkofler - Travel Agency desk"
    wb.save(synthetic_workbook)
    wb.close()
    view = load_workbook_view(synthetic_config)

    plan = propose_fixups(view, synthetic_config)
    emails = {r.row: r.new for r in plan.cells if r.column == "Sales_B2B_Email"}
    assert emails == {2: "office@one.example"}  # only the unambiguous one
    assert any("holds 2 addresses" in item for item in plan.review)
    assert any("holds no address at all" in item for item in plan.review)


def test_phone_normalisation_uses_the_row_region_and_skips_the_ambiguous(
    synthetic_config, synthetic_workbook
):
    wb = load_workbook(synthetic_workbook)
    ws = wb["PARTNERS"]
    ws["K2"] = "+41 27 966 03 03"  # spaces only
    ws["K3"] = "089 1234567"  # national, Country=DE on row 3
    ws["K4"] = "(55) 6279 2102 / 2103"  # two numbers
    ws["K5"] = "+39 393896911110"  # not a valid number
    wb.save(synthetic_workbook)
    wb.close()
    view = load_workbook_view(synthetic_config)

    plan = propose_fixups(view, synthetic_config)
    phones = {r.row: r.new for r in plan.cells if r.column == "Phone"}
    assert phones[2] == "+41279660303"
    assert phones[3] == "+49891234567"  # region came from the row's Country
    assert 4 not in phones and 5 not in phones
    assert any("more than one number" in item for item in plan.review)
    assert any("not a valid dialable number" in item for item in plan.review)

    out = write_fixups(synthetic_config, view, plan, run_id="test-phones")
    wb = load_workbook(out)
    try:
        assert wb["PARTNERS"]["K2"].value == "+41279660303"
        assert isinstance(wb["PARTNERS"]["K2"].value, str)
    finally:
        wb.close()


def test_nothing_to_fix_is_refused(synthetic_config):
    view = load_workbook_view(synthetic_config)
    plan = _plan_of(view, synthetic_config)
    with pytest.raises(VerificationError, match="nothing to fix"):
        write_fixups(synthetic_config, view, plan, run_id="test-empty")


# ---------------------------------------------------------------------------
# The plan file and the CLI
# ---------------------------------------------------------------------------


def test_plan_round_trips_through_the_csv(synthetic_config, tmp_path):
    view = load_workbook_view(synthetic_config)
    plan = _plan_of(view, synthetic_config, _renumber(9, "EFE-0008", "EFE-0009", "No Website Row"))
    path = write_fixup_plan(tmp_path / "plan.csv", plan)
    back = read_fixup_plan(path, synthetic_config)
    assert back.input_sha256 == view.fingerprint and back.input_name == view.path.name
    assert len(back.renumbers) == 1 and back.renumbers[0].new == "EFE-0009"
    assert back.id_map == {"EFE-0008": "EFE-0009"}


def test_the_plan_carries_what_it_refused_to_automate(
    synthetic_config, synthetic_workbook, tmp_path
):
    """The applied report must be as honest as the proposal about what it left alone."""
    wb = load_workbook(synthetic_workbook)
    wb["PARTNERS"]["J3"] = "a.person@two.example / events@two.example"
    wb.save(synthetic_workbook)
    wb.close()
    view = load_workbook_view(synthetic_config)
    plan = propose_fixups(view, synthetic_config)
    assert plan.review

    path = write_fixup_plan(tmp_path / "plan.csv", plan)
    assert read_fixup_plan(path, synthetic_config).review == plan.review


def test_plan_csv_shape_is_strict(synthetic_config, tmp_path):
    bad = tmp_path / "bad.csv"
    bad.write_text("Kind,Row\ncell,2\n", encoding="utf-8")
    with pytest.raises(VerificationError, match="header is"):
        read_fixup_plan(bad, synthetic_config)

    unknown = tmp_path / "unknown.csv"
    unknown.write_text(
        "Kind,Row,Column,Old,New,Guard_Column,Guard_Value,Why\ncell,2,Bogus,a,b,,,\n",
        encoding="utf-8",
    )
    with pytest.raises(VerificationError, match="not a PARTNERS column"):
        read_fixup_plan(unknown, synthetic_config)

    misdirected = tmp_path / "mis.csv"
    misdirected.write_text(
        "Kind,Row,Column,Old,New,Guard_Column,Guard_Value,Why\nrenumber,2,Phone,a,b,,,\n",
        encoding="utf-8",
    )
    with pytest.raises(VerificationError, match="a renumber must target the 'ID' column"):
        read_fixup_plan(misdirected, synthetic_config)


def test_cli_fixup_propose_then_dry_run_then_write(
    synthetic_config, synthetic_workbook, capsys, monkeypatch
):
    monkeypatch.setenv("COLUMNS", "200")
    _duplicate_id(synthetic_workbook)

    rc = main(["--config", str(synthetic_config.config_path), "fixup", "--propose"])
    text = capsys.readouterr().out
    assert rc == 0, text
    assert "renumber :" in text and "NOT fixed by this tool" in text
    plan_path = next(synthetic_config.artifacts_directory.glob("fixup_plan_*.csv"))
    assert list(synthetic_config.artifacts_directory.glob("PROPOSED_fixup_*.md"))
    assert not list(synthetic_config.output_directory.glob("*.xlsx"))

    rc = main(["--config", str(synthetic_config.config_path), "fixup", str(plan_path), "--dry-run"])
    text = capsys.readouterr().out
    assert rc == 0, text
    assert "DRY RUN" in text
    assert not list(synthetic_config.output_directory.glob("*.xlsx"))
    dry = next(synthetic_config.artifacts_directory.glob("DRYRUN_fixup_*.md")).read_text(
        encoding="utf-8"
    )
    assert "DRY RUN - nothing written" in dry
    assert "## Not fixed by this tool" in dry

    rc = main(["--config", str(synthetic_config.config_path), "fixup", str(plan_path)])
    text = capsys.readouterr().out
    assert rc == 0, text
    written = list(synthetic_config.output_directory.glob("*_v02.xlsx"))
    assert len(written) == 1
    report = next(
        p
        for p in synthetic_config.artifacts_directory.glob("fixup_*.md")
        if not p.name.startswith(("DRYRUN", "PROPOSED", "REFUSED"))
    )
    assert f"WRITTEN: {written[0].name}" in report.read_text(encoding="utf-8")

    saved = load_state(state_path(synthetic_config.state_directory))
    assert saved is not None and saved.version == 2 and saved.data_rows == 8
    assert saved.command == "efe fixup (output)"
    assert saved.file_sha256

    # The point of the release: the duplicate is gone.
    synthetic_config.output_dir = None
    Path(written[0]).replace(synthetic_config.workbook_directory / written[0].name)
    nxt = load_workbook_view(synthetic_config)
    assert nxt.schema is not None and nxt.schema.duplicate_ids == {}


def test_cli_fixup_needs_a_plan_or_propose(synthetic_config):
    with pytest.raises(SystemExit, match="needs a plan CSV"):
        main(["--config", str(synthetic_config.config_path), "fixup"])


def test_verify_expect_plan_proves_the_narrow_claim(
    synthetic_config, synthetic_workbook, capsys, monkeypatch
):
    monkeypatch.setenv("COLUMNS", "200")
    _duplicate_id(synthetic_workbook)
    view = load_workbook_view(synthetic_config)
    plan = _plan_of(view, synthetic_config, _renumber(9, "EFE-0003", "EFE-0008", "No Website Row"))
    plan_path = write_fixup_plan(synthetic_config.artifacts_directory / "p.csv", plan)
    plan.source = plan_path
    out = write_fixups(synthetic_config, view, plan, run_id="test-verify")

    rc = main(
        [
            "--config",
            str(synthetic_config.config_path),
            "verify",
            str(out),
            "--against",
            str(view.path),
            "--expect-plan",
            str(plan_path),
        ]
    )
    text = capsys.readouterr().out
    assert rc == 0, text
    assert "fidelity gate: PASS" in text

    # A change the plan does not name is caught, unlike --allow-partners-changes.
    wb = load_workbook(out)
    wb["PARTNERS"]["B2"] = "renamed behind the plan's back"
    wb.save(out)
    wb.close()
    rc = main(
        [
            "--config",
            str(synthetic_config.config_path),
            "verify",
            str(out),
            "--against",
            str(view.path),
            "--expect-plan",
            str(plan_path),
        ]
    )
    assert rc == 1
