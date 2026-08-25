"""The workbook path: guards, selection, the write, and the fidelity gate.

Runs entirely against the synthetic workbook from `conftest.py`. The real file on
Google Drive is never opened by the test suite.
"""

from __future__ import annotations

import shutil
import zipfile
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from efe.models import (
    CellChange,
    Confidence,
    DataClass,
    DriveSyncError,
    SchemaMismatchError,
    VerificationError,
    WorkbookLockedError,
)
from efe.workbook.reader import (
    WorkbookGuardOutputExists,
    guard_readable,
    guard_writable,
    load_workbook_view,
    select_candidates,
)
from efe.workbook.verify import compare, snapshot
from efe.workbook.writer import (
    assert_no_precedents_touched,
    build_provenance_changes,
    next_version_path,
    write_enriched,
)
from efe.workbook.xmlutil import read_cached_values, reinject_cached_values
from tests.conftest import inject_cached_values

#: Later than the fixture's own Date_Verified (2026-08-21), so provenance updates
#: are genuine changes rather than no-ops the writer correctly skips.
NOW = datetime(2026, 8, 23, 12, 0, 0)


def change(row, column, field, new_value, *, old_value="TBD", url="https://x.example/contact"):
    return CellChange(
        row=row, column=column, field=field, entity_id=f"EFE-{row - 1:04d}",
        entity_name="Test Entity", old_value=old_value, new_value=new_value,
        confidence=Confidence.HIGH, data_class=DataClass.CORPORATE_ROLE,
        source_url=url, fetched_at=NOW, extractor="tests",
    )


# ---------------------------------------------------------------------------
# Guards
# ---------------------------------------------------------------------------

def test_missing_file_reports_a_drive_sync_problem(tmp_path):
    with pytest.raises(DriveSyncError, match="not found"):
        guard_readable(tmp_path / "nope.xlsx")


def test_zero_byte_file_reports_a_drive_sync_problem(tmp_path):
    empty = tmp_path / "empty.xlsx"
    empty.write_bytes(b"")
    with pytest.raises(DriveSyncError, match="ZERO BYTES"):
        guard_readable(empty)


def test_truncated_placeholder_is_refused(tmp_path):
    stub = tmp_path / "stub.xlsx"
    stub.write_bytes(b"x" * 100)
    with pytest.raises(DriveSyncError, match="below the"):
        guard_readable(stub)


def test_excel_lock_file_stops_the_run(synthetic_workbook):
    lock = synthetic_workbook.with_name("~$" + synthetic_workbook.name)
    lock.write_bytes(b"lock")
    try:
        with pytest.raises(WorkbookLockedError, match="OPEN IN EXCEL"):
            guard_readable(synthetic_workbook, min_bytes=4000)
    finally:
        lock.unlink()


def test_non_zip_payload_is_refused(tmp_path):
    fake = tmp_path / "fake.xlsx"
    fake.write_bytes(b"not a zip file " * 5000)
    with pytest.raises(DriveSyncError, match="not a readable"):
        guard_readable(fake)


def test_existing_output_is_never_overwritten(synthetic_workbook):
    with pytest.raises(WorkbookGuardOutputExists, match="Refusing to overwrite"):
        guard_writable(synthetic_workbook)


def test_schema_mismatch_is_caught(synthetic_config, synthetic_workbook):
    wb = load_workbook(synthetic_workbook)
    wb["PARTNERS"]["I1"] = "Renamed_Column"
    wb.save(synthetic_workbook)
    wb.close()
    with pytest.raises(SchemaMismatchError, match="General_Email"):
        load_workbook_view(synthetic_config)


# ---------------------------------------------------------------------------
# Selection
# ---------------------------------------------------------------------------

def test_selection_skips_gold_rows_and_rows_without_a_website(synthetic_config):
    view = load_workbook_view(synthetic_config)
    candidates, skipped = select_candidates(view, synthetic_config)
    names = {c.name for c in candidates}

    assert "Already Contacted Agency" not in names, "Contacted=YES row must never be fetched"
    assert "No Website Row" not in names
    assert "Summit Lodge Verbier" in names
    # A row with one filled field but others TBD is still worth enriching.
    assert "Filled Row" in names

    reasons = " ".join(skipped.values())
    assert "human-verified row" in reasons
    assert "no usable Website_URL" in reasons


def test_candidate_carries_current_cell_values(synthetic_config):
    view = load_workbook_view(synthetic_config)
    candidates, _ = select_candidates(view, synthetic_config)
    filled = next(c for c in candidates if c.name == "Filled Row")
    assert filled.existing["general_email"] == "info@filled.example"
    assert filled.existing["phone"] == "TBD"


def test_shared_domain_is_detected_from_the_sheet(synthetic_config):
    view = load_workbook_view(synthetic_config)
    assert view.domain_row_counts["grandclass.example"] == 2
    assert "grandclass.example" in view.duplicate_domains


def test_source_ledger_is_read(synthetic_config):
    view = load_workbook_view(synthetic_config)
    assert view.ledger_domains["summitlodge.example"]["exclude_next_round"] == "YES"


# ---------------------------------------------------------------------------
# Cached-value reinjection
# ---------------------------------------------------------------------------

def test_fixture_really_has_cached_values(synthetic_workbook):
    cached = read_cached_values(synthetic_workbook)
    assert cached[("PARTNERS", "AC2")] == "14"
    assert cached[("DASHBOARD", "B2")] == "4"


def test_openpyxl_roundtrip_loses_cached_values_and_reinjection_restores_them(
    synthetic_workbook, tmp_path
):
    copy = tmp_path / "roundtrip.xlsx"
    shutil.copy2(synthetic_workbook, copy)
    wb = load_workbook(copy)
    wb.save(copy)
    wb.close()

    after = read_cached_values(copy)
    assert after[("PARTNERS", "AC2")] is None, "openpyxl should have dropped it"

    repaired = reinject_cached_values(copy, synthetic_workbook)
    assert repaired > 0
    restored = read_cached_values(copy)
    assert restored[("PARTNERS", "AC2")] == "14"
    assert restored[("DASHBOARD", "B2")] == "4"
    assert restored == read_cached_values(synthetic_workbook)


# ---------------------------------------------------------------------------
# The write
# ---------------------------------------------------------------------------

def test_write_fills_only_the_intended_cells(synthetic_config):
    view = load_workbook_view(synthetic_config)
    changes = [
        change(2, "I", "general_email", "info@summitlodge.example"),
        change(2, "J", "sales_b2b_email", "traveltrade@summitlodge.example"),
        change(2, "K", "phone", "+41274720000"),
    ]
    out = write_enriched(synthetic_config, view, changes, run_id="test-001")

    wb = load_workbook(out)
    partners = wb["PARTNERS"]
    assert partners["I2"].value == "info@summitlodge.example"
    assert partners["J2"].value == "traveltrade@summitlodge.example"
    assert partners["K2"].value == "+41274720000"
    # Untouched row keeps its TBDs.
    assert partners["I3"].value == "TBD"
    wb.close()


def test_write_preserves_formulas_validations_and_cached_values(synthetic_config):
    view = load_workbook_view(synthetic_config)
    before = snapshot(view.path)
    out = write_enriched(
        synthetic_config, view,
        [change(2, "I", "general_email", "info@summitlodge.example")],
        run_id="test-002",
    )
    after = snapshot(out)

    assert len(after.formulas) == len(before.formulas)
    assert after.formulas == before.formulas
    assert after.data_validations["PARTNERS"] == before.data_validations["PARTNERS"]
    assert after.autofilter["PARTNERS"] == before.autofilter["PARTNERS"]
    assert after.freeze_panes["PARTNERS"] == before.freeze_panes["PARTNERS"]
    for key, value in before.cached_values.items():
        if value is not None:
            assert after.cached_values[key] == value, f"{key} lost its cached result"


def test_crm_columns_are_untouched_by_a_write(synthetic_config):
    view = load_workbook_view(synthetic_config)
    before = snapshot(view.path)
    out = write_enriched(
        synthetic_config, view,
        [change(2, "I", "general_email", "info@summitlodge.example")],
        run_id="test-003",
    )
    after = snapshot(out)
    crm = set(synthetic_config.workbook.crm_columns)
    for (sheet, coordinate), value in before.values.items():
        if sheet != "PARTNERS":
            continue
        letter = "".join(ch for ch in coordinate if ch.isalpha())
        if letter in crm:
            assert after.values.get((sheet, coordinate)) == value, coordinate


def test_provenance_is_written_only_on_changed_rows(synthetic_config):
    view = load_workbook_view(synthetic_config)
    changes = [change(2, "I", "general_email", "info@summitlodge.example")]
    provenance = build_provenance_changes(changes, synthetic_config, view, "2026-08-22")
    assert {p.column for p in provenance} == {"AK", "AL", "AM"}
    assert all(p.row == 2 for p in provenance)

    out = write_enriched(synthetic_config, view, changes, run_id="test-004")
    wb = load_workbook(out)
    partners = wb["PARTNERS"]
    assert partners["AK2"].value == "https://x.example/contact"
    assert partners["AM2"].value == synthetic_config.selection.round_tag
    # Row 3 changed nothing, so its provenance is byte-identical.
    assert partners["AK3"].value == "TBD"
    assert partners["AM3"].value == "R1"
    wb.close()


def test_existing_source_url_is_appended_not_replaced(synthetic_config, synthetic_workbook):
    wb = load_workbook(synthetic_workbook)
    wb["PARTNERS"]["AK2"] = "https://earlier.example/research"
    wb.save(synthetic_workbook)
    wb.close()
    # openpyxl just dropped the cached results; restore them so this test measures
    # the append behaviour rather than re-testing the gate.
    inject_cached_values(
        synthetic_workbook,
        {("PARTNERS", f"AC{r}"): "14" for r in range(2, 10)}
        | {("DASHBOARD", "B2"): "4", ("DASHBOARD", "B3"): "1", ("DASHBOARD", "B5"): "1"},
    )

    view = load_workbook_view(synthetic_config)
    out = write_enriched(
        synthetic_config, view,
        [change(2, "I", "general_email", "info@summitlodge.example")],
        run_id="test-005",
    )
    wb = load_workbook(out)
    assert wb["PARTNERS"]["AK2"].value == (
        "https://earlier.example/research; https://x.example/contact"
    )
    wb.close()


def test_changelog_row_and_detail_sheet_are_added(synthetic_config):
    view = load_workbook_view(synthetic_config)
    held = [
        change(2, "I", "general_email", "reservations@summitlodge.example",
               url="https://x.example/contact")
    ]
    out = write_enriched(
        synthetic_config, view,
        [change(2, "I", "general_email", "info@summitlodge.example")],
        run_id="test-006", held_back=held,
    )
    wb = load_workbook(out)
    assert synthetic_config.workbook.changelog_detail_sheet in wb.sheetnames

    detail = wb[synthetic_config.workbook.changelog_detail_sheet]
    headers = [detail.cell(1, c).value for c in range(1, 16)]
    assert headers[:7] == ["Timestamp", "Run_ID", "Row", "Entity_ID", "Entity_Name",
                           "Column", "Field"]
    recorded = {detail.cell(r, 9).value for r in range(2, detail.max_row + 1)}
    assert "info@summitlodge.example" in recorded
    assert "reservations@summitlodge.example" in recorded, "held-back values are logged too"

    changelog = wb["CHANGELOG"]
    last = [changelog.cell(changelog.max_row, c).value for c in range(1, 5)]
    assert last[0] == out.stem.rsplit("_", 1)[-1]      # the emitted version label
    assert "test-006" in last[2]
    assert "CHANGELOG_DETAIL" in last[2]
    wb.close()


# ---------------------------------------------------------------------------
# Refusals
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    ("column", "match"),
    [
        ("Z", "human-owned CRM"),
        ("AI", "human-owned CRM"),
        ("AC", "human-owned CRM"),   # the Next_Follow_Up formula
        ("B", "not in writable_columns"),
    ],
)
def test_writer_refuses_forbidden_columns(synthetic_config, column, match):
    view = load_workbook_view(synthetic_config)
    with pytest.raises(VerificationError, match=match):
        write_enriched(
            synthetic_config, view, [change(2, column, "nope", "x")], run_id="test-007"
        )


def test_writer_refuses_to_overwrite_a_filled_cell(synthetic_config):
    view = load_workbook_view(synthetic_config)
    # Row 8 is `Filled Row`, whose General_Email already holds a real address.
    with pytest.raises(VerificationError, match="refusing to overwrite"):
        write_enriched(
            synthetic_config, view,
            [change(8, "I", "general_email", "other@filled.example",
                    old_value="info@filled.example")],
            run_id="test-008",
        )


def test_writer_refuses_a_value_without_provenance(synthetic_config):
    view = load_workbook_view(synthetic_config)
    with pytest.raises(VerificationError, match="without provenance"):
        write_enriched(
            synthetic_config, view,
            [change(2, "I", "general_email", "info@x.example", url="")],
            run_id="test-009",
        )


def test_precedent_guard_blocks_columns_that_feed_formulas(synthetic_config):
    """Cached-value reinjection is only sound while no written column feeds a formula.

    The column map already makes this unreachable through `write_enriched`; the guard
    exists so that loosening `writable_columns` in config.yaml fails loudly instead of
    silently preserving a stale DASHBOARD total.
    """
    for column in ("C", "G", "Y", "Z", "AI", "AA", "AB"):
        with pytest.raises(VerificationError, match="feed live formulas"):
            assert_no_precedents_touched([change(2, column, "x", "y")], synthetic_config)

    # A column the enricher actually writes is not a precedent.
    assert_no_precedents_touched([change(2, "I", "general_email", "a@b.example")],
                                 synthetic_config)


def test_nothing_is_written_when_the_change_set_is_illegal(synthetic_config):
    view = load_workbook_view(synthetic_config)
    before = synthetic_config.output_directory.glob("*.xlsx")
    assert list(before) == []
    with pytest.raises(VerificationError):
        write_enriched(
            synthetic_config, view, [change(2, "Z", "contacted", "YES")], run_id="test-011"
        )
    assert list(synthetic_config.output_directory.glob("*.xlsx")) == []


def test_input_is_never_modified(synthetic_config):
    view = load_workbook_view(synthetic_config)
    original = view.path.read_bytes()
    write_enriched(
        synthetic_config, view,
        [change(2, "I", "general_email", "info@summitlodge.example")],
        run_id="test-012",
    )
    assert view.path.read_bytes() == original


# ---------------------------------------------------------------------------
# The fidelity gate itself
# ---------------------------------------------------------------------------

def test_gate_detects_a_corrupted_formula(synthetic_config, tmp_path):
    view = load_workbook_view(synthetic_config)
    broken = tmp_path / "broken.xlsx"
    shutil.copy2(view.path, broken)
    wb = load_workbook(broken)
    wb["PARTNERS"]["AC2"] = "=BROKEN()"
    wb.save(broken)
    wb.close()

    problems = compare(snapshot(view.path), snapshot(broken))
    assert any("formula" in p for p in problems)


def test_gate_detects_a_lost_data_validation(synthetic_config, tmp_path):
    view = load_workbook_view(synthetic_config)
    stripped = tmp_path / "stripped.xlsx"
    shutil.copy2(view.path, stripped)
    wb = load_workbook(stripped)
    wb["PARTNERS"].data_validations.dataValidation = []
    wb.save(stripped)
    wb.close()

    problems = compare(snapshot(view.path), snapshot(stripped))
    assert any("data validation" in p for p in problems)


def test_gate_detects_missing_cached_values(synthetic_config, tmp_path):
    view = load_workbook_view(synthetic_config)
    stale = tmp_path / "stale.xlsx"
    shutil.copy2(view.path, stale)
    wb = load_workbook(stale)
    wb.save(stale)          # openpyxl drops the cached results
    wb.close()

    problems = compare(snapshot(view.path), snapshot(stale))
    assert any("cached result" in p for p in problems)


def test_gate_detects_an_unintended_value_change(synthetic_config, tmp_path):
    view = load_workbook_view(synthetic_config)
    tampered = tmp_path / "tampered.xlsx"
    shutil.copy2(view.path, tampered)
    wb = load_workbook(tampered)
    wb["PARTNERS"]["B3"] = "Renamed Entity"
    wb.save(tampered)
    wb.close()

    problems = compare(
        snapshot(view.path), snapshot(tampered), require_cached_values=False
    )
    assert any("without being intended" in p for p in problems)


def test_gate_passes_a_clean_write(synthetic_config):
    view = load_workbook_view(synthetic_config)
    out = write_enriched(
        synthetic_config, view,
        [change(2, "I", "general_email", "info@summitlodge.example")],
        run_id="test-013",
    )
    allowed = {("PARTNERS", c) for c in ("I2", "AK2", "AL2", "AM2")}
    allowed |= {key for key in snapshot(out).values if key[0] == "CHANGELOG"}
    assert compare(
        snapshot(view.path),
        snapshot(out),
        allowed_value_changes=allowed,
        allowed_new_sheets={synthetic_config.workbook.changelog_detail_sheet},
    ) == []


# ---------------------------------------------------------------------------
# Versioned output naming
# ---------------------------------------------------------------------------

def test_next_version_increments(synthetic_config):
    """The output version is the input's plus one -- never a folder scan."""
    from efe.models import VersionConflictError

    directory = synthetic_config.output_directory
    base = synthetic_config.output_basename
    assert next_version_path(synthetic_config, 3, when="2026-08-22").name == (
        f"2026-08-22_{base}_v04.xlsx"
    )
    # A higher version anywhere in the workbook or output folder blocks the write:
    # version numbers only go up, and the resolver would read that file next time.
    (directory / f"2026-08-20_{base}_v07.xlsx").write_bytes(b"x")
    with pytest.raises(VersionConflictError, match="v07 already exists"):
        next_version_path(synthetic_config, 3, when="2026-08-22")
    assert next_version_path(synthetic_config, 7, when="2026-08-22").name.endswith("_v08.xlsx")


def test_output_version_needs_a_versioned_input(synthetic_config):
    """No vNN on the input means no way to derive the output: refuse, never guess v01."""
    from efe.models import VersionConflictError

    with pytest.raises(VersionConflictError, match="carries no vNN"):
        next_version_path(synthetic_config, 0, when="2026-08-22")
    assert next_version_path(synthetic_config, 1, when="2026-08-22").name.endswith("_v02.xlsx")


def test_written_output_is_a_valid_xlsx(synthetic_config):
    view = load_workbook_view(synthetic_config)
    out = write_enriched(
        synthetic_config, view,
        [change(2, "P", "instagram_handle", "@summitlodgeverbier")],
        run_id="test-014",
    )
    with zipfile.ZipFile(out) as zf:
        assert zf.testzip() is None
        assert "xl/workbook.xml" in zf.namelist()
    assert Path(out).stat().st_size > 5_000


def test_no_staging_file_is_left_behind(synthetic_config):
    view = load_workbook_view(synthetic_config)
    write_enriched(
        synthetic_config, view,
        [change(2, "P", "instagram_handle", "@summitlodgeverbier")],
        run_id="test-015",
    )
    staging = synthetic_config.state_directory / "staging"
    assert not list(staging.glob("*.xlsx"))


# ---------------------------------------------------------------------------
# Dropped parts are judged by content, not by filename
# ---------------------------------------------------------------------------

_EMPTY_DRAWING = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/'
    'spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/'
    'main"/>'
)
_DRAWING_WITH_AN_IMAGE = _EMPTY_DRAWING.replace(
    "/>", "><xdr:twoCellAnchor><xdr:pic/></xdr:twoCellAnchor></xdr:wsDr>"
)


def _add_parts(source, target, parts: dict[str, str]) -> None:
    with zipfile.ZipFile(source) as zin:
        contents = {n: zin.read(n) for n in zin.namelist()}
    contents.update({k: v.encode() for k, v in parts.items()})
    with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, blob in contents.items():
            zout.writestr(name, blob)


def test_an_empty_drawing_may_be_dropped(synthetic_workbook, tmp_path):
    with_stub = tmp_path / "with_stub.xlsx"
    _add_parts(synthetic_workbook, with_stub,
               {"xl/drawings/drawing1.xml": _EMPTY_DRAWING})

    before = snapshot(with_stub)
    assert before.drawing_anchor_counts["xl/drawings/drawing1.xml"] == 0

    without = tmp_path / "without.xlsx"
    shutil.copy2(synthetic_workbook, without)
    problems = compare(before, snapshot(without), require_cached_values=False)
    assert not any("drawing" in p for p in problems)


def test_a_drawing_holding_an_object_may_not_be_dropped(synthetic_workbook, tmp_path):
    """The whitelist used to match on filename, which would have hidden this."""
    with_image = tmp_path / "with_image.xlsx"
    _add_parts(synthetic_workbook, with_image,
               {"xl/drawings/drawing1.xml": _DRAWING_WITH_AN_IMAGE})

    before = snapshot(with_image)
    assert before.drawing_anchor_counts["xl/drawings/drawing1.xml"] == 1

    without = tmp_path / "without.xlsx"
    shutil.copy2(synthetic_workbook, without)
    problems = compare(before, snapshot(without), require_cached_values=False)
    assert any("anchored object" in p for p in problems)


def test_google_sheets_metadata_blob_is_benign(synthetic_workbook, tmp_path):
    """A Google Sheets export carries an inert `xl/metadata` part; Excel ignores it."""
    with_meta = tmp_path / "with_meta.xlsx"
    _add_parts(synthetic_workbook, with_meta, {"xl/metadata": "\x00binary blob"})

    without = tmp_path / "without.xlsx"
    shutil.copy2(synthetic_workbook, without)
    problems = compare(snapshot(with_meta), snapshot(without),
                       require_cached_values=False)
    assert not any("metadata" in p for p in problems)


def test_an_unknown_dropped_part_still_fails_the_gate(synthetic_workbook, tmp_path):
    with_extra = tmp_path / "with_extra.xlsx"
    _add_parts(synthetic_workbook, with_extra,
               {"xl/pivotCache/pivotCacheDefinition1.xml": "<pivot/>"})

    without = tmp_path / "without.xlsx"
    shutil.copy2(synthetic_workbook, without)
    problems = compare(snapshot(with_extra), snapshot(without),
                       require_cached_values=False)
    assert any("pivotCache" in p for p in problems)


def test_date_verified_comes_from_the_evidence_not_the_clock(synthetic_config):
    """A run that crosses midnight must not stamp rows with a date it did not verify."""
    from datetime import datetime as _dt

    view = load_workbook_view(synthetic_config)
    change_a = change(2, "I", "general_email", "info@summitlodge.example")
    change_a.fetched_at = _dt(2026, 8, 23, 23, 55, 0)

    # The run is written on the 24th; the evidence was gathered on the 23rd.
    provenance = build_provenance_changes(
        [change_a], synthetic_config, view, run_date="2026-08-24"
    )
    by_column = {p.column: p.new_value for p in provenance}
    assert by_column["AL"] == "2026-08-23", "should follow the fetch, not the write"


def test_date_verified_uses_the_most_recent_contributing_fetch(synthetic_config):
    from datetime import datetime as _dt

    view = load_workbook_view(synthetic_config)
    older = change(2, "I", "general_email", "info@summitlodge.example")
    older.fetched_at = _dt(2026, 8, 23, 10, 0)
    newer = change(2, "K", "phone", "+41274720000")
    newer.fetched_at = _dt(2026, 8, 24, 9, 0)

    provenance = build_provenance_changes(
        [older, newer], synthetic_config, view, run_date="2026-09-01"
    )
    by_column = {p.column: p.new_value for p in provenance}
    assert by_column["AL"] == "2026-08-24"


def test_an_unchanged_provenance_cell_is_not_rewritten(synthetic_config):
    """The fixture is already dated 2026-08-21; re-stamping it would be a no-op write."""
    from datetime import datetime as _dt

    view = load_workbook_view(synthetic_config)
    same_day = change(2, "I", "general_email", "info@summitlodge.example")
    same_day.fetched_at = _dt(2026, 8, 21, 9, 0)

    provenance = build_provenance_changes(
        [same_day], synthetic_config, view, run_date="2026-08-21"
    )
    assert "AL" not in {p.column for p in provenance}
    assert {"AK", "AM"} <= {p.column for p in provenance}


# ---------------------------------------------------------------------------
# Only the workbook goes to Drive; every process artifact stays local
# ---------------------------------------------------------------------------

def test_artifacts_and_workbook_have_separate_destinations(synthetic_config):
    assert synthetic_config.artifacts_directory != synthetic_config.output_directory


def test_config_refuses_artifacts_in_the_drive_folder(synthetic_config):
    """A hand-edited config must not be able to sync reports to a shared folder."""
    synthetic_config.artifacts_dir = str(synthetic_config.output_directory)
    synthetic_config.dry_run_dir = None
    problems = synthetic_config.sanity_check()
    assert any("artifacts_dir must not be the same as output_dir" in p for p in problems)


def test_write_outputs_puts_everything_in_one_local_directory(synthetic_config):
    from datetime import datetime as _dt

    from efe.pipeline import RunOutcome
    from efe.report import build_summary, write_outputs

    view = load_workbook_view(synthetic_config)
    outcome = RunOutcome()
    summary = build_summary(
        synthetic_config, view, outcome, run_id="test-artifacts", round_id="TEST",
        started_at=_dt(2026, 8, 22, 9, 0), dry_run=True, selected=0, skipped={},
    )
    written = write_outputs(synthetic_config.artifacts_directory, "stem", summary, outcome)

    assert set(written) == {"report", "json", "changes", "review", "decisions"}
    for label, path in written.items():
        assert path.parent == synthetic_config.artifacts_directory, label
        assert path.is_file(), label
    # Nothing reached the Drive folder.
    assert list(synthetic_config.output_directory.glob("*.md")) == []
    assert list(synthetic_config.output_directory.glob("*.csv")) == []
    assert list(synthetic_config.output_directory.glob("*.json")) == []


def test_a_real_write_sends_only_the_xlsx_to_the_output_dir(synthetic_config):
    view = load_workbook_view(synthetic_config)
    written = write_enriched(
        synthetic_config, view,
        [change(2, "I", "general_email", "info@summitlodge.example")],
        run_id="test-split",
    )
    assert written.parent == synthetic_config.output_directory
    emitted = sorted(p.name for p in synthetic_config.output_directory.iterdir())
    assert emitted == [written.name], f"Drive folder should hold only the workbook: {emitted}"


# ---------------------------------------------------------------------------
# Spreadsheet apps pad the used range; that is cosmetic
# ---------------------------------------------------------------------------

def test_trailing_empty_rows_do_not_fail_the_schema_check(synthetic_config,
                                                          synthetic_workbook):
    """A Google Sheets round trip left the real workbook reporting 1000 rows."""
    from efe.workbook.reader import last_data_row

    wb = load_workbook(synthetic_workbook)
    ws = wb["PARTNERS"]
    real_last = ws.max_row
    # Pad the used range the way a spreadsheet app does: cells that exist in the
    # sheet XML but hold nothing. The key columns (ID, Entity_Name) stay untouched.
    for row in range(real_last + 1, real_last + 400):
        ws.cell(row, 4).value = ""
    wb.save(synthetic_workbook)
    wb.close()

    reopened = load_workbook(synthetic_workbook)
    assert reopened["PARTNERS"].max_row > real_last, "the fixture must actually be padded"
    assert last_data_row(reopened["PARTNERS"], synthetic_config.workbook) == real_last
    reopened.close()

    # The schema check passes, and only the real rows are loaded.
    view = load_workbook_view(synthetic_config)
    assert len(view.rows) == real_last - 1
    assert all(r.get("B") for r in view.rows), "no blank rows should be loaded"


def test_a_genuinely_short_sheet_still_fails(synthetic_config, synthetic_workbook):
    """Tolerating trailing blanks must not tolerate missing data.

    There is no configured row count any more; the baseline is what the last run
    recorded. Fewer rows than that is an old file or a half-finished sync.
    """
    from efe.models import ContinuityError

    assert load_workbook_view(synthetic_config).data_rows == 8

    wb = load_workbook(synthetic_workbook)
    ws = wb["PARTNERS"]
    for column in range(1, ws.max_column + 1):
        ws.cell(ws.max_row, column).value = None
    wb.save(synthetic_workbook)
    wb.close()

    with pytest.raises(ContinuityError, match="fewer data rows than last run"):
        load_workbook_view(synthetic_config)
    # Re-baselining is possible, but only when asked for explicitly.
    assert load_workbook_view(synthetic_config, reset_state=True).data_rows == 7
