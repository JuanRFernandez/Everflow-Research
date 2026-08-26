"""Shared test fixtures.

The suite is fully offline: no network, and no dependency on the real workbook.
`synthetic_workbook` builds a stand-in with the same structure the writer relies on
-- the PARTNERS column layout, a DASHBOARD of COUNTIF formulas over it, data
validations, an autofilter and cached formula results -- so the whole write and
verify path is exercised without touching Google Drive.

The HTML fixtures are invented. Real pages from these companies name real people;
committing them would put GDPR personal data into git history permanently.
"""

from __future__ import annotations

import re
import shutil
import zipfile
from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from efe import config as config_mod

FIXTURES = Path(__file__).parent / "fixtures"

PARTNERS_HEADERS = [
    "ID", "Entity_Name", "Category", "Subcategory", "Resort_Base", "Region_Valley",
    "Country", "Website_URL", "General_Email", "Sales_B2B_Email", "Phone", "WhatsApp",
    "Contact_Person_Name", "Contact_Person_Role", "LinkedIn_URL", "Instagram_Handle",
    "Segment_Tier", "Star_Rating_or_Class", "Capacity_Keys_or_Beds",
    "B2B_Program_Exists", "Commission_or_Partner_Terms", "Languages_Served",
    "Owner_or_Group_Affiliation", "Strategic_Fit_Note", "Priority_Score", "Contacted",
    "Contact_Date", "Follow_Up_Days", "Next_Follow_Up", "Email_Sent", "Call_Made",
    "WhatsApp_Sent", "Meeting_Booked", "Agreement_Signed", "Status", "Next_Action",
    "Source_URL", "Date_Verified", "Round", "Material_Sent",
]

#: (name, category, country, website, contacted, general_email, resort)
#: Rows 5 and 6 deliberately share a domain, so the scope guard has a group case.
SYNTHETIC_ROWS = [
    ("Summit Lodge Verbier", "1. Hotels", "CH",
     "https://summitlodge.example", "NO", "TBD", "Verbier"),
    ("Berghotel Silberdistel", "1. Hotels", "DE",
     "https://silberdistel-berghotel.example", "NO", "TBD", "Garmisch"),
    ("Chalet Belle Etoile", "2. Chalets & Chalet Management", "FR",
     "https://chalet-belle-etoile.example", "NO", "TBD", "Courchevel"),
    ("W Verbier", "1. Hotels", "CH",
     "https://grandclass.example", "NO", "TBD", "Verbier"),
    ("Grandclass Courchevel", "1. Hotels", "FR",
     "https://grandclass.example", "NO", "TBD", "Courchevel"),
    ("Already Contacted Agency", "6. Distribution & Sales Agencies", "Brasil",
     "https://contacted.example", "YES", "hello@contacted.example", ""),
    ("Filled Row", "1. Hotels", "AT",
     "https://filled.example", "NO", "info@filled.example", "Kitzbuehel"),
    ("No Website Row", "10. Catering & Private Chefs", "FR",
     "TBD", "NO", "TBD", ""),
]

_CELL_RE = re.compile(r"<c\b([^>]*?)(/>|>((?:(?!</c>).)*?)</c>)", re.S)


def inject_cached_values(path: Path, values: dict[tuple[str, str], str]) -> None:
    """Give formula cells a cached `<v>`, the way Excel and LibreOffice do.

    openpyxl cannot write cached results, so the fixture would otherwise not
    reproduce the condition `reinject_cached_values` exists to repair.
    """
    from efe.workbook.xmlutil import sheet_part_map

    parts = sheet_part_map(path)
    with zipfile.ZipFile(path) as zf:
        contents = {n: zf.read(n) for n in zf.namelist()}
        order = list(zf.namelist())

    for title, part in parts.items():
        wanted = {ref: v for (t, ref), v in values.items() if t == title}
        if not wanted or part not in contents:
            continue
        xml = contents[part].decode("utf-8")

        def _fix(m: re.Match[str], *, wanted: dict[str, str] = wanted) -> str:
            attrs, closing, body = m.group(1), m.group(2), m.group(3) or ""
            if closing == "/>" or "<f" not in body:
                return m.group(0)
            ref = re.search(r'r="([A-Z]+\d+)"', attrs)
            if not ref or ref.group(1) not in wanted:
                return m.group(0)
            value = wanted[ref.group(1)]
            body = re.sub(r"<v\s*/>|<v></v>", "", body)
            return f'<c{attrs} t="n">{body}<v>{value}</v></c>'

        contents[part] = _CELL_RE.sub(_fix, xml).encode("utf-8")

    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zf:
        for name in order:
            zf.writestr(name, contents[name])
    shutil.move(str(tmp), str(path))


@pytest.fixture
def synthetic_workbook(tmp_path: Path) -> Path:
    """A stand-in workbook with the real structure: formulas, DV, autofilter, cache."""
    wb = Workbook()
    readme = wb.active
    readme.title = "READ_ME"
    readme["A1"] = "SYNTHETIC TEST WORKBOOK"

    dashboard = wb.create_sheet("DASHBOARD")
    dashboard["A1"] = "Category"
    dashboard["B1"] = "Total"
    for offset, category in enumerate(["1. Hotels", "2. Chalets & Chalet Management"]):
        row = 2 + offset
        dashboard.cell(row, 1).value = category
        dashboard.cell(row, 2).value = f"=COUNTIF(PARTNERS!$C$2:$C$9,A{row})"
    dashboard["B5"] = '=COUNTIFS(PARTNERS!$Z$2:$Z$9,"YES")'

    partners = wb.create_sheet("PARTNERS")
    for index, header in enumerate(PARTNERS_HEADERS, start=1):
        partners.cell(1, index).value = header

    for offset, (name, category, country, website, contacted, email, resort) in enumerate(
        SYNTHETIC_ROWS
    ):
        row = 2 + offset
        partners.cell(row, 1).value = f"EFE-{offset + 1:04d}"
        partners.cell(row, 2).value = name
        partners.cell(row, 3).value = category
        partners.cell(row, 5).value = resort
        partners.cell(row, 7).value = country
        partners.cell(row, 8).value = website
        partners.cell(row, 9).value = email
        for column in (10, 11, 12, 13, 14, 15, 16, 21):
            partners.cell(row, column).value = "TBD"
        partners.cell(row, 20).value = "Unknown"
        partners.cell(row, 25).value = 3
        partners.cell(row, 26).value = contacted
        partners.cell(row, 28).value = 14
        partners.cell(row, 29).value = f'=IFERROR(AA{row}+AB{row},"")'
        partners.cell(row, 35).value = "Contacted" if contacted == "YES" else "Not started"
        partners.cell(row, 37).value = "TBD"
        partners.cell(row, 38).value = "2026-08-21"
        partners.cell(row, 39).value = "Pre-existing" if contacted == "YES" else "R1"

    partners.auto_filter.ref = "A1:AM9"
    partners.freeze_panes = "C2"
    partners.add_data_validation(_dv('"NO,YES"', "Z2:Z9"))
    partners.add_data_validation(_dv('"1,2,3,4,5"', "Y2:Y9"))

    for name in ("RESORTS_SBI", "PRICING_BENCH", "REGULATORY", "_SOURCES",
                 "_GAPS_ROUND2", "CHANGELOG"):
        wb.create_sheet(name)

    sources = wb["_SOURCES"]
    sources["A4"] = "#"
    sources["B4"] = "Domain"
    sources["C4"] = "Category_Covered"
    sources["D4"] = "Round"
    sources["E4"] = "Exclude_Next_Round"
    sources["A5"], sources["B5"] = 1, "summitlodge.example"
    sources["C5"], sources["D5"], sources["E5"] = "Hotels", "R1", "YES"

    changelog = wb["CHANGELOG"]
    changelog["A1"] = "VERSION HISTORY"
    changelog["A3"], changelog["B3"] = "Version", "Date"
    changelog["C3"], changelog["D3"] = "Change", "Author"
    changelog["A4"], changelog["B4"] = "v01", "2026-08-21"
    changelog["C4"], changelog["D4"] = "Initial synthetic build.", "tests"

    # The real workbook already carries the audit sheet from the previous run
    # (a required sheet since v05); the writer appends to it rather than creating it.
    from efe.workbook.writer import CHANGELOG_DETAIL_HEADERS

    detail = wb.create_sheet("CHANGELOG_DETAIL")
    for index, title in enumerate(CHANGELOG_DETAIL_HEADERS, start=1):
        detail.cell(1, index).value = title
    prior = ["2026-08-21T10:00:00", "20260821-100000", 2, "EFE-0001", "Summit Lodge Verbier",
             "K", "phone", "TBD", "+41 27 000 00 00", "high", "corporate_role",
             "https://summitlodge.example/contact", "2026-08-21T09:59:00",
             "phones.text", "prior run"]
    for index, value in enumerate(prior, start=1):
        detail.cell(2, index).value = value

    path = tmp_path / "2026-08-21_EFE_Alpine_Partner_Database_v01.xlsx"
    wb.save(path)
    wb.close()

    cached = {("PARTNERS", f"AC{2 + i}"): "14" for i in range(len(SYNTHETIC_ROWS))}
    cached[("DASHBOARD", "B2")] = "4"
    cached[("DASHBOARD", "B3")] = "1"
    cached[("DASHBOARD", "B5")] = "1"
    inject_cached_values(path, cached)
    return path


def _dv(formula: str, sqref: str) -> DataValidation:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.sqref = sqref
    return validation


@pytest.fixture
def synthetic_config(tmp_path: Path, synthetic_workbook: Path):
    """Real `config.yaml`, repointed at the synthetic workbook and a temp data dir."""
    repo_config = Path(__file__).resolve().parents[1] / "config.yaml"
    raw = yaml.safe_load(repo_config.read_text(encoding="utf-8"))

    # The synthetic workbook sits alone in tmp_path as `..._v01.xlsx`, so the
    # resolver picks it exactly as it would pick the real one in the Drive folder.
    raw.pop("workbook_path", None)
    raw["workbook_dir"] = str(synthetic_workbook.parent).replace("\\", "/")
    raw["output_dir"] = str(tmp_path / "out").replace("\\", "/")
    raw["cache_dir"] = str(tmp_path / "cache").replace("\\", "/")
    raw["state_dir"] = str(tmp_path / "state").replace("\\", "/")
    raw["log_dir"] = str(tmp_path / "logs").replace("\\", "/")
    raw["artifacts_dir"] = str(tmp_path / "artifacts").replace("\\", "/")
    raw.pop("dry_run_dir", None)
    # The synthetic workbook is ~10 KB; the real one is ~700 KB.
    raw["workbook"]["min_plausible_bytes"] = 4000
    # The live config ships with hotel targeting active; the legacy fixtures test
    # the unfiltered pipeline, so the synthetic config clears the filters.
    raw["selection"]["categories"] = []
    raw["selection"]["resorts"] = []
    raw["scope"]["chain_domains"] = ["grandclass.example"]

    (tmp_path / "out").mkdir(exist_ok=True)
    config_path = tmp_path / "config.yaml"
    config_path.write_text(yaml.safe_dump(raw, allow_unicode=True), encoding="utf-8")
    return config_mod.load(config_path)


@pytest.fixture
def real_config():
    """The repo's real config, for tests that only read vocabulary and column maps."""
    return config_mod.load(Path(__file__).resolve().parents[1] / "config.yaml")


def fixture_text(name: str) -> str:
    return (FIXTURES / name).read_text(encoding="utf-8")
