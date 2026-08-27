"""Range drift: what the sheet's rectangles no longer cover, measured from the file."""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from efe.models import VerificationError
from efe.workbook.ranges import (
    assert_touches_no_formula_input,
    inspect_ranges,
    inspect_sheet_ranges,
    referenced_columns,
    render_sheets_handoff,
)


def test_dashboard_categories_are_read_from_the_formulas_not_hardcoded(synthetic_config):
    """The fixture's DASHBOARD counts two categories via `COUNTIF(...,A2)` criteria."""
    ranges = inspect_sheet_ranges(
        synthetic_config.workbook_directory / _only(synthetic_config), synthetic_config
    )
    assert ranges.dashboard_categories == ["1. Hotels", "2. Chalets & Chalet Management"]


def test_category_gap_names_a_partners_category_no_dashboard_row_counts(synthetic_config):
    ranges = inspect_sheet_ranges(
        synthetic_config.workbook_directory / _only(synthetic_config), synthetic_config
    )
    assert ranges.uncounted_categories == {
        "6. Distribution & Sales Agencies": 1,
        "10. Catering & Private Chefs": 1,
    }
    drift = "\n".join(ranges.drift())
    assert "is on 1 PARTNERS row(s) and on no DASHBOARD row" in drift


def test_the_three_numbers_promote_reports_are_unchanged(synthetic_config):
    """`inspect_ranges` moved module; its shape is what `efe promote` still depends on."""
    path = synthetic_config.workbook_directory / _only(synthetic_config)
    assert inspect_ranges(path, synthetic_config) == {
        "autofilter_last_row": 9,
        "validation_last_row": 9,
        "formula_reach_last_row": 9,
    }


def test_handoff_names_the_stale_autofilter_column(synthetic_config):
    """The fixture's autofilter ends at AM while the header reaches AN (Material_Sent)."""
    path = synthetic_config.workbook_directory / _only(synthetic_config)
    ranges = inspect_sheet_ranges(path, synthetic_config)
    assert ranges.autofilter_last_col == "AM" and ranges.header_last_col == "AN"
    handoff = "\n".join(render_sheets_handoff(ranges, synthetic_config))
    assert "Autofilter. Select A1:AN9" in handoff


def test_drift_is_quiet_when_every_rectangle_covers_the_data(synthetic_config, synthetic_workbook):
    wb = load_workbook(synthetic_workbook)
    ws = wb["PARTNERS"]
    ws.auto_filter.ref = "A1:AN9"
    dash = wb["DASHBOARD"]
    dash["A4"], dash["B4"] = "6. Distribution & Sales Agencies", "=COUNTIF(PARTNERS!$C$2:$C$9,A4)"
    dash["A5"] = "10. Catering & Private Chefs"
    dash["B5"] = "=COUNTIF(PARTNERS!$C$2:$C$9,A5)"
    wb.save(synthetic_workbook)
    wb.close()
    ranges = inspect_sheet_ranges(synthetic_workbook, synthetic_config)
    assert ranges.drift() == []
    assert render_sheets_handoff(ranges, synthetic_config) == []


def test_formula_shape_gap_finds_the_hand_edited_rows(synthetic_config, synthetic_workbook):
    wb = load_workbook(synthetic_workbook)
    wb["PARTNERS"]["AC9"] = "=AA9+AB9"
    wb.save(synthetic_workbook)
    wb.close()
    ranges = inspect_sheet_ranges(synthetic_workbook, synthetic_config)
    dominant, rows = ranges.formula_shape_gaps["Next_Follow_Up"]
    assert rows == [9] and dominant == '=IFERROR(AA2+AB2,"")'
    assert any("different Next_Follow_Up formula" in line for line in ranges.drift())


def test_referenced_columns_sees_both_qualified_and_bare_references(synthetic_config):
    path = synthetic_config.workbook_directory / _only(synthetic_config)
    columns = referenced_columns(path, synthetic_config)
    assert {"C", "Z"} <= columns  # DASHBOARD COUNTIFs over PARTNERS
    assert {"AA", "AB"} <= columns  # PARTNERS!AC reads them without a sheet prefix
    assert "A" not in columns  # nothing reads the ID column -- which is why fixup may write it


def test_assert_touches_no_formula_input_catches_a_new_reference(
    synthetic_config, synthetic_workbook
):
    path = synthetic_workbook
    assert_touches_no_formula_input(path, synthetic_config, {"A"})  # nobody reads A today

    wb = load_workbook(path)
    wb["DASHBOARD"]["B9"] = "=COUNTA(PARTNERS!A:A)"
    wb.save(path)
    wb.close()
    with pytest.raises(
        VerificationError, match=r"Columns \['ID'\] \(\['A'\]\) are read by a formula"
    ):
        assert_touches_no_formula_input(path, synthetic_config, {"A"})


def _only(cfg) -> str:
    """The single workbook the synthetic fixture put in the folder."""
    return next(p.name for p in cfg.workbook_directory.glob("*.xlsx"))
