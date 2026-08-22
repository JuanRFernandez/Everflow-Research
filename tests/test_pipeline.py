"""End-to-end enrichment, offline.

httpx is driven by a `MockTransport` serving the synthetic fixtures, so the whole
path -- robots, discovery, fetch, cache, extract, classify, scope, choose, write --
runs without a single packet leaving the machine.
"""

from __future__ import annotations

import asyncio

import httpx
import pytest

from efe.fetch.cache import PageCache
from efe.fetch.client import Fetcher
from efe.models import Confidence, DataClass
from efe.pipeline import RunLedger, run_enrichment
from efe.report import build_summary, render_markdown, render_review_csv
from efe.workbook.reader import load_workbook_view, select_candidates
from efe.workbook.writer import write_enriched
from tests.conftest import fixture_text

ROUTES = {
    # Summit Lodge Verbier -- its own domain, a real contact page and a trade page.
    "summitlodge.example/": "homepage_with_links.html",
    "summitlodge.example/robots.txt": "robots_disallow.txt",
    "summitlodge.example/sitemap.xml": "sitemap.xml",
    "summitlodge.example/en/contact": "socials_footer.html",
    "summitlodge.example/en/trade": "trade_terms.html",
    "summitlodge.example/en/about-us": "team_roles.html",
    "summitlodge.example/impressum": "impressum_de.html",
    # Berghotel Silberdistel -- German, so the Impressum is probed first.
    "silberdistel-berghotel.example/": "homepage_with_links.html",
    "silberdistel-berghotel.example/impressum": "impressum_de.html",
    # Chalet Belle Etoile -- French contact page.
    "chalet-belle-etoile.example/": "homepage_with_links.html",
    "chalet-belle-etoile.example/contact": "contact_fr.html",
    # A group domain shared by two rows: one page names a property, one does not.
    "grandclass.example/": "chain_group_contact.html",
    "grandclass.example/contact": "chain_group_contact.html",
    "grandclass.example/hotels/w-verbier": "chain_property_page.html",
    "filled.example/": "homepage_with_links.html",
    "filled.example/kontakt": "kontakt_at.html",
}


def _handler(request: httpx.Request) -> httpx.Response:
    key = f"{request.url.host}{request.url.path}"
    name = ROUTES.get(key)
    if name is None:
        return httpx.Response(404, text="not found")
    body = fixture_text(name)
    media = "application/xml" if name.endswith(".xml") else (
        "text/plain" if name.endswith(".txt") else "text/html; charset=utf-8"
    )
    return httpx.Response(200, text=body, headers={"content-type": media})


@pytest.fixture
def fast_config(synthetic_config):
    """Same config without the politeness waits, so the suite runs in seconds.

    The robots fixture declares `Crawl-delay: 5`, which the fetcher rightly honours
    in production. Obeying it here would make this module take three minutes; the
    parsing and honouring of Crawl-delay is covered directly in the robots tests.
    """
    synthetic_config.fetch.per_domain_delay_seconds = 0.0
    synthetic_config.fetch.honour_crawl_delay = False
    synthetic_config.fetch.max_pages_per_entity = 6
    return synthetic_config


def run(cfg, candidates, **kwargs):
    view = kwargs.pop("view")

    async def go():
        cache = PageCache(cfg.cache_directory, enabled=True)
        cfg.cache_directory.mkdir(parents=True, exist_ok=True)
        transport = httpx.MockTransport(_handler)
        client = httpx.AsyncClient(transport=transport, follow_redirects=True)
        async with Fetcher(cfg.fetch, cache, client=client) as fetcher:
            try:
                return await run_enrichment(
                    cfg, view, candidates, round_id="TEST", run_id="test-run",
                    fetcher=fetcher, **kwargs
                )
            finally:
                await client.aclose()

    return asyncio.run(go())


@pytest.fixture
def outcome(fast_config):
    view = load_workbook_view(fast_config)
    candidates, _ = select_candidates(view, fast_config)
    return view, candidates, run(fast_config, candidates, view=view)


# ---------------------------------------------------------------------------

def test_pipeline_processes_every_selected_row(outcome):
    _, candidates, result = outcome
    assert len(result.results) == len(candidates) == 6


def test_role_addresses_are_written_to_the_right_columns(fast_config, outcome):
    _, _, result = outcome
    written = {(c.row, c.column): c.new_value for c in result.changes}

    # Summit Lodge is row 2. Its trade page publishes traveltrade@ and info@.
    assert written.get((2, "J")) == "traveltrade@summitlodge.example"
    assert written.get((2, "I")) == "info@summitlodge.example"


def test_an_address_from_another_domain_is_never_written(outcome):
    """The routes serve a foreign Impressum on summitlodge.example on purpose.

    Its `info@silberdistel-berghotel.example` is genuinely published on the page,
    and is genuinely the wrong contact for Summit Lodge. It must be held, not
    written -- the same way a parent company's or web agency's address would be.
    """
    _, _, result = outcome
    foreign = "info@silberdistel-berghotel.example"
    assert not [c for c in result.changes if c.row == 2 and c.new_value == foreign]
    held = [c for c in result.held if c.row == 2 and c.new_value == foreign]
    assert held
    assert "does not match the site domain" in held[0].note


def test_socials_are_captured(outcome):
    _, _, result = outcome
    written = {(c.row, c.column): c.new_value for c in result.changes}
    assert written.get((2, "P")) == "@summitlodgeverbier"
    assert written.get((2, "O")) == "https://www.linkedin.com/company/summit-lodge-verbier"


def test_published_trade_terms_are_captured_verbatim(outcome):
    _, _, result = outcome
    terms = [c for c in result.changes if c.column == "U"]
    assert terms, "the trade page states a commission and should have been captured"
    assert "commission of 10%" in terms[0].new_value


def test_impressum_yields_a_named_representative_with_a_role(outcome):
    _, _, result = outcome
    by_row = {(c.row, c.column): c for c in result.changes}
    # Row 3 is the German entity; its Impressum names the Geschaeftsfuehrer.
    assert by_row[(3, "M")].new_value == "Katharina Meier"
    assert by_row[(3, "N")].new_value in ("Geschäftsführer", "Vertretungsberechtigter (Impressum)")
    assert by_row[(3, "M")].data_class is DataClass.PERSONAL_NAMED


def test_person_name_and_role_are_written_together_or_not_at_all(outcome):
    _, _, result = outcome
    rows_with_name = {c.row for c in result.changes if c.column == "M"}
    rows_with_role = {c.row for c in result.changes if c.column == "N"}
    assert rows_with_name == rows_with_role


def test_group_page_does_not_fill_the_property_row(outcome):
    """Row 5 is W Verbier on a shared group domain. Nothing may be written."""
    _, _, result = outcome
    assert not [c for c in result.changes if c.row == 5], (
        "a group-level contact must never be written onto a property row"
    )
    held = [c for c in result.held if c.row == 5]
    assert held
    assert any("group/chain domain" in c.note for c in held)


def test_named_individual_addresses_never_reach_the_workbook(outcome):
    _, _, result = outcome
    for change in result.changes:
        if change.column in ("I", "J"):
            assert change.data_class is DataClass.CORPORATE_ROLE


def test_only_high_confidence_values_are_written(outcome):
    _, _, result = outcome
    assert all(c.confidence is Confidence.HIGH for c in result.changes)


def test_every_written_value_has_provenance(outcome):
    _, _, result = outcome
    for change in result.changes:
        assert change.source_url.startswith("http")
        assert change.fetched_at is not None


def test_filled_cells_are_never_contested(outcome):
    """Row 8 already has a General_Email, so nothing may be written to column I."""
    _, _, result = outcome
    assert not [c for c in result.changes if c.row == 8 and c.column == "I"]
    assert any(c.row == 8 and c.column == "I" for c in result.held)


def test_robots_disallowed_paths_are_never_fetched(fast_config, outcome):
    _, _, result = outcome
    for url in [u for r in result.results for u in r.pages_fetched]:
        assert "/private/" not in url
        assert "/booking-engine/" not in url


def test_ledger_domain_revisits_are_logged(outcome):
    _, _, result = outcome
    domains = {entry["domain"] for entry in result.revisited}
    assert "summitlodge.example" in domains
    assert all("discovery" in entry["reason"] for entry in result.revisited)


# ---------------------------------------------------------------------------
# Resume
# ---------------------------------------------------------------------------

def test_run_is_resumable(fast_config):
    view = load_workbook_view(fast_config)
    candidates, _ = select_candidates(view, fast_config)
    ledger = RunLedger(fast_config.state_directory, "TEST", "run-1")

    first = run(fast_config, candidates[:2], view=view, ledger=ledger)
    assert len(first.results) == 2
    assert ledger.completed_entities() == {c.entity_id for c in candidates[:2]}

    # A second run over the full list skips what the first one finished.
    second = run(fast_config, candidates, view=view, ledger=ledger)
    assert len(second.results) == len(candidates) - 2
    assert ledger.path.is_file()


def test_ledger_records_are_entity_field_shaped(fast_config):
    view = load_workbook_view(fast_config)
    candidates, _ = select_candidates(view, fast_config)
    ledger = RunLedger(fast_config.state_directory, "TEST", "run-1")
    run(fast_config, candidates[:1], view=view, ledger=ledger)

    import json

    lines = [json.loads(x) for x in ledger.path.read_text(encoding="utf-8").splitlines()]
    assert lines
    for record in lines:
        # The Phase-1 `entity_field` columns must all be present.
        for key in ("entity_id", "field", "value", "confidence", "source_url",
                    "fetched_at", "round_id"):
            assert key in record
        assert record["round_id"] == "TEST"


# ---------------------------------------------------------------------------
# Reporting and the full write
# ---------------------------------------------------------------------------

def test_reports_render(fast_config, outcome):
    view, candidates, result = outcome
    _, skipped = select_candidates(view, fast_config)
    from datetime import datetime

    summary = build_summary(
        fast_config, view, result, run_id="test-run", round_id="TEST",
        started_at=datetime(2026, 8, 21, 9, 0), dry_run=True,
        selected=len(candidates), skipped=skipped,
    )
    markdown = render_markdown(summary, result)
    assert "# EFE enrichment run" in markdown
    assert "Round-1 ledger domains revisited" in markdown
    assert "Rows needing a property-level Website_URL" in markdown

    csv_text = render_review_csv(result.held)
    assert "Held_Back_Because" in csv_text
    assert csv_text.count("\n") > 1


def test_full_write_passes_the_fidelity_gate(fast_config, outcome):
    view, _, result = outcome
    written = write_enriched(
        fast_config, view, result.changes,
        run_id="test-run", held_back=result.held,
    )
    assert written.is_file()

    from openpyxl import load_workbook

    wb = load_workbook(written)
    assert wb["PARTNERS"]["J2"].value == "traveltrade@summitlodge.example"
    assert wb["PARTNERS"]["I5"].value == "TBD", "group row must stay TBD"
    assert fast_config.workbook.changelog_detail_sheet in wb.sheetnames
    wb.close()


def test_cache_makes_a_second_run_free(fast_config):
    view = load_workbook_view(fast_config)
    candidates, _ = select_candidates(view, fast_config)
    first = run(fast_config, candidates[:1], view=view)
    second = run(fast_config, candidates[:1], view=view)
    assert second.cache_hits > 0
    assert second.pages_fetched < first.pages_fetched
