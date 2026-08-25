"""Ad-hoc hotel discovery for the GaPa · Arlberg · Kitzbühel · Innsbruck · Stubai ·
Zell-Kaprun corridor, from official tourism-board pages.

This is the one-off precursor to the Phase-2 `resort_directory` source plugin
(ARCHITECTURE.md §3), not the plugin itself. It follows the same rule as the
enricher: **nothing is invented**. A candidate row carries only what the source page
literally publishes -- the property's name, the website it links to, the stars it
states, the postcode line it prints -- plus the URL of the page it came from.

Output is a CSV of *candidates* in the exact PARTNERS column layout. Per
ARCHITECTURE §1, discovery produces candidates and a human promotes them: nothing is
written to the workbook. Contact fields stay TBD; the enricher fills them after
promotion.

Run:  uv run python scripts/discover_hotels.py
"""

from __future__ import annotations

import asyncio
import csv
import io
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from urllib.parse import urljoin, urlparse

from selectolax.parser import HTMLParser

from efe import config
from efe.dedupe import fold, name_tokens
from efe.fetch.cache import PageCache
from efe.fetch.client import Fetcher
from efe.fetch.discovery import parse_sitemap
from efe.workbook.reader import domain_of, load_workbook_view

FIRST_ID = 359  # 0254-0280 = Juan's hotels (v05); 0281-0358 = repo-queue merge CSV
ROUND = "R3-discovery"
PER_RESORT_CAP = 20

# ---------------------------------------------------------------------------
# The corridor: town -> (Resort_Base, Region_Valley, Country). Matched on the
# postcode line a source prints, so "St. Anton am Arlberg" and "St Anton" both land
# on Arlberg, and a bare postcode still resolves.
# ---------------------------------------------------------------------------
CORRIDOR: dict[str, tuple[str, str, str]] = {
    "garmisch-partenkirchen": ("Garmisch-Partenkirchen", "Zugspitz Region / Werdenfels", "DE"),
    "garmisch": ("Garmisch-Partenkirchen", "Zugspitz Region / Werdenfels", "DE"),
    "grainau": ("Grainau", "Zugspitz Region / Werdenfels", "DE"),
    "krün": ("Krün / Elmau", "Zugspitz Region / Werdenfels", "DE"),
    "kruen": ("Krün / Elmau", "Zugspitz Region / Werdenfels", "DE"),
    "elmau": ("Krün / Elmau", "Zugspitz Region / Werdenfels", "DE"),
    "klais": ("Krün / Elmau", "Zugspitz Region / Werdenfels", "DE"),
    "lech am arlberg": ("Lech", "Arlberg", "AT"),
    "lech": ("Lech", "Arlberg", "AT"),
    "oberlech": ("Lech", "Arlberg", "AT"),
    "zürs": ("Zürs", "Arlberg", "AT"),
    "zuers": ("Zürs", "Arlberg", "AT"),
    "st. anton am arlberg": ("St. Anton am Arlberg", "Arlberg", "AT"),
    "st. anton": ("St. Anton am Arlberg", "Arlberg", "AT"),
    "st anton": ("St. Anton am Arlberg", "Arlberg", "AT"),
    "st. christoph": ("St. Christoph am Arlberg", "Arlberg", "AT"),
    "st christoph": ("St. Christoph am Arlberg", "Arlberg", "AT"),
    "kitzbühel": ("Kitzbühel", "Kitzbüheler Alpen", "AT"),
    "kitzbuehel": ("Kitzbühel", "Kitzbüheler Alpen", "AT"),
    "jochberg": ("Jochberg", "Kitzbüheler Alpen", "AT"),
    "reith bei kitzbühel": ("Kitzbühel", "Kitzbüheler Alpen", "AT"),
    "aurach bei kitzbühel": ("Kitzbühel", "Kitzbüheler Alpen", "AT"),
    "aurach": ("Kitzbühel", "Kitzbüheler Alpen", "AT"),
    "going am wilden kaiser": ("Going", "Wilder Kaiser", "AT"),
    "going": ("Going", "Wilder Kaiser", "AT"),
    "seefeld in tirol": ("Seefeld", "Seefeld Plateau", "AT"),
    "seefeld": ("Seefeld", "Seefeld Plateau", "AT"),
    "leutasch": ("Seefeld – Leutasch", "Seefeld Plateau", "AT"),
    "mösern": ("Seefeld – Mösern", "Seefeld Plateau", "AT"),
    "moesern": ("Seefeld – Mösern", "Seefeld Plateau", "AT"),
    "reith bei seefeld": ("Seefeld – Reith", "Seefeld Plateau", "AT"),
    "innsbruck": ("Innsbruck", "Innsbruck", "AT"),
    "igls": ("Innsbruck – Igls", "Innsbruck", "AT"),
    "neustift im stubaital": ("Neustift / Stubai", "Stubaital", "AT"),
    "neustift": ("Neustift / Stubai", "Stubaital", "AT"),
    "fulpmes": ("Stubai – Fulpmes", "Stubaital", "AT"),
    "mieders": ("Stubai – Mieders", "Stubaital", "AT"),
    "telfes": ("Stubai – Telfes", "Stubaital", "AT"),
    "schönberg im stubaital": ("Stubai – Schönberg", "Stubaital", "AT"),
    "zell am see": ("Zell am See", "Zell am See-Kaprun", "AT"),
    "kaprun": ("Kaprun", "Zell am See-Kaprun", "AT"),
}
#: Postcode -> town, for sources that print a code but abbreviate the town.
POSTCODES = {
    "82467": "garmisch-partenkirchen",
    "82491": "grainau",
    "82494": "krün",
    "82493": "klais",
    "6764": "lech am arlberg",
    "6763": "zürs",
    "6580": "st. anton am arlberg",
    "6370": "kitzbühel",
    "6373": "jochberg",
    "6371": "aurach",
    "6353": "going",
    "6100": "seefeld in tirol",
    "6105": "leutasch",
    "6103": "reith bei seefeld",
    "6020": "innsbruck",
    "6080": "igls",
    "6167": "neustift im stubaital",
    "6166": "fulpmes",
    "6142": "mieders",
    "6165": "telfes",
    "6141": "schönberg im stubaital",
    "5700": "zell am see",
    "5710": "kaprun",
}
#: Corridor towns as they appear in tirol.at detail-page slugs.
TIROL_SLUG_TOWNS = (
    "st-anton",
    "anton-am-arlberg",
    "st-christoph",
    "kitzbuehel",
    "kitzbühel",
    "jochberg",
    "going",
    "seefeld",
    "leutasch",
    "moesern",
    "innsbruck",
    "igls",
    "neustift",
    "stubai",
    "fulpmes",
    "mieders",
    "telfes",
)

#: Hosts that are never a hotel's own website: platforms, boards, vendors, socials.
SOCIAL = (
    "facebook", "instagram", "youtube", "twitter", "linkedin", "google", "apple", "tiktok",
    "pinterest", "vimeo", "booking", "feratel", "deskline", "tomas", "cookiebot", "onetrust",
    "addtoany", "pixxio", "goo.gl", "maps.", "newsletter", "spotify", "whatsapp", "m.me",
    "trustyou", "tour.gix", "moments.tirol", "tyrol", "tirolo", "visittirol", "table4u",
    "tmona", "huettenland", "erlebe.bayern", "zugspitz-region", "skiarlberg",
    "vorarlberg.travel", "bestofthealps", "remus.eu", "head.com", "landrover",
    "creativecommons", "static.et4", "meta.et4", "top.oberbayern", "tirolwerbung",
    "cine.tirol", "convention.tirol", "pixelart", "web-crossing", "cookiehub",
    "enable-javascript", "bergfex", "vorarlberg.at", "v-card", "golf", "gletscher",
    "ironman", "audi", "stubai.at", "kitzbuehel.com", "seefeld.com", "tirol.at",
    "lechzuers.com", "zellamsee-kaprun.com", "alpenwelt-karwendel", "gapa",
    "eichenheim.com", "zielhaus.at", "felbermayer", "streamdiver", "job-offers",
    "onepagebooking", "seekda", "hoteliers", "urlaubambauernhof", "bavaria.travel",
    "tripadvisor", "holidaycheck", "expedia", "hrs.", "trivago", "airbnb",
)  # fmt: skip
STAR_RE = re.compile(
    r"(?P<n>[3-5])\s?(?:\*|[sS]terne?)[\s-]?(?P<s>[sS]uperior|[sS]\b)?"
    r"|(?P<a>\*{3,5})\s?(?P<as>[sS]\b|[sS]uperior)?"
)
#: A postcode followed by a town, anywhere in a line ("Zugspitzstr. 8, 82491 Grainau").
POSTCODE_ANY = re.compile(r"(?<!\d)(\d{4,5})\s+([A-Za-zÄÖÜäöüß][^\d\n|,;()]{2,40}?)\s*(?:[,;|]|$)")
NOT_A_PROPERTY = re.compile(
    r"wanderung|tour\b|touren|\balm\b|kaffee|konditorei|bankerl|nebenhaus|\bfewo\b|hütte|huette"
    r"|bahn\b|lift|museum|kirche|loipe|route|parkplatz|gästehaus|gaestehaus|verleih|schule"
    r"|restaurant$|bar$|café$|cafe$|spa$|day spa|skyspace|mitfahr",
    re.I,
)
PROPERTY_WORD = re.compile(
    r"hotel|chalet|resort|lodge|residen[cz]|suite|villa|apartm|appartem|ferienwohnung"
    r"|landhaus|\bhof\b|haus\b|schloss|palais|alpin",
    re.I,
)


def host(u: str) -> str:
    h = urlparse(u).netloc.lower()
    return h[4:] if h.startswith("www.") else h


def root(u: str) -> str:
    """scheme://host of a linked site. The path is a deep link; the host is the site."""
    p = urlparse(u)
    return f"{p.scheme}://{p.netloc}" if p.scheme and p.netloc else u


def flat(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", fold(s or ""))


def flat_ue(s: str) -> str:
    """Flattened with German umlaut expansion (ü -> ue), for domain affinity."""
    t = (s or "").lower().replace("ä", "ae").replace("ö", "oe").replace("ü", "ue")
    return re.sub(r"[^a-z0-9]", "", fold(t.replace("ß", "ss")))


def stars_of(*texts: str) -> str:
    """Published star rating normalised to 5*S / 5* / 4*S / 4* / 3*, or ''."""
    for t in texts:
        if not t:
            continue
        for m in STAR_RE.finditer(t):
            n = m.group("n") or (str(len(m.group("a"))) if m.group("a") else "")
            s = "S" if (m.group("s") or m.group("as")) else ""
            if n in ("3", "4", "5"):
                return f"{n}*{s}"
    return ""


def strip_stars(name: str) -> str:
    name = re.sub(r"\s*\*{3,5}\s?(?:[sS]uperior|[sS])?(?=\s|$)", " ", name)
    name = re.sub(r"\s*\b[3-5]\s?\*\s?(?:[sS]uperior|[sS])?(?=\s|$)", " ", name)
    return re.sub(r"\s{2,}", " ", name).strip(" –-·|")


VENUE_WORD = re.compile(
    r"restaurant|stube|stüberl|stueberl|esszimmer|seminar|day\s*spa|\bspa\b|\bbar\b"
    r"|caf[eé]|ladestation|sauna|bistro|lounge|kaffee|wirtshaus|gasthaus|pool|kitchen"
    r"|küche|kueche|terrasse|eisstock",
    re.I,
)
VENUE_LEAD = (
    r"^(?:seminarraum|restaurant|day\s*spa|bar|caf[eé]|bistro|lounge|sauna|wellness|spa"
    r"|e-bike[- ]ladestation|ladestation|wirtshaus|pool)"
)
W = r"[\w'`´’‘.&-]"
HOTEL_PHRASE = (
    r"((?:" + W + r"+\s+){0,2}" + W + r"*(?:hotel|resort|lodge|residenz|chalet)" + W + r"*"
    r"(?:\s+" + W + r"+){0,4})$"
)
HOTEL_WORD = re.compile(r"hotel|resort|lodge|residenz|chalet|landhaus|pension|gasthof", re.I)
VENUE_SUFFIX = re.compile(
    r"apr[eè]s|\bski\b|\bbar\b|lounge|restaurant|terrasse|stube|kitchen|küche|day\s*spa",
    re.I,
)


def clean_name(name: str) -> str:
    """Trim editorial framing so the property name is what is left."""
    name = strip_stars(name)
    name = re.sub(r"^[3-5]-?sterne-?(?:superior-?)?(?=hotel)", "", name, flags=re.I)
    name = re.sub(
        r"^(?:wellness(?:\s*&\s*spa)?|day\s*spa|spa|urlaub|genuss)\s+(?:im|in|at|beim)\s+",
        "",
        name,
        flags=re.I,
    )
    # "Esszimmer am Bichl – Restaurant im Landhotel zum Bad" -> "Landhotel zum Bad"
    m = re.search(r"\b(?:im|in|des|der|by|at)\s+" + HOTEL_PHRASE, name, flags=re.I)
    if m and VENUE_WORD.search(name[: m.start()]):
        name = m.group(1)
    # "Seminarraum Hotel Tauernhof" -> "Hotel Tauernhof"
    m = re.match(VENUE_LEAD + r"\s+" + HOTEL_PHRASE, name, flags=re.I)
    if m:
        name = m.group(1)
    # "Sporthotel Alpin - Apres Ski" -> "Sporthotel Alpin"
    m = re.match(r"^(.{4,}?)\s+[-–|]\s+(.+)$", name)
    if m and HOTEL_WORD.search(m.group(1)) and not HOTEL_WORD.search(m.group(2)):
        if VENUE_SUFFIX.search(m.group(2)):
            name = m.group(1)
    return re.sub(r"\s{2,}", " ", name).strip(" –-·|")


def resort_of(town: str, postcode: str = "") -> tuple[str, str, str] | None:
    key = fold(town).strip(" .,")
    for k, v in sorted(CORRIDOR.items(), key=lambda kv: -len(kv[0])):
        fk = fold(k)
        if key == fk or key.startswith(fk + " ") or key.startswith(fk + "-"):
            return v
    if postcode in POSTCODES:
        return CORRIDOR[POSTCODES[postcode]]
    return None


TOWN_STOPLIST = {
    "aktivitaten",
    "aktivitaeten",
    "deutsch",
    "english",
    "italiano",
    "sommer",
    "winter",
    "gmbh",
    "werbung",
    "tourismus",
    "tourismusverband",
    "info",
    "kontakt",
    "impressum",
}


def address_of(html: str) -> tuple[str, str]:
    """(postcode, town) of the property: the FIRST plausible address line in document
    order, after nav/footer chrome is removed (headers stay: hero headers carry
    the address on some boards). The caller decides
    whether that town is in the corridor -- a page whose own address is outside it is
    dropped, instead of being rescued by a board's footer address."""
    tree = HTMLParser(html)
    for t in (
        "script",
        "style",
        "noscript",
        "nav",
        "footer",
        "[role=contentinfo]",
        "[role=navigation]",
        "[role=banner]",
    ):
        for node in tree.css(t):
            node.decompose()
    text = tree.text(separator="\n")
    for line in text.splitlines():
        for m in POSTCODE_ANY.finditer(line.strip()):
            pc, town = m.group(1), m.group(2).strip()
            if re.match(r"^(19|20)\d\d$", pc) or len(set(pc)) == 1:
                continue
            if len(pc) == 4 and pc < "1000":
                continue
            first = fold(town).split(" ")[0].strip(".,")
            if not first or first in TOWN_STOPLIST or not town[0].isupper():
                continue
            return pc, town
    return "", ""


def external_links(html: str, page_url: str) -> list[tuple[str, str]]:
    """(href, anchor text) for every link off the site that could be a property."""
    tree = HTMLParser(html)
    bh = host(page_url)
    out: list[tuple[str, str]] = []
    for a in tree.css("a[href]"):
        href = (a.attributes.get("href") or "").strip()
        if not href.startswith("http"):
            continue
        h = host(href)
        if not h or h == bh or h.endswith("." + bh) or any(s in h for s in SOCIAL):
            continue
        out.append((href.split("?")[0], a.text(strip=True) or ""))
    return out


VENDOR_DOMAIN = re.compile(
    r"architekt|untertrifaller|holzbox|agentur|agency|media|studio|design\b|webdesign|print"
    r"|photo|foto|bau\b|holzbau|tischlerei|zimmerei|consult|marketing|software|it-",
    re.I,
)


def external_site(html: str, page_url: str, name: str = "") -> str:
    """The property's own website on a single-property page, or ''. Never guessed.

    Preference: a link the page itself labels as the website, then a domain that
    shares a distinguishing token with the property name, then the first external
    link -- unless that looks like a vendor credit (architect, agency, builder)."""
    links = external_links(html, page_url)
    if not links:
        return ""
    for href, text in links:
        if re.search(r"website|webseite|homepage|zur seite|www\.|^http", text, re.I):
            return root(href)
    if name:
        hit = affine_site(name, links)
        if hit:
            return hit
    href = links[0][0]
    if VENDOR_DOMAIN.search(host(href)):
        return ""
    return root(href)


def affine_site(name: str, links: list[tuple[str, str]]) -> str:
    """On a multi-property page, a link only counts as the property's site when a
    distinguishing token of the name appears in the domain."""
    generic = {
        "hotel", "chalet", "resort", "spa", "lech", "arlberg", "seefeld", "tirol",
        "alpin", "alpen", "wellness", "gourmet", "burg", "romantik", "boutique",
    }  # fmt: skip
    tokens = [t for t in re.split(r"[^a-z0-9]+", fold(name)) if len(t) >= 4 and t not in generic]
    tokens += [
        t for t in re.split(r"[^a-z0-9]+", flat_ue(name)) if len(t) >= 4 and t not in generic
    ]
    for href, _ in links:
        h = flat(host(href))
        if any(t in h for t in tokens):
            return root(href)
    return ""


def category_for(name: str, hint: str = "") -> str:
    """PARTNERS category from the published name (and the website, for the
    apartment/holiday-flat signal that German 'Landhaus' listings hide in the domain)."""
    blob = fold(name) + " " + fold(hint)
    if re.search(
        r"apart|appartement|ferienwohnung|fewo|residenz|suites?\b", blob
    ) and not re.search(r"hotel", fold(name)):
        return "3. Apartments & Residences"
    if re.search(r"chalet|lodge|hütte|huette", fold(name)) and not re.search(r"hotel", fold(name)):
        return "2. Chalets & Chalet Management"
    return "1. Hotels"


def looks_like_property(name: str) -> bool:
    return (
        bool(name)
        and 4 <= len(name) <= 90
        and not NOT_A_PROPERTY.search(name)
        and bool(PROPERTY_WORD.search(name))
    )


@dataclass
class Candidate:
    name: str
    website: str
    town: str
    postcode: str
    stars: str
    source_url: str
    source_label: str
    snippet: str
    hint: str = ""
    resort: tuple[str, str, str] | None = None
    fetched_at: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))


# ---------------------------------------------------------------------------
# Sources
# ---------------------------------------------------------------------------


class Sources:
    def __init__(self, fetcher: Fetcher):
        self.f = fetcher
        self.log: list[str] = []

    async def sitemap(self, home: str, max_children: int = 8) -> list[str]:
        await self.f.ensure_robots(host(home))
        pol = self.f.robots.get(host(home))
        seeds = list(dict.fromkeys((pol.sitemaps if pol else []) + [urljoin(home, "sitemap.xml")]))
        urls, queue, n = [], seeds[:3], 0
        while queue and n < max_children:
            u = queue.pop(0)
            n += 1
            p = await self.f.get(u)
            if not p.ok:
                continue
            us, kids = parse_sitemap(p.body)
            urls += us
            queue += kids[:6]
        return urls

    async def detail(self, url: str) -> tuple[str, str, str, str, str] | None:
        """(h1, website, postcode, town, text) from a single-property page."""
        p = await self.f.get(url)
        if not p.ok or not p.body:
            return None
        tree = HTMLParser(p.body)
        h1 = tree.css_first("h1")
        name = (h1.text(strip=True) if h1 else "") or ""
        pc, town = address_of(p.body)
        site = external_site(p.body, p.final_url or url, name)
        for t in ("script", "style", "noscript"):
            for n in tree.css(t):
                n.decompose()
        return name, site, pc, town, tree.text(separator=" ")[:6000]

    # -- tirol.at: editorial listings + sitemap detail pages by corridor town -----
    async def tirol(self) -> list[Candidate]:
        pages = [
            "luxushotels",
            "best-of-wellnesshotels",
            "chalets",
            "design",
            "adults-only-wellnesshotels",
            "an-der-piste",
            "wellness-skifahren",
            "hotels-mit-pool",
            "romantische-unterkuenfte",
            "beheizte-hotelpools",
            "wellness-kulinarik",
        ]
        entries: dict[str, tuple[str, str]] = {}
        for slug in pages:
            url = f"https://www.tirol.at/unterkuenfte/{slug}"
            p = await self.f.get(url)
            if not p.ok:
                self.log.append(f"tirol.at/{slug}: {p.error or p.status}")
                continue
            tree = HTMLParser(p.body)
            for a in tree.css("a[href]"):
                href = urljoin(url, a.attributes.get("href") or "")
                if "/unterkuenfte/urlaub-buchen/" not in href:
                    continue
                href = href.split("#")[0].rstrip("/")
                entries.setdefault(href, (a.text(strip=True) or "", f"tirol.at/{slug}"))
        # Widen with every sitemap detail page whose slug names a corridor town.
        urls = await self.sitemap("https://www.tirol.at/")
        extra = [
            u.rstrip("/")
            for u in urls
            if "/unterkuenfte/urlaub-buchen/" in u
            and "/en/" not in u
            and any(t in u.lower() for t in TIROL_SLUG_TOWNS)
        ]
        for u in extra:
            entries.setdefault(u, ("", "tirol.at/sitemap"))
        self.log.append(f"tirol.at: {len(entries)} detail pages ({len(extra)} via sitemap)")
        out: list[Candidate] = []
        for href, (text, label) in entries.items():
            d = await self.detail(href)
            if not d:
                continue
            name, site, pc, town, body = d
            name = re.sub(r"^Unterkunft", "", name).strip() or text
            resort = resort_of(town, pc)
            if not resort or not looks_like_property(name):
                continue
            out.append(
                Candidate(
                    clean_name(name),
                    site,
                    town,
                    pc,
                    stars_of(name, text, body),
                    href,
                    label,
                    town,
                    resort=resort,
                )  # fmt: skip
            )
        return out

    # -- kitzbuehel.com: golfhotels list -> detail pages ------------------------
    async def kitzbuehel(self) -> list[Candidate]:
        url = "https://www.kitzbuehel.com/aktivitaeten/golf/golfhotels-in-kitzbuehel/"
        p = await self.f.get(url)
        if not p.ok:
            self.log.append(f"kitzbuehel.com: {p.error or p.status}")
            return []
        tree = HTMLParser(p.body)
        links: dict[str, str] = {}
        for hd in tree.css("h2, h3, h4"):
            t = hd.text(strip=True) or ""
            node, link = hd, ""
            for _ in range(5):
                if node is None:
                    break
                a = node.css_first("a[href]")
                if a is not None and "/unterkunft-buchen/details/" in (
                    a.attributes.get("href") or ""
                ):
                    link = urljoin(url, a.attributes["href"])
                    break
                node = node.parent
            if link and 3 < len(t) < 80:
                links.setdefault(link, t)
        out: list[Candidate] = []
        for link, heading in links.items():
            d = await self.detail(link)
            if not d:
                continue
            name, site, pc, town, body = d
            resort = resort_of(town, pc) or CORRIDOR["kitzbühel"]
            out.append(
                Candidate(
                    clean_name(name or heading),
                    site,
                    town,
                    pc,
                    stars_of(heading, body),
                    link,
                    "kitzbuehel.com",
                    heading,
                    resort=resort,
                )  # fmt: skip
            )
        return out

    # -- seefeld.com: editorial pages, link-driven -------------------------------
    async def seefeld(self) -> list[Candidate]:
        pages = {
            "5-sterne-wellness-hotels-region-seefeld.html": "5*",
            "4-sterne-superior-wellness-hotels-region-seefeld.html": "4*S",
            "4-sterne-wellness-hotels.html": "4*",
        }
        out: list[Candidate] = []
        for slug, page_stars in pages.items():
            url = f"https://www.seefeld.com/de/{slug}"
            p = await self.f.get(url)
            if not p.ok:
                self.log.append(f"seefeld.com/{slug}: {p.error or p.status}")
                continue
            best: dict[str, str] = {}
            for href, text in external_links(p.body, url):
                h = host(href)
                name = clean_name(text)
                if looks_like_property(name) and (h not in best or len(name) > len(best[h])):
                    best[h] = name
            for h, name in best.items():
                town = "leutasch" if "leutasch" in fold(name) else "seefeld in tirol"
                out.append(
                    Candidate(
                        name,
                        f"https://{h}",
                        town,
                        "",
                        stars_of(name) or page_stars,
                        url,
                        "seefeld.com",
                        name,
                        resort=CORRIDOR[town],
                    )  # fmt: skip
                )
        return out

    # -- zellamsee-kaprun.com: golfhotels cards ---------------------------------
    async def zell(self) -> list[Candidate]:
        url = "https://www.zellamsee-kaprun.com/golfhotels"
        p = await self.f.get(url)
        if not p.ok:
            self.log.append(f"zellamsee-kaprun.com: {p.error or p.status}")
            return []
        out: list[Candidate] = []
        seen: set[str] = set()
        for href, text in external_links(p.body, url):
            h = host(href)
            if h in seen or not re.search(r"hotel|resort|spa|chalet", text, re.I):
                continue
            m = re.match(r"^(.*?\*{3,5}\s?[sS]?)", text)
            head = m.group(1) if m else text[:80]
            name = clean_name(head)
            if not looks_like_property(name):
                continue
            seen.add(h)
            town = "kaprun" if "kaprun" in fold(text[:160]) else "zell am see"
            out.append(
                Candidate(
                    name.title() if name.isupper() else name,
                    root(href),
                    town,
                    "",
                    stars_of(head),
                    url,
                    "zellamsee-kaprun.com",
                    text[:120],
                    resort=CORRIDOR[town],
                )  # fmt: skip
            )
        return out

    # -- zugspitz-region.de: gastro pages of hotels (GaPa / Grainau / Krün) ------
    async def zell_poi(self) -> list[Candidate]:
        """POI fiches (day spa, restaurant, seminar room) hosted by a hotel: static pages
        carrying the hotel's own address and a labelled website link. Stars only from
        the heading, never from surrounding page text."""
        urls = await self.sitemap("https://www.zellamsee-kaprun.com/")
        pois = [
            u
            for u in urls
            if "/de/poi/" in u and re.search(r"hotel|resort|residenz|chalet|lodge", u)
        ]
        self.log.append(f"zellamsee-kaprun.com: {len(pois)} hotel-hosted POI pages")
        out: list[Candidate] = []
        for u in pois[:120]:
            d = await self.detail(u)
            if not d:
                continue
            h1, site, pc, town, body = d
            name = clean_name(h1)
            resort = resort_of(town, pc)
            if not resort or not site or not looks_like_property(name):
                continue
            if not re.search(r"hotel|resort|residenz|chalet|lodge", name, re.I):
                continue
            out.append(
                Candidate(
                    name,
                    site,
                    town,
                    pc,
                    stars_of(h1),
                    u,
                    "zellamsee-kaprun.com/poi",
                    body[:120],
                    resort=resort,
                )  # fmt: skip
            )
        return out

    async def zugspitz(self) -> list[Candidate]:
        urls = await self.sitemap("https://www.zugspitz-region.de/")
        slugs = [
            u
            for u in urls
            if "/gastro/" in u
            and "/en/" not in u
            and re.search(r"hotel|resort|chalet|lodge|residenz", u)
            and not re.search(r"gasthof|berggasthof|alm\b|huette|hütte", u)
        ]
        self.log.append(f"zugspitz-region.de: {len(slugs)} hotel-ish gastro pages")
        out: list[Candidate] = []
        for u in slugs[:60]:
            d = await self.detail(u)
            if not d:
                continue
            name, site, pc, town, body = d
            resort = resort_of(town, pc)
            if not resort or not site:
                continue
            name = clean_name(name)
            if re.match(r"^restaurant\s+", name, re.I):
                name = re.sub(r"^restaurant\s+", "", name, flags=re.I)
            if not looks_like_property(name):
                continue
            out.append(
                Candidate(
                    name,
                    site,
                    town,
                    pc,
                    stars_of(body),
                    u,
                    "zugspitz-region.de",
                    d[0],
                    resort=resort,
                )  # fmt: skip
            )
        return out

    # -- alpenwelt-karwendel.de: accommodation fiches (Krün / Elmau / Klais) -----
    async def alpenwelt(self) -> list[Candidate]:
        urls = await self.sitemap("https://www.alpenwelt-karwendel.de/")
        fichas = [
            u
            for u in urls
            if re.search(r"/a-[a-z0-9\-]+$", u)
            and "/en/" not in u
            and re.search(r"hotel|chalet|resort|lodge|residenz|villa|suite|landhaus", u)
        ]
        self.log.append(f"alpenwelt-karwendel.de: {len(fichas)} hotel-ish fiches")
        out: list[Candidate] = []
        for u in fichas[:150]:
            d = await self.detail(u)
            if not d:
                continue
            name, site, pc, town, body = d
            resort = resort_of(town, pc)
            if not resort or not site or not looks_like_property(name):
                continue
            out.append(
                Candidate(
                    clean_name(name),
                    site,
                    town,
                    pc,
                    stars_of(body),
                    u,
                    "alpenwelt-karwendel.de",
                    body[:120],
                    resort=resort,
                )  # fmt: skip
            )
        return out

    # -- Lech Zürs: editorial hotel pages (name<->domain affinity) + vorarlberg ---
    async def lech(self) -> list[Candidate]:
        out: list[Candidate] = []
        urls = await self.sitemap("https://www.lechzuers.com/")
        editorial = [
            u
            for u in urls
            if "/de/" in u
            and "/buchen/" not in u
            and "/presse/" not in u
            and re.search(r"-im-hotel-|-hotel-|im-burg-hotel|chalet|-hotel$", u)
        ]
        self.log.append(f"lechzuers.com: {len(editorial)} editorial hotel pages")
        for u in editorial[:40]:
            p = await self.f.get(u)
            if not p.ok:
                continue
            tree = HTMLParser(p.body)
            h1 = tree.css_first("h1")
            title = (h1.text(strip=True) if h1 else "") or ""
            m = re.search(
                r"(?:im|in|at|beim)\s+((?:Burg\s+|Romantik\s+|Boutique\s+)?Hotel[^,.–|]{2,50}"
                r"|Chalet[^,.–|]{2,50})",
                title,
            )
            name = m.group(1).strip() if m else ""
            if not name and re.match(r"^(Hotel|Chalet|Burg Hotel)\b", title):
                name = title.split(",")[0].strip()
            if not looks_like_property(name):
                continue
            site = affine_site(name, external_links(p.body, u))
            town = "zürs" if re.search(r"z[üu]e?rs", fold(title)) else "lech am arlberg"
            out.append(
                Candidate(
                    clean_name(name),
                    site,
                    town,
                    "",
                    stars_of(title),
                    u,
                    "lechzuers.com",
                    title,
                    resort=CORRIDOR[town],
                )  # fmt: skip
            )
        vurls = await self.sitemap("https://www.vorarlberg.travel/")
        hotel_pages = [u for u in vurls if "/hotel/" in u and "/en/" not in u]
        self.log.append(f"vorarlberg.travel: {len(hotel_pages)} /hotel/ pages")
        for u in hotel_pages[:150]:
            d = await self.detail(u)
            if not d:
                continue
            name, site, pc, town, body = d
            resort = resort_of(town, pc)
            if not resort or resort[1] != "Arlberg" or not looks_like_property(name):
                continue
            out.append(
                Candidate(
                    clean_name(name),
                    site,
                    town,
                    pc,
                    stars_of(name, body),
                    u,
                    "vorarlberg.travel",
                    body[:120],
                    resort=resort,
                )  # fmt: skip
            )
        return out


# ---------------------------------------------------------------------------
# Assembly
# ---------------------------------------------------------------------------


async def discover(cfg) -> tuple[list[Candidate], list[str]]:
    cache = PageCache(cfg.cache_directory, enabled=True)
    cfg.cache_directory.mkdir(parents=True, exist_ok=True)
    async with Fetcher(cfg.fetch, cache) as f:
        s = Sources(f)
        groups = await asyncio.gather(
            s.tirol(),
            s.kitzbuehel(),
            s.seefeld(),
            s.zell(),
            s.zell_poi(),
            s.zugspitz(),
            s.alpenwelt(),
            s.lech(),
        )
    return [c for g in groups for c in g], s.log


def dedupe(cands: list[Candidate], view, cfg):
    """Drop candidates already in PARTNERS (by domain or folded name) and internal dupes.
    Among internal dupes the best-evidenced one (stars known, website known) wins."""
    spec = cfg.workbook
    stop = {fold(w) for w in cfg.dedupe.name_stopwords}
    existing_domains, existing_names = set(), {}
    for pr in view.rows:
        d = domain_of(pr.get(spec.column_for("website_url")))
        if d:
            existing_domains.add(d)
        n = pr.get(spec.column_for("entity_name"))
        existing_names[name_tokens(n, stop)] = n
    ordered = sorted(
        cands,
        key=lambda c: (c.stars == "", c.website == "", not HOTEL_WORD.search(c.name), -len(c.name)),
    )
    kept, dropped, seen_dom, seen_name = [], [], set(), set()
    seen_pairs: list[tuple[set[str], object]] = []
    for c in ordered:
        d = domain_of(c.website) if c.website else ""
        toks = name_tokens(c.name, stop)
        if d and d in existing_domains:
            dropped.append((c, f"ya en PARTNERS (dominio {d})"))
        elif toks in existing_names:
            dropped.append((c, f"ya en PARTNERS ({existing_names[toks]})"))
        elif d and d in seen_dom:
            dropped.append((c, "duplicado interno (dominio)"))
        elif toks in seen_name:
            dropped.append((c, "duplicado interno (nombre)"))
        elif toks and any(
            r == c.resort and (set(toks) <= t or t <= set(toks)) for t, r in seen_pairs
        ):
            dropped.append((c, "duplicado interno (nombre contenido, mismo resort)"))
        else:
            if d:
                seen_dom.add(d)
            seen_name.add(toks)
            seen_pairs.append((set(toks), c.resort))
            kept.append(c)
    return kept, dropped


def tier_for(stars: str, vocabulary: set[str]) -> str:
    """Only from published stars, and only with a value the sheet already uses."""
    want = {"5*S": "luxury", "5*": "luxury", "4*S": "premium", "4*": "mid-premium"}.get(stars, "")
    if want and (not vocabulary or want in vocabulary):
        return want
    return "TBD"


def main() -> int:
    # Windows consoles default to cp1252; hotel names carry umlauts.
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    cfg = config.load()
    view = load_workbook_view(cfg, record_state=False, command="discover_hotels")
    spec = cfg.workbook
    import openpyxl

    wb = openpyxl.load_workbook(view.path, read_only=True)
    ws = wb[spec.sheet]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()
    tier_vocab = {
        str(pr.get(spec.column_for("segment_tier"))).strip().lower()
        for pr in view.rows
        if pr.get(spec.column_for("segment_tier")).strip() not in ("", "TBD")
    }

    print(f"PARTNERS: {len(view.rows)} filas | tier vocabulary: {sorted(tier_vocab)}")
    cands, log = asyncio.run(discover(cfg))
    print("\n".join("  " + line for line in log))
    print(f"\ncandidatos crudos: {len(cands)}")

    kept, dropped = dedupe(cands, view, cfg)
    by_resort: dict[str, list[Candidate]] = defaultdict(list)
    for c in kept:
        by_resort[c.resort[0]].append(c)
    final: list[Candidate] = []
    capped: list[tuple[Candidate, str]] = []
    for resort, cs in by_resort.items():
        final += cs[:PER_RESORT_CAP]
        capped += [
            (c, f"tope por resort ({PER_RESORT_CAP}) en {resort}") for c in cs[PER_RESORT_CAP:]
        ]
    final.sort(key=lambda c: (c.resort[2], c.resort[1], c.resort[0], c.stars == "", c.name))

    today = datetime.now().date().isoformat()
    out_dir = cfg.artifacts_directory
    out_dir.mkdir(parents=True, exist_ok=True)
    main_path = out_dir / f"{today}_hotel_candidates.csv"
    detail_path = out_dir / f"{today}_hotel_candidates_detail.csv"

    rows, details = [], []
    next_id = FIRST_ID
    for c in final:
        rid = f"EFE-{next_id:04d}"
        next_id += 1
        resort_base, region, country = c.resort
        row = {h: "TBD" for h in hdr}
        row.update(
            {
                "ID": rid,
                "Entity_Name": c.name,
                "Category": category_for(c.name, c.website),
                "Subcategory": "TBD",
                "Resort_Base": resort_base,
                "Region_Valley": region,
                "Country": country,
                "Website_URL": c.website or "TBD",
                "Segment_Tier": tier_for(c.stars, tier_vocab),
                "Star_Rating_or_Class": c.stars or "TBD",
                "B2B_Program_Exists": "Unknown",
                "Priority_Score": "",
                "Contacted": "NO",
                "Contact_Date": "",
                "Follow_Up_Days": 14,
                "Next_Follow_Up": "",
                "Email_Sent": "",
                "Call_Made": "",
                "WhatsApp_Sent": "",
                "Meeting_Booked": "",
                "Agreement_Signed": "",
                "Status": "Not started",
                "Next_Action": "TBD",
                "Source_URL": c.source_url,
                "Date_Verified": today,
                "Round": ROUND,
                "Strategic_Fit_Note": (
                    f"[CANDIDATO] {c.source_label} — "
                    f"{c.stars or 'sin estrellas publicadas'}"
                    + ("" if c.website else " — [SIN WEB]")
                ),
            }
        )
        rows.append([row[h] for h in hdr])
        details.append(
            [
                rid,
                c.name,
                resort_base,
                country,
                c.stars,
                c.website,
                c.source_label,
                c.source_url,
                c.postcode,
                c.town,
                c.fetched_at,
                c.snippet[:160],
            ]  # fmt: skip
        )

    with open(main_path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(hdr)
        w.writerows(rows)
    with open(detail_path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(
            [
                "ID",
                "Entity_Name",
                "Resort_Base",
                "Country",
                "Stars",
                "Website_URL",
                "Source",
                "Source_URL",
                "Postcode",
                "Town_as_published",
                "Fetched_At",
                "Snippet",
            ]  # fmt: skip
        )
        w.writerows(details)
        w.writerow([])
        w.writerow(["-- descartados --"])
        for c, why in dropped + capped:
            w.writerow(
                [
                    "",
                    c.name,
                    c.resort[0] if c.resort else "",
                    "",
                    c.stars,
                    c.website,
                    c.source_label,
                    c.source_url,
                    c.postcode,
                    c.town,
                    c.fetched_at,
                    why,
                ]  # fmt: skip
            )

    print(
        f"\ncandidatos finales: {len(final)}  "
        f"(descartados: ya en PARTNERS/dupes {len(dropped)}, tope {len(capped)})"
    )
    print("por resort:", dict(Counter(c.resort[0] for c in final)))
    print("por estrellas:", dict(Counter(c.stars or "n/d" for c in final)))
    print("por fuente:", dict(Counter(c.source_label for c in final)))
    print("sin web:", sum(1 for c in final if not c.website))
    print(f"IDs: EFE-{FIRST_ID:04d}..EFE-{next_id - 1:04d}")
    print(f"\n{main_path}\n{detail_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
