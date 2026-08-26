"""Promotion: candidate rows become PARTNERS rows.

Discovery emits candidates as a CSV shaped like PARTNERS (`scripts/discover_hotels.py`).
Promotion appends the ones that are not already in the sheet -- by ID, by domain, by
normalised name -- as new data rows after the last real row, gives each the live
`Next_Follow_Up` formula the sheet expects on every data row, records the event in
CHANGELOG / CHANGELOG_DETAIL, and emits the next version through the same fidelity
gate as enrichment.

No existing cell is ever touched: every write lands on a row above the last data row
of the input, and those rows are proven empty across every column before the first
write -- Google Sheets pads the used range with hundreds of blank rows, and a note
typed on one of them must stop the run, not vanish under a new row.

What promotion cannot make true, it says: the DASHBOARD's COUNTIF totals carry the
cached results of the input version until Sheets or Excel recompute them on open,
and new rows beyond the DASHBOARD's ranges are not counted at all.
"""

from __future__ import annotations

import csv
import logging
import re
import shutil
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries

from efe.config import Config
from efe.models import CellChange, Confidence, DataClass, VerificationError, today_iso
from efe.workbook.reader import (
    WorkbookView,
    domain_of,
    file_fingerprint,
    guard_readable,
    guard_writable,
)
from efe.workbook.resolve import parse_version
from efe.workbook.verify import compare, format_report, snapshot
from efe.workbook.writer import (
    _append_changelog,
    _write_changelog_detail,
    assert_version_free,
    deliver,
    next_version_path,
)
from efe.workbook.xmlutil import read_cached_values, reinject_cached_values

log = logging.getLogger(__name__)

ID_PATTERN = re.compile(r"^EFE-\d{4}$")
#: Columns the sheet stores as numbers. Everything else stays text, so phone
#: numbers, postcodes and codes keep their leading zeros.
NUMERIC_COLUMNS = ("Follow_Up_Days", "Priority_Score")
#: Second-level labels under which a registrable domain has three labels.
_SECOND_LEVEL = {"co", "com", "net", "org", "ac", "gov", "edu", "or", "ne"}
_RANGE_RE = re.compile(r"PARTNERS!\$?[A-Z]+\$?(\d+):\$?[A-Z]+\$?(\d+)")


@dataclass
class PromotionPlan:
    """What would be appended, what would not, and why."""

    source: Path
    first_row: int
    #: (worksheet row, header name -> value) for every accepted candidate, in order.
    accepted: list[tuple[int, dict[str, str]]] = field(default_factory=list)
    #: (ID, Entity_Name, reason) for every candidate left out.
    rejected: list[tuple[str, str, str]] = field(default_factory=list)
    #: Facts that do not block but the human should know.
    notices: list[str] = field(default_factory=list)

    @property
    def ids(self) -> list[str]:
        return [row.get("ID", "") for _, row in self.accepted]

    @property
    def last_row(self) -> int:
        return self.first_row + len(self.accepted) - 1

    def describe(self) -> str:
        if self.accepted:
            head = (
                f"promotion from {self.source.name}: {len(self.accepted)} row(s) to append "
                f"at worksheet rows {self.first_row}..{self.last_row}, "
                f"{len(self.rejected)} left out"
            )
        else:
            head = (
                f"promotion from {self.source.name}: no rows to append, "
                f"{len(self.rejected)} left out"
            )
        lines = [head]
        for entity_id, name, reason in self.rejected:
            lines.append(f"  skipped : {entity_id} {name} -- {reason}")
        for note in self.notices:
            lines.append(f"  note    : {note}")
        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------


def read_candidates(path: Path, cfg: Config) -> list[dict[str, str]]:
    """Rows of a PARTNERS-shaped CSV, keyed by header name.

    Strict on shape: every column must be a PARTNERS column, no column twice, every
    data row exactly as wide as the header (an unquoted comma would otherwise shift
    a whole row sideways in silence). PARTNERS columns the CSV lacks -- a column
    added to the sheet after the CSV was made -- are filled blank.
    """
    spec = cfg.workbook
    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError as exc:
        raise VerificationError(
            f"{path.name} is not UTF-8 ({exc.reason} at byte {exc.start}); save it as "
            "UTF-8 CSV and re-run. Nothing has been written."
        ) from exc
    import io

    with io.StringIO(text, newline="") as fh:
        reader = csv.reader(fh)
        try:
            columns = [c.strip() for c in next(reader)]
        except StopIteration:
            raise VerificationError(f"{path.name} is empty. Nothing has been written.") from None
        records: list[dict[str, str]] = []
        for line_no, record in enumerate(reader, start=2):
            if not any(cell.strip() for cell in record):
                continue
            if len(record) != len(columns):
                raise VerificationError(
                    f"{path.name} line {line_no}: {len(record)} fields, the header has "
                    f"{len(columns)}. Fix the CSV. Nothing has been written."
                )
            records.append(dict(zip(columns, record, strict=True)))
    repeated = sorted(c for c, k in Counter(columns).items() if k > 1)
    if repeated:
        raise VerificationError(
            f"{path.name} names a column twice: {repeated}. Nothing has been written."
        )
    unknown = [c for c in columns if c not in spec.header]
    if unknown:
        raise VerificationError(
            f"{path.name} has columns that are not PARTNERS columns: {unknown}. "
            "Nothing has been written."
        )
    for required in ("ID", "Entity_Name"):
        if required not in columns:
            raise VerificationError(
                f"{path.name} lacks the {required!r} column. Nothing has been written."
            )
    return [{name: (row.get(name) or "").strip() for name in spec.header} for row in records]


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------


def _normalise(text: str) -> str:
    text = text.lower().replace("ä", "ae").replace("ö", "oe").replace("ü", "ue")
    text = text.replace("ß", "ss")
    text = re.sub(r"\([^)]*\)", " ", text)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def name_keys(name: str) -> set[str]:
    """Comparison keys for an entity name: the whole name and the part before a
    dash or pipe, lowercased, umlauts expanded, accents stripped, parentheticals and
    punctuation dropped. `Schlosshotel Kitzbühel – THE SPA MOMENT` and
    `Schlosshotel Kitzbuehel (ex A-ROSA)` share the key `schlosshotel kitzbuehel`."""
    keys: set[str] = set()
    full = _normalise(name)
    if full:
        keys.add(full)
    head = _normalise(re.split(r"\s+[–—|-]\s+", name, maxsplit=1)[0])
    if head:
        keys.add(head)
    return keys


def registrable(domain: str) -> str:
    """`de.kristiania.at` -> `kristiania.at`; `x.example.co.uk` -> `example.co.uk`."""
    labels = domain.split(".")
    if len(labels) <= 2:
        return domain
    if labels[-2] in _SECOND_LEVEL:
        return ".".join(labels[-3:])
    return ".".join(labels[-2:])


def _resort_key(text: str) -> str:
    return _normalise(text)


# ---------------------------------------------------------------------------
# Planning
# ---------------------------------------------------------------------------


def inspect_ranges(path: Path, cfg: Config) -> dict[str, int | None]:
    """Where the sheet's range-bound structures end: the PARTNERS autofilter, its
    data validations, and the highest PARTNERS row any other sheet's formula reads."""
    spec = cfg.workbook
    wb = load_workbook(path, data_only=False)
    try:
        ws = wb[spec.sheet]
        autofilter = None
        if ws.auto_filter.ref:
            _, _, _, autofilter = range_boundaries(ws.auto_filter.ref.replace("$", ""))
        validation = None
        for dv in ws.data_validations.dataValidation:
            for rng in str(dv.sqref).split():
                _, _, _, bottom = range_boundaries(rng.replace("$", ""))
                validation = bottom if validation is None else max(validation, bottom)
        formula_reach = None
        for other in wb.worksheets:
            if other.title == spec.sheet:
                continue
            for row in other.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        for _, end in _RANGE_RE.findall(cell.value):
                            reach = int(end)
                            formula_reach = (
                                reach if formula_reach is None else max(formula_reach, reach)
                            )
        return {
            "autofilter_last_row": autofilter,
            "validation_last_row": validation,
            "formula_reach_last_row": formula_reach,
        }
    finally:
        wb.close()


def plan_promotion(
    view: WorkbookView,
    cfg: Config,
    rows: list[dict[str, str]],
    source: Path,
    ranges: dict[str, int | None] | None = None,
) -> PromotionPlan:
    """Decide which candidates may become rows. Detection only: nothing is merged.

    A candidate is left out when its ID is malformed, not above the sheet's highest
    ID, or already used; when its domain (exact or registrable) is already in the
    sheet; when its normalised name matches a row -- unless it clearly is another
    property (its own domain AND another resort), which is kept and noted; and when
    it repeats an earlier candidate.
    """
    spec = cfg.workbook
    id_col = spec.column_for("id")
    name_col = spec.column_for("entity_name")
    web_col = spec.column_for("website_url")
    resort_col = spec.column_for("resort_base")

    existing_ids: set[str] = set()
    highest = 0
    exact_domains: dict[str, int] = {}
    reg_domains: dict[str, int] = {}
    names: dict[str, int] = {}
    row_domain: dict[int, str] = {}
    row_resort: dict[int, str] = {}
    for pr in view.rows:
        rid = pr.get(id_col)
        existing_ids.add(rid)
        if ID_PATTERN.match(rid):
            highest = max(highest, int(rid[4:]))
        d = domain_of(pr.get(web_col))
        row_domain[pr.row] = d
        row_resort[pr.row] = _resort_key(pr.get(resort_col))
        if d:
            exact_domains.setdefault(d, pr.row)
            reg_domains.setdefault(registrable(d), pr.row)
        for key in name_keys(pr.get(name_col)):
            names.setdefault(key, pr.row)

    last_row = view.schema.last_row if view.schema else (view.rows[-1].row if view.rows else 1)
    plan = PromotionPlan(source=source, first_row=last_row + 1)
    seen_ids: set[str] = set()
    seen_domains: set[str] = set()
    seen_names: set[str] = set()
    next_row = last_row + 1
    for row in rows:
        entity_id, name = row.get("ID", ""), row.get("Entity_Name", "")
        domain = domain_of(row.get("Website_URL", ""))
        keys = name_keys(name)
        reason = ""
        matched = next((names[k] for k in keys if k in names), None)
        if not entity_id or not name:
            reason = "ID or Entity_Name missing"
        elif not ID_PATTERN.match(entity_id):
            reason = f"ID {entity_id!r} is not of the form EFE-dddd"
        elif int(entity_id[4:]) <= highest:
            reason = f"ID {entity_id} is not above the sheet's highest ID EFE-{highest:04d}"
        elif entity_id in existing_ids:
            reason = f"ID {entity_id} already in PARTNERS"
        elif entity_id in seen_ids:
            reason = f"ID {entity_id} repeated in the CSV"
        elif domain and domain in exact_domains:
            reason = f"domain {domain} already in PARTNERS (row {exact_domains[domain]})"
        elif domain and registrable(domain) in reg_domains:
            reason = (
                f"domain {domain} is under {registrable(domain)}, already in PARTNERS "
                f"(row {reg_domains[registrable(domain)]})"
            )
        elif matched is not None and (
            not domain
            or domain == row_domain.get(matched)
            or _resort_key(row.get("Resort_Base", "")) == row_resort.get(matched)
        ):
            reason = f"name already in PARTNERS (row {matched})"
        elif domain and domain in seen_domains:
            reason = f"domain {domain} repeated in the CSV"
        elif keys & seen_names:
            reason = "name repeated in the CSV"
        if reason:
            plan.rejected.append((entity_id, name, reason))
            continue
        if matched is not None:
            plan.notices.append(
                f"{entity_id} {name}: name matches row {matched} but has its own site "
                f"({domain}) and another resort -- kept as a different property"
            )
        seen_ids.add(entity_id)
        if domain:
            seen_domains.add(domain)
        seen_names |= keys
        plan.accepted.append((next_row, row))
        next_row += 1

    if plan.accepted and ranges:
        last = plan.last_row
        reach = ranges.get("formula_reach_last_row")
        if reach is not None and last > reach:
            plan.notices.append(
                f"the DASHBOARD's formulas read PARTNERS only down to row {reach}; rows "
                f"{max(plan.first_row, reach + 1)}..{last} would NOT be counted until "
                "the ranges are extended in Sheets"
            )
        if reach is not None:
            plan.notices.append(
                "the DASHBOARD's cached totals in the output reflect the input version "
                "until Sheets or Excel recompute them on open"
            )
        af = ranges.get("autofilter_last_row")
        if af is not None and last > af:
            plan.notices.append(
                f"the PARTNERS autofilter ends at row {af}; the new rows sit outside it "
                "(as do all rows added since)"
            )
        dv = ranges.get("validation_last_row")
        if dv is not None and last > dv:
            plan.notices.append(
                f"the PARTNERS data validations (dropdowns) end at row {dv}; the new "
                "rows get none (as do all rows added since)"
            )
    return plan


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------


def _cell_value(name: str, text: str):
    """CSV text -> cell value. Blank stays empty; only the sheet's numeric columns
    become numbers, and only for plain ASCII integers without a leading zero."""
    if text == "":
        return None
    if (
        name in NUMERIC_COLUMNS
        and text.isascii()
        and text.isdecimal()
        and not (len(text) > 1 and text.startswith("0"))
    ):
        return int(text)
    return text


def _formula_templates(ws, spec, last_row: int) -> dict[str, tuple[str, str, int, int]]:
    """Formula column -> (origin, canonical formula, rows using it, rows using another).

    The column's DOMINANT pattern, not whatever the nearest row holds: every formula
    is translated to the first data row and the most common form wins, so a single
    hand-edited cell cannot seed every new row."""
    out: dict[str, tuple[str, str, int, int]] = {}
    for name in spec.formula_columns:
        letter = spec.letter_of(name)
        col = column_index_from_string(letter)
        origin = f"{letter}{spec.first_data_row}"
        forms: Counter[str] = Counter()
        for r in range(spec.first_data_row, last_row + 1):
            value = ws.cell(r, col).value
            if isinstance(value, str) and value.startswith("="):
                forms[Translator(value, origin=f"{letter}{r}").translate_formula(origin)] += 1
        if not forms:
            raise VerificationError(
                f"no data row holds a formula in {name} ({letter}); cannot derive the "
                "formula for the new rows. Nothing has been written."
            )
        canonical, count = forms.most_common(1)[0]
        out[name] = (origin, canonical, count, sum(forms.values()) - count)
    return out


def _translated(template: tuple[str, str, int, int], letter: str, row_number: int, spec) -> str:
    origin, canonical, _, _ = template
    formula = Translator(canonical, origin=origin).translate_formula(f"{letter}{row_number}")
    anchor = str(spec.first_data_row)
    if re.search(rf"[A-Z]\$?{anchor}(?!\d)", canonical) and not re.search(
        rf"[A-Z]{row_number}(?!\d)", formula
    ):
        raise VerificationError(
            f"{letter}{row_number}: the translated formula {formula!r} does not reference "
            "its own row. Nothing has been written."
        )
    return formula


def _assert_target_rows_empty(ws, spec, first_row: int, last_row: int) -> None:
    """The rows about to be written must hold nothing, in any column. A value typed
    on a padded row is data, and it must stop the run rather than disappear."""
    stray: list[str] = []
    width = len(spec.header)
    for r in range(first_row, last_row + 1):
        for c in range(1, width + 1):
            value = ws.cell(r, c).value
            if value not in (None, ""):
                stray.append(f"{get_column_letter(c)}{r}={value!r}")
    if stray:
        raise VerificationError(
            f"rows {first_row}..{last_row} are not empty; promotion would overwrite "
            f"{len(stray)} cell(s), e.g. {stray[:8]}. Move or clear them in Sheets and "
            "re-run. Nothing has been written."
        )


def write_promoted(
    cfg: Config,
    view: WorkbookView,
    plan: PromotionPlan,
    *,
    run_id: str,
    output_path: Path | None = None,
    workdir: Path | None = None,
) -> Path:
    """Append the accepted rows and emit the next version. Returns the written path.

    Raises:
        VerificationError: the target rows are not empty, the fidelity gate failed,
            or there is nothing to promote.
        VersionConflictError: the next version already exists somewhere.
    """
    if not plan.accepted:
        raise VerificationError(
            "nothing to promote: every candidate was left out. Nothing has been written."
        )
    spec = cfg.workbook
    source = view.path
    guard_readable(source, spec.min_plausible_bytes)
    if view.fingerprint and file_fingerprint(source) != view.fingerprint:
        raise VerificationError(
            f"The input workbook changed since it was read:\n  {source}\n"
            "Re-run from the current file. Nothing has been written."
        )
    last_row = view.schema.last_row if view.schema else view.rows[-1].row
    if plan.first_row <= last_row:
        raise VerificationError(
            f"promotion would start at row {plan.first_row}, but data runs to row "
            f"{last_row}. Nothing has been written."
        )

    destination = output_path or next_version_path(cfg, view.version)
    target_version = parse_version(destination, cfg.output_basename) or view.version + 1
    assert_version_free(cfg, target_version)
    guard_writable(destination)

    staging_dir = workdir or (cfg.state_directory / "staging")
    staging_dir.mkdir(parents=True, exist_ok=True)
    candidate = staging_dir / f"{run_id}_{destination.name}"
    if candidate.exists():
        candidate.unlink()
    shutil.copy2(source, candidate)

    allowed: set[tuple[str, str]] = set()
    sheet_name = spec.sheet
    detail_sheet = spec.changelog_detail_sheet
    run_date = today_iso()
    version_label = destination.stem.rsplit("_", 1)[-1]
    detail_created = False
    now = datetime.now()

    try:
        wb = load_workbook(candidate, data_only=False)
        try:
            ws = wb[sheet_name]
            _assert_target_rows_empty(ws, spec, plan.first_row, plan.last_row)
            templates = _formula_templates(ws, spec, last_row)
            audit: list[CellChange] = []
            for row_number, values in plan.accepted:
                for name in spec.header:
                    if name in templates:
                        continue
                    value = _cell_value(name, values.get(name, ""))
                    if value is None:
                        continue
                    if isinstance(value, str) and value.lstrip().startswith("="):
                        raise VerificationError(
                            f"{values.get('ID')} {name}: value {value!r} would be stored as a "
                            "formula. Nothing has been written."
                        )
                    cell = ws.cell(row_number, column_index_from_string(spec.letter_of(name)))
                    if cell.value not in (None, ""):  # proven empty above; belt and braces
                        raise VerificationError(
                            f"{cell.coordinate} is not empty. Nothing has been written."
                        )
                    cell.value = value
                    allowed.add((sheet_name, cell.coordinate))
                for name, template in templates.items():
                    letter = spec.letter_of(name)
                    cell = ws.cell(row_number, column_index_from_string(letter))
                    cell.value = _translated(template, letter, row_number, spec)
                    allowed.add((sheet_name, cell.coordinate))
                audit.append(
                    CellChange(
                        row=row_number,
                        column=spec.column_for("id"),
                        field="row",
                        entity_id=values.get("ID", ""),
                        entity_name=values.get("Entity_Name", ""),
                        old_value="",
                        new_value=f"row promoted from {plan.source.name}",
                        confidence=Confidence.HIGH,
                        data_class=DataClass.CORPORATE_ROLE,
                        source_url=values.get("Source_URL", ""),
                        fetched_at=now,
                        extractor="promote",
                        note=values.get("Strategic_Fit_Note", ""),
                    )
                )

            ids = plan.ids
            formula_note = "; ".join(
                f"{name} formula from the column's dominant pattern ({count} rows"
                + (f", {others} other form(s) ignored" if others else "")
                + ")"
                for name, (_, _, count, others) in templates.items()
            )
            message = (
                f"Promotion (efe promote, run {run_id}): {len(plan.accepted)} candidate rows "
                f"appended from {plan.source.name} ({ids[0]}..{ids[-1]}) at rows "
                f"{plan.first_row}..{plan.last_row}; {len(plan.rejected)} left out as already "
                f"present or repeated. {formula_note}. Contacts stay TBD for the enricher; "
                "no existing cell was modified. DASHBOARD cached totals reflect the input "
                "version until recomputed on open."
            )
            for coordinate in _append_changelog(
                wb[spec.changelog_sheet], version_label, run_date, message, author="efe promote"
            ):
                allowed.add((spec.changelog_sheet, coordinate))
            coordinates, detail_created = _write_changelog_detail(wb, detail_sheet, audit, run_id)
            for coordinate in coordinates:
                allowed.add((detail_sheet, coordinate))
            wb.save(candidate)
        finally:
            wb.close()

        repaired = reinject_cached_values(candidate, source)
        problems = compare(
            snapshot(source),
            snapshot(candidate),
            allowed_value_changes=allowed,
            allowed_new_sheets={detail_sheet} if detail_created else set(),
            allowed_autofilter_changes={detail_sheet},
            require_cached_values=True,
        )
        if problems:
            raise VerificationError(
                format_report(problems)
                + f"\n\nThe candidate output was DELETED. {source.name} is untouched."
            )
        source_cached = sum(1 for v in read_cached_values(source).values() if v is not None)
        if repaired == 0 and source_cached:
            raise VerificationError(
                "No cached formula results were reinjected although the input carries "
                f"{source_cached}. The candidate output was DELETED."
            )

        assert_version_free(cfg, target_version)
        guard_writable(destination)
        deliver(candidate, destination)
        return destination
    finally:
        if candidate.exists():
            candidate.unlink()
