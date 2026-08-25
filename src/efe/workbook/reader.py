"""Guarded reading of the partner workbook.

Which file to read is decided by `efe.workbook.resolve` (highest version in the
Drive folder), never by a filename in config. What the file must look like is a
*contract*, never a count: the ordered PARTNERS header, the sheets that must exist,
a formula in every formula-column cell of every data row. And every load is
compared with the previous one (`efe.workbook.state`): fewer rows or a lower
version is an old file or a half-finished sync, and the run stops.

Two environment failure modes get their own hard stop, because retrying either one
in a loop makes things worse rather than better:

1. Google Drive hands back a zero-byte file or a placeholder stub.
2. The workbook is open in Excel and therefore exclusively locked.

Both raise immediately with an instruction, never a retry.
"""

from __future__ import annotations

import logging
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.formula import ArrayFormula

from efe.config import ENRICHABLE_LOGICAL_FIELDS, Config
from efe.models import (
    Candidate,
    ContinuityError,
    DriveSyncError,
    SchemaMismatchError,
    VersionConflictError,
    WorkbookLockedError,
)
from efe.workbook.resolve import (
    Resolution,
    WorkbookCandidate,
    candidate_for,
    filename_pattern,
    resolve_workbook,
)
from efe.workbook.state import (
    WorkbookState,
    continuity_problems,
    load_state,
    save_state,
    state_path,
)
from efe.workbook.xmlutil import read_cached_values

log = logging.getLogger(__name__)

#: A workbook this small cannot be the real thing; Drive is still syncing.
MIN_PLAUSIBLE_BYTES = 50_000

#: Logical fields the enricher can fill, in workbook column order.
ENRICHABLE_FIELDS = ENRICHABLE_LOGICAL_FIELDS


# ---------------------------------------------------------------------------
# Guards
# ---------------------------------------------------------------------------


def _lock_file_for(path: Path) -> Path:
    """Excel's owner file: `Book.xlsx` -> `~$Book.xlsx` in the same directory."""
    return path.with_name("~$" + path.name)


def guard_readable(path: Path, min_bytes: int = MIN_PLAUSIBLE_BYTES) -> None:
    """Stop the run if the workbook cannot be trusted to be the real file.

    Args:
        path: the workbook to check.
        min_bytes: size below which the file must be a Drive placeholder rather
            than the real thing. Comes from `workbook.min_plausible_bytes`.

    Raises:
        DriveSyncError: missing, empty, truncated, or not a valid xlsx container.
        WorkbookLockedError: Excel holds an exclusive lock on it.
    """
    if not path.exists():
        raise DriveSyncError(
            f"Workbook not found:\n  {path}\n"
            "If this path is on Google Drive, the file may not be synced to this "
            "machine yet. Open the Drive folder, wait for the sync to finish, then "
            "re-run. Nothing has been changed."
        )

    size = path.stat().st_size
    if size == 0:
        raise DriveSyncError(
            f"Workbook read back as ZERO BYTES:\n  {path}\n"
            "This is a Google Drive sync problem, not a data problem. Let Drive "
            "finish syncing (the file should be several hundred KB), then re-run. "
            "Nothing has been changed."
        )
    if size < min_bytes:
        raise DriveSyncError(
            f"Workbook is only {size:,} bytes:\n  {path}\n"
            f"That is below the {min_bytes:,}-byte floor for this file and "
            "looks like a Drive placeholder or a truncated download. Let Drive "
            "finish syncing, then re-run. Nothing has been changed."
        )

    lock = _lock_file_for(path)
    if lock.exists():
        raise WorkbookLockedError(
            f"The workbook is OPEN IN EXCEL:\n  {path}\n"
            f"Excel's lock file is present ({lock.name}). Close the workbook in "
            "Excel and re-run. Nothing has been changed."
        )

    try:
        with path.open("rb") as fh:
            fh.read(4)
    except PermissionError as exc:
        raise WorkbookLockedError(
            f"The workbook is LOCKED and cannot be read:\n  {path}\n"
            "It is almost certainly open in Excel. Close it and re-run. "
            "Nothing has been changed."
        ) from exc

    try:
        with zipfile.ZipFile(path) as zf:
            names = set(zf.namelist())
    except zipfile.BadZipFile as exc:
        raise DriveSyncError(
            f"Workbook is not a readable .xlsx container:\n  {path}\n"
            "Drive may have delivered a placeholder or a partial file. Let the sync "
            "finish, then re-run. Nothing has been changed."
        ) from exc

    if "xl/workbook.xml" not in names:
        raise DriveSyncError(
            f"Workbook container is missing xl/workbook.xml:\n  {path}\n"
            "The file is corrupt or still syncing. Nothing has been changed."
        )


def guard_writable(path: Path) -> None:
    """Stop before writing if the destination is locked or already taken."""
    if path.exists():
        raise WorkbookGuardOutputExists(
            f"Refusing to overwrite an existing file:\n  {path}\n"
            "Version numbers only ever go up. Nothing has been changed."
        )
    lock = _lock_file_for(path)
    if lock.exists():
        raise WorkbookLockedError(
            f"The intended output file is OPEN IN EXCEL:\n  {path}\n"
            "Close it and re-run. Nothing has been changed."
        )
    parent = path.parent
    if not parent.is_dir():
        raise DriveSyncError(
            f"Output directory does not exist:\n  {parent}\n"
            "If this is a Google Drive path, check the folder is synced offline."
        )


class WorkbookGuardOutputExists(VersionConflictError):
    """The versioned output filename is already taken."""


def last_writer_of(path: Path) -> str:
    """The application that last saved the file, from docProps/app.xml ("" if unknown)."""
    import re

    try:
        with zipfile.ZipFile(path) as zf:
            if "docProps/app.xml" not in zf.namelist():
                return ""
            xml = zf.read("docProps/app.xml").decode("utf-8", "replace")
    except (OSError, zipfile.BadZipFile):
        return ""
    match = re.search("<Application>(.*?)</Application>", xml)
    return match.group(1).strip() if match else ""


def file_fingerprint(path: Path) -> str:
    """SHA-256 of the file as read. The writer refuses an input that no longer
    matches it: a Drive re-sync or a Sheets download between load and write
    would otherwise be applied to rows the run never looked at."""
    import hashlib

    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


# ---------------------------------------------------------------------------
# Schema: invariants, not constants
# ---------------------------------------------------------------------------


def last_data_row(ws, spec) -> int:
    """The last row that actually holds an entity.

    `ws.max_row` is the used range, not the data. Excel and Google Sheets both pad a
    saved sheet out to their default grid -- a round trip through Sheets left this
    workbook reporting 1000 rows with 746 of them empty. Trailing blanks are
    cosmetic; the data extent is what the schema check cares about.
    """
    id_col = column_index_from_string(spec.column_for("id"))
    name_col = column_index_from_string(spec.column_for("entity_name"))
    for row in range(ws.max_row, spec.first_data_row - 1, -1):
        if (
            str(ws.cell(row, id_col).value or "").strip()
            or str(ws.cell(row, name_col).value or "").strip()
        ):
            return row
    return spec.header_row


def read_header(ws, spec) -> list[str]:
    """The header row as text, trailing blank cells trimmed."""
    values = [ws.cell(spec.header_row, c).value for c in range(1, ws.max_column + 1)]
    while values and values[-1] in (None, ""):
        values.pop()
    return ["" if v is None else str(v) for v in values]


def header_diff(expected: list[str], found: list[str]) -> list[str]:
    """Positional diff of two header rows, one line per disagreement."""
    lines: list[str] = []
    for index in range(max(len(expected), len(found))):
        want = expected[index] if index < len(expected) else None
        have = found[index] if index < len(found) else None
        if want == have:
            continue
        letter = get_column_letter(index + 1)
        if want is None:
            lines.append(f"{letter}: unexpected extra column {have!r}")
        elif have is None:
            lines.append(f"{letter}: missing column {want!r}")
        else:
            lines.append(f"{letter}: expected {want!r}, found {have!r}")
    missing = [n for n in expected if n not in found]
    extra = [n for n in found if n not in expected]
    if missing:
        lines.append(f"names missing entirely: {missing}")
    if extra:
        lines.append(f"names not in the contract: {extra}")
    return lines


def formula_gaps(ws, spec, last_row: int) -> dict[str, list[int]]:
    """Formula column -> data rows whose cell does not hold a formula.

    A per-row `=...` string counts, and so does an array formula (Google Sheets'
    ARRAYFORMULA exports as one anchor cell whose `ref` spans the column): every
    row inside its range is covered.
    """
    gaps: dict[str, list[int]] = {}
    for name in spec.formula_columns:
        col = column_index_from_string(spec.letter_of(name))
        covered: set[int] = set()
        for r in range(spec.first_data_row, last_row + 1):
            value = ws.cell(r, col).value
            if isinstance(value, ArrayFormula) and value.ref:
                _, top, _, bottom = range_boundaries(value.ref)
                covered.update(range(top, bottom + 1))
        rows = []
        for r in range(spec.first_data_row, last_row + 1):
            value = ws.cell(r, col).value
            if r in covered or (isinstance(value, str) and value.startswith("=")):
                continue
            rows.append(r)
        if rows:
            gaps[name] = rows
    return gaps


@dataclass
class SchemaReport:
    """What the schema check saw. `problems()` is empty when the contract holds."""

    sheets_found: list[str]
    sheets_missing: list[str]
    header_found: list[str]
    header_problems: list[str]
    last_row: int
    data_rows: int
    formula_gaps: dict[str, list[int]]
    padded_rows: int

    def problems(self) -> list[str]:
        out: list[str] = []
        if self.sheets_missing:
            out.append(
                f"required sheets missing: {self.sheets_missing}\n      found: {self.sheets_found}"
            )
        if self.header_problems:
            out.append(
                "the PARTNERS header does not match the contract in config.yaml:\n      "
                + "\n      ".join(self.header_problems)
            )
        for name, rows in self.formula_gaps.items():
            out.append(
                f"{name} must hold a formula on every data row; "
                f"{len(rows)} row(s) do not, e.g. {rows[:10]}"
            )
        if self.data_rows <= 0:
            out.append("PARTNERS holds no data rows")
        return out


def inspect_schema(wb: Workbook, cfg: Config, label: str = "") -> SchemaReport:
    """Check the workbook against the contract without raising."""
    spec = cfg.workbook
    if spec.sheet not in wb.sheetnames:
        raise SchemaMismatchError(
            f"{label or 'the workbook'}: sheet {spec.sheet!r} is missing (found {wb.sheetnames})"
        )
    ws = wb[spec.sheet]
    header = read_header(ws, spec)
    header_problems = header_diff(spec.header, header)
    populated = last_data_row(ws, spec)
    gaps = formula_gaps(ws, spec, populated) if not header_problems else {}
    padded = ws.max_row - populated
    if padded > 0:
        # Tolerated: a spreadsheet app padded the used range. Say so rather than
        # failing, so the difference is visible if it ever means something else.
        log.info(
            "%s: used range runs to row %d but data ends at %d; %d trailing empty rows ignored",
            spec.sheet,
            ws.max_row,
            populated,
            padded,
        )
    return SchemaReport(
        sheets_found=list(wb.sheetnames),
        sheets_missing=[s for s in spec.required_sheets if s not in wb.sheetnames],
        header_found=header,
        header_problems=header_problems,
        last_row=populated,
        data_rows=max(0, populated - spec.first_data_row + 1),
        formula_gaps=gaps,
        padded_rows=max(0, padded),
    )


def assert_schema(wb: Workbook, cfg: Config, label: str = "") -> SchemaReport:
    """Confirm the workbook honours the contract; raise with the full diff if not.

    Reading a workbook whose columns have moved and writing to hardcoded letters is
    the one mistake that silently destroys data. This makes it impossible.
    """
    report = inspect_schema(wb, cfg, label)
    problems = report.problems()
    if problems:
        joined = "\n  - ".join(problems)
        raise SchemaMismatchError(
            f"{label or 'The workbook'} does not match the contract in config.yaml:\n  - "
            + joined
            + "\n\nNothing has been changed. Fix the workbook (or, if the contract "
            "really changed, config.yaml) and re-run."
        )
    return report


def count_formulas(wb: Workbook) -> int:
    return sum(
        1
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )


# ---------------------------------------------------------------------------
# Row model
# ---------------------------------------------------------------------------


@dataclass(slots=True)
class PartnerRow:
    """One PARTNERS data row as read, keyed by column letter."""

    row: int
    cells: dict[str, str | None]

    def get(self, letter: str) -> str:
        value = self.cells.get(letter)
        return "" if value is None else str(value).strip()

    def logical(self, cfg: Config, logical_field: str) -> str:
        return self.get(cfg.workbook.column_for(logical_field))


@dataclass(slots=True)
class WorkbookView:
    """Everything read from the workbook that the pipeline needs."""

    path: Path
    rows: list[PartnerRow]
    formula_count: int
    domain_row_counts: Counter[str] = field(default_factory=Counter)
    ledger_domains: dict[str, dict[str, str]] = field(default_factory=dict)
    duplicate_domains: dict[str, list[int]] = field(default_factory=dict)
    #: domain -> the Resort_Base of every row using it, for the scope guard.
    domain_resorts: dict[str, list[str]] = field(default_factory=dict)
    #: header text -> column letter, read from the sheet rather than configured,
    #: so the CRM columns can be shown by name without eleven config entries.
    header_letters: dict[str, str] = field(default_factory=dict)
    #: vNN parsed from the filename; 0 when the name carries no version.
    version: int = 0
    #: The header row as read (equal to the contract once the schema check passed).
    header: list[str] = field(default_factory=list)
    sheets: list[str] = field(default_factory=list)
    resolution: Resolution | None = None
    schema: SchemaReport | None = None
    #: What the previous run recorded, if anything.
    previous_state: WorkbookState | None = None
    #: SHA-256 of the file as read; the writer refuses to work from a changed file.
    fingerprint: str = ""
    #: Formula cells in the file, and how many carry a cached result. A file last
    #: written by openpyxl carries none; Sheets and Excel recompute on open.
    formula_cells: int = 0
    cached_results: int = 0
    #: docProps/app.xml Application, e.g. 'Microsoft Excel' or 'Openpyxl 3.1.5'.
    last_writer: str = ""

    @property
    def data_rows(self) -> int:
        return len(self.rows)


def domain_of(url: str) -> str:
    """Registrable host for a website URL, lowercased, `www.` stripped.

    Returns an empty string for anything that is not a usable URL.
    """
    if not url:
        return ""
    text = url.strip()
    if not text or text.upper() == "TBD":
        return ""
    if "://" not in text:
        text = "https://" + text
    try:
        host = urlparse(text).netloc.lower()
    except ValueError:
        return ""
    host = host.split("@")[-1].split(":")[0]
    return host[4:] if host.startswith("www.") else host


def normalise_website(url: str) -> str:
    """Canonical absolute URL for a Website_URL cell, or '' if unusable."""
    domain = domain_of(url)
    if not domain:
        return ""
    text = url.strip()
    if "://" not in text:
        text = "https://" + text
    parsed = urlparse(text)
    scheme = "https" if parsed.scheme not in ("http", "https") else parsed.scheme
    path = parsed.path or "/"
    return f"{scheme}://{parsed.netloc}{path}"


def _read_source_ledger(wb: Workbook) -> dict[str, dict[str, str]]:
    """Round-1 source ledger, keyed by domain.

    Used to log revisits. The `Exclude_Next_Round` flag governs *discovery*
    (Phase 2); enrichment revisits these domains on purpose, because Round 1 visited
    them to find entities, not to extract contacts.
    """
    if "_SOURCES" not in wb.sheetnames:
        return {}
    ws = wb["_SOURCES"]
    header_row = None
    for r in range(1, min(ws.max_row, 12) + 1):
        if str(ws.cell(r, 2).value or "").strip().lower() == "domain":
            header_row = r
            break
    if header_row is None:
        return {}

    headers = {
        str(ws.cell(header_row, c).value or "").strip(): c for c in range(1, ws.max_column + 1)
    }
    out: dict[str, dict[str, str]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        raw = ws.cell(r, headers.get("Domain", 2)).value
        if not raw:
            continue
        key = domain_of(str(raw)) or str(raw).strip().lower()
        out[key] = {
            "domain": key,
            "category_covered": str(ws.cell(r, headers.get("Category_Covered", 3)).value or ""),
            "round": str(ws.cell(r, headers.get("Round", 4)).value or ""),
            "exclude_next_round": str(ws.cell(r, headers.get("Exclude_Next_Round", 5)).value or "")
            .strip()
            .upper(),
        }
    return out


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------


def resolve_input(cfg: Config, path: Path | None = None) -> tuple[WorkbookCandidate, Resolution]:
    """The file to read: an explicit override, or the highest version in the folder."""
    basename = cfg.output_basename
    if path is not None:
        target = Path(path)
        guard_readable(target, cfg.workbook.min_plausible_bytes)
        chosen = candidate_for(target, basename)
        resolution = Resolution(
            directory=target.parent,
            pattern=filename_pattern(basename).pattern,
            chosen=chosen,
            override=True,
        )
        log.info("%s", resolution.describe())
        return chosen, resolution

    resolution = resolve_workbook(
        cfg.workbook_directory,
        basename=basename,
        min_bytes=cfg.workbook.min_plausible_bytes,
        exclude_tokens=tuple(cfg.workbook.exclude_name_tokens),
        # An output folder that differs from the input folder still holds the
        # file the writer just emitted; it must be the one read next.
        extra_directories=[cfg.output_directory],
    )
    assert resolution.chosen is not None  # resolve_workbook raises otherwise
    guard_readable(resolution.chosen.path, cfg.workbook.min_plausible_bytes)
    return resolution.chosen, resolution


def load_workbook_view(
    cfg: Config,
    path: Path | None = None,
    *,
    reset_state: bool = False,
    record_state: bool = True,
    command: str = "efe",
    resolved: tuple[WorkbookCandidate, Resolution] | None = None,
) -> WorkbookView:
    """Resolve, guard, open, validate, compare with the last run, and read.

    Args:
        cfg: configuration.
        path: explicit workbook (`--workbook`); skips the folder scan.
        reset_state: accept this file as the new baseline even if it is smaller or
            older than what the last run recorded.
        record_state: write `data/state/workbook.json` after a successful load
            (never from an unversioned `--workbook` file: see `record_input_state`).
        command: label stored in the state file.
        resolved: a resolution already made (and reported) by the caller.

    Raises:
        DriveSyncError / WorkbookLockedError: the environment is not safe to read.
        SchemaMismatchError: the contract does not hold (full diff in the message).
        ContinuityError: this file goes backwards relative to the last run.
    """
    chosen, resolution = resolved or resolve_input(cfg, path)
    target = chosen.path
    spec = cfg.workbook

    cached = read_cached_values(target)
    wb = load_workbook(target, data_only=False)
    try:
        report = assert_schema(wb, cfg, target.name)
        # The header now equals the contract; bind the letters from what the sheet
        # actually holds, so nothing downstream depends on config-typed letters.
        spec.bind_header(report.header_found)
        ws = wb[spec.sheet]
        width = len(report.header_found)
        letters = [get_column_letter(c) for c in range(1, width + 1)]
        header_letters = {
            str(ws.cell(spec.header_row, c).value): letters[c - 1]
            for c in range(1, width + 1)
            if ws.cell(spec.header_row, c).value
        }

        rows: list[PartnerRow] = []
        for r in range(spec.first_data_row, report.last_row + 1):
            cells = {letters[c - 1]: ws.cell(r, c).value for c in range(1, width + 1)}
            rows.append(PartnerRow(row=r, cells=cells))

        try:
            previous = load_state(state_path(cfg.state_directory))
        except ContinuityError:
            if not reset_state:
                raise
            previous = None  # unreadable baseline, explicitly discarded
        problems = (
            []
            if reset_state
            else continuity_problems(
                previous,
                file=target.name,
                version=chosen.version,
                data_rows=len(rows),
                header=report.header_found,
            )
        )
        if problems:
            joined = "\n  - ".join(problems)
            raise ContinuityError(
                f"{target.name} goes backwards relative to the last run:\n  - "
                + joined
                + f"\n\nBaseline: {state_path(cfg.state_directory)}\n"
                "If Drive is still syncing, wait and re-run. If reading this file is "
                "deliberate, re-run with --reset-state to make it the new baseline. "
                "Nothing has been changed."
            )

        website_col = spec.column_for("website_url")
        resort_col = spec.column_for("resort_base")
        domain_counts: Counter[str] = Counter()
        by_domain: dict[str, list[int]] = defaultdict(list)
        resorts: dict[str, list[str]] = defaultdict(list)
        for pr in rows:
            d = domain_of(pr.get(website_col))
            if d:
                domain_counts[d] += 1
                by_domain[d].append(pr.row)
                resorts[d].append(pr.get(resort_col))

        view = WorkbookView(
            path=target,
            rows=rows,
            formula_count=count_formulas(wb),
            domain_row_counts=domain_counts,
            ledger_domains=_read_source_ledger(wb),
            duplicate_domains={d: rs for d, rs in by_domain.items() if len(rs) > 1},
            domain_resorts=dict(resorts),
            header_letters=header_letters,
            version=chosen.version,
            header=list(report.header_found),
            sheets=list(wb.sheetnames),
            resolution=resolution,
            schema=report,
            previous_state=previous,
            fingerprint=file_fingerprint(target),
            formula_cells=len(cached),
            cached_results=sum(1 for v in cached.values() if v is not None),
            last_writer=last_writer_of(target),
        )
    finally:
        wb.close()

    if record_state:
        record_input_state(cfg, view, command)
    return view


def record_input_state(cfg: Config, view: WorkbookView, command: str) -> bool:
    """Record `view` as the continuity baseline. Returns False when it was not.

    An unversioned `--workbook` file is never recorded: a baseline of "v00" would
    switch the version-regression guard off for every run after it.
    """
    if not view.version:
        log.warning(
            "baseline NOT updated: %s carries no vNN, so it cannot serve as the "
            "continuity baseline",
            view.path.name,
        )
        return False
    save_state(
        state_path(cfg.state_directory),
        WorkbookState.now(
            file=view.path.name,
            version=view.version,
            data_rows=len(view.rows),
            header=view.header,
            command=command,
        ),
    )
    return True


# ---------------------------------------------------------------------------
# Selection
# ---------------------------------------------------------------------------


def _resort_forms(text: str) -> tuple[str, str]:
    """Two comparison forms of a resort/place name, flattened to alnum.

    German places are spelled both ways in the sheet -- `Kitzbühel` and
    `Kitzbuehel`, `Zürs` and `Zurs` -- so a name matches if either its
    umlaut-expanded form (ü->ue) or its plain accent-stripped form (ü->u) is a
    substring of either form of the cell.
    """
    import re as _re
    import unicodedata as _ud

    lowered = (text or "").lower()
    expanded = lowered.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")

    def flat(s: str) -> str:
        s = _ud.normalize("NFKD", s)
        s = "".join(ch for ch in s if not _ud.combining(ch))
        return _re.sub(r"[^a-z0-9]", "", s)

    return flat(expanded), flat(lowered)


def resort_matches(cell_text: str, wanted: list[str]) -> bool:
    """Whether a Resort_Base / Region_Valley cell names one of the target resorts."""
    if not wanted:
        return True
    cell_forms = _resort_forms(cell_text)
    for name in wanted:
        for want in _resort_forms(name):
            if want and any(want in c for c in cell_forms):
                return True
    return False


NEEDS_MANUAL_URL_REASON = (
    "needs_manual_url: this domain refuses automated access (403/TLS), so the row "
    "needs a property-level Website_URL before it can be enriched"
)


def select_candidates(view: WorkbookView, cfg: Config) -> tuple[list[Candidate], dict[int, str]]:
    """Split rows into what will be processed and what is skipped, with reasons.

    Skip reasons are reported verbatim in the run report, so a row never silently
    disappears from the pipeline.
    """
    spec = cfg.workbook
    candidates: list[Candidate] = []
    skipped: dict[int, str] = {}

    website_col = spec.column_for("website_url")
    name_col = spec.column_for("entity_name")
    id_col = spec.column_for("id")
    skip_rules = [(rule, spec.letter_of(rule.column)) for rule in cfg.selection.skip_when]

    for pr in view.rows:
        skip_reason = ""
        for rule, letter in skip_rules:
            current = pr.get(letter).upper()
            if current in {v.upper() for v in rule.values}:
                skip_reason = (
                    f"human-verified row ({rule.column}={pr.get(letter)!r}) "
                    "- never fetched, never written"
                )
                break
        if skip_reason:
            skipped[pr.row] = skip_reason
            continue

        website = pr.get(website_col)
        domain = domain_of(website)
        if cfg.selection.require_website and not domain:
            skipped[pr.row] = "no usable Website_URL - needs discovery, not enrichment"
            continue

        blocked = {d.lower() for d in cfg.scope.needs_manual_url}
        if domain in blocked or any(domain.endswith("." + d) for d in blocked):
            skipped[pr.row] = NEEDS_MANUAL_URL_REASON
            continue

        if cfg.selection.categories:
            category = pr.logical(cfg, "category").strip()
            if not any(category.startswith(p) for p in cfg.selection.categories):
                skipped[pr.row] = (
                    f"outside the target categories {cfg.selection.categories} "
                    "(selection.categories; override with --categories all)"
                )
                continue
        if cfg.selection.resorts:
            place = f"{pr.logical(cfg, 'resort_base')} {pr.logical(cfg, 'region_valley')}"
            if not resort_matches(place, cfg.selection.resorts):
                skipped[pr.row] = (
                    "outside the target resorts (selection.resorts; override with --resorts all)"
                )
                continue

        existing = {f: pr.logical(cfg, f) for f in ENRICHABLE_FIELDS}
        if all(not spec.is_empty(v) for v in existing.values()):
            skipped[pr.row] = "every enrichable field is already filled"
            continue

        candidates.append(
            Candidate(
                entity_id=pr.get(id_col) or f"ROW-{pr.row}",
                row=pr.row,
                name=pr.get(name_col),
                website_url=normalise_website(website),
                domain=domain_of(website),
                category=pr.logical(cfg, "category"),
                country=pr.logical(cfg, "country"),
                resort_base=pr.logical(cfg, "resort_base"),
                existing=existing,
            )
        )

    return candidates, skipped
