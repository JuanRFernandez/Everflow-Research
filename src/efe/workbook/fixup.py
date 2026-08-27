"""Corrective releases: repairing what a human, or an earlier run, left wrong.

Enrichment fills empty cells and promotion appends rows. Neither can renumber an ID
or rewrite a value somebody already typed -- and deliberately so: `ID` is in no
writable set, and `assert_changes_legal` refuses to overwrite a non-empty cell.
That is the right default, and it is also why a workbook can end up with 39 IDs used
twice and no supported way to fix it.

`efe fixup` is that way, kept narrow on purpose:

* Every edit is a line in a CSV a human reads and approves before anything runs.
* Every line carries the value it expects to find (`Old`) plus an independent witness
  (`Guard_Column`/`Guard_Value`); if either has moved, the whole run refuses rather
  than half-applying a plan built against bytes that no longer exist.
* The plan pins the input's SHA-256, so it can only ever be applied to the file it
  was computed from.
* Only `writable ∪ provenance ∪ {ID}` may be touched. CRM columns, formula columns
  and formula precedents stay out of reach, exactly as in every other write path.
* Every old value lands in CHANGELOG_DETAIL, which makes the tool its own undo:
  swap `Old` and `New` in the plan and run it again.
"""

from __future__ import annotations

import csv
import io
import logging
import re
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import phonenumbers
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from efe.config import Config
from efe.extract.phones import region_for, to_e164
from efe.models import CellChange, Confidence, DataClass, VerificationError, today_iso
from efe.workbook.promote import ID_PATTERN
from efe.workbook.ranges import assert_touches_no_formula_input
from efe.workbook.reader import (
    WorkbookView,
    domain_of,
    file_fingerprint,
    guard_readable,
    guard_writable,
)
from efe.workbook.resolve import parse_version
from efe.workbook.verify import compare, format_report, snapshot
from efe.workbook.writer import (
    _append_changelog,
    _write_changelog_detail,
    assert_version_free,
    deliver,
    next_version_path,
)
from efe.workbook.xmlutil import read_cached_values, reinject_cached_values

log = logging.getLogger(__name__)

FIXUP_PLAN_HEADERS = ["Kind", "Row", "Column", "Old", "New", "Guard_Column", "Guard_Value", "Why"]
KINDS = ("meta", "renumber", "cell", "review")
#: Default ceiling on cell edits in one plan, so a runaway proposal cannot quietly
#: rewrite the sheet.
MAX_CELLS = 300

_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
_MULTI_PHONE_RE = re.compile(r"[/;]|,\s*\+|\s+(?:o|or|und|and)\s+", re.I)
#: E.164 is an international address; these kinds are not reachable from abroad.
_NOT_DIALABLE_ABROAD = (
    phonenumbers.PhoneNumberType.TOLL_FREE,
    phonenumbers.PhoneNumberType.PREMIUM_RATE,
    phonenumbers.PhoneNumberType.SHARED_COST,
)


@dataclass
class FixupRecord:
    kind: str
    row: int
    column: str
    old: str
    new: str
    guard_column: str = ""
    guard_value: str = ""
    why: str = ""

    def as_csv_row(self) -> dict[str, str]:
        return {
            "Kind": self.kind,
            "Row": str(self.row),
            "Column": self.column,
            "Old": self.old,
            "New": self.new,
            "Guard_Column": self.guard_column,
            "Guard_Value": self.guard_value,
            "Why": self.why,
        }


@dataclass
class FixupPlan:
    """What a corrective release would change, and against which bytes."""

    source: Path
    input_name: str = ""
    input_sha256: str = ""
    renumbers: list[FixupRecord] = field(default_factory=list)
    cells: list[FixupRecord] = field(default_factory=list)
    #: Things the proposer refused to automate, for a human to settle.
    review: list[str] = field(default_factory=list)

    @property
    def records(self) -> list[FixupRecord]:
        return self.renumbers + self.cells

    @property
    def id_map(self) -> dict[str, str]:
        return {r.old: r.new for r in self.renumbers}

    def renumber_summary(self) -> str:
        if not self.renumbers:
            return ""
        rows = [r.row for r in self.renumbers]
        olds = [r.old for r in self.renumbers]
        news = [r.new for r in self.renumbers]
        return (
            f"{len(self.renumbers)} IDs, rows {min(rows)}..{max(rows)}, "
            f"{min(olds)}..{max(olds)} -> {min(news)}..{max(news)}"
        )

    def cell_summary(self) -> dict[str, int]:
        out: dict[str, int] = {}
        for record in self.cells:
            out[record.column] = out.get(record.column, 0) + 1
        return out

    def describe(self) -> str:
        lines: list[str] = []
        if self.renumbers:
            lines.append(f"renumber : {self.renumber_summary()}")
        if self.cells:
            lines.append(f"cells    : {len(self.cells)} edit(s)")
            for column, count in sorted(self.cell_summary().items()):
                lines.append(f"           {column:28s} {count:4d}")
        if not self.records:
            lines.append("nothing to fix: the plan holds no records")
        for item in self.review:
            lines.append(f"review   : {item}")
        return "\n".join(lines)


# ---------------------------------------------------------------------------
# The plan file
# ---------------------------------------------------------------------------


def write_fixup_plan(path: Path, plan: FixupPlan) -> Path:
    """Emit the reviewable CSV. The meta row pins the bytes it was computed from."""
    path.parent.mkdir(parents=True, exist_ok=True)
    meta = FixupRecord(
        kind="meta",
        row=0,
        column="ID",
        old=plan.input_sha256,
        new=plan.input_name,
        why="input this plan was computed from; it may only be applied to these bytes",
    )
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=FIXUP_PLAN_HEADERS)
        writer.writeheader()
        for record in [meta, *plan.records]:
            writer.writerow(record.as_csv_row())
        for item in plan.review:
            writer.writerow(
                FixupRecord(
                    kind="review", row=0, column="ID", old="", new="", why=item
                ).as_csv_row()
            )
    return path


def read_fixup_plan(path: Path, cfg: Config) -> FixupPlan:
    """Parse and shape-check a plan CSV. Strict, for the same reason candidates are."""
    spec = cfg.workbook
    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError as exc:
        raise VerificationError(
            f"{path.name} is not UTF-8 ({exc.reason} at byte {exc.start}); save it as "
            "UTF-8 CSV and re-run. Nothing has been written."
        ) from exc

    with io.StringIO(text, newline="") as fh:
        reader = csv.reader(fh)
        try:
            columns = [c.strip() for c in next(reader)]
        except StopIteration:
            raise VerificationError(f"{path.name} is empty. Nothing has been written.") from None
        if columns != FIXUP_PLAN_HEADERS:
            raise VerificationError(
                f"{path.name} header is {columns}, expected {FIXUP_PLAN_HEADERS}. "
                "Nothing has been written."
            )
        raw: list[tuple[int, dict[str, str]]] = []
        for line_no, record in enumerate(reader, start=2):
            if not any(cell.strip() for cell in record):
                continue
            if len(record) != len(columns):
                raise VerificationError(
                    f"{path.name} line {line_no}: {len(record)} fields, the header has "
                    f"{len(columns)}. Fix the CSV. Nothing has been written."
                )
            raw.append((line_no, {k: v.strip() for k, v in zip(columns, record, strict=True)}))

    plan = FixupPlan(source=path)
    for line_no, row in raw:
        kind = row["Kind"].lower()
        if kind not in KINDS:
            raise VerificationError(
                f"{path.name} line {line_no}: Kind {row['Kind']!r} is not one of {KINDS}. "
                "Nothing has been written."
            )
        if kind == "meta":
            plan.input_sha256, plan.input_name = row["Old"], row["New"]
            continue
        if kind == "review":
            plan.review.append(row["Why"])
            continue
        try:
            row_number = int(row["Row"])
        except ValueError:
            raise VerificationError(
                f"{path.name} line {line_no}: Row {row['Row']!r} is not an integer. "
                "Nothing has been written."
            ) from None
        if row_number < spec.first_data_row:
            raise VerificationError(
                f"{path.name} line {line_no}: row {row_number} is above the first data row. "
                "Nothing has been written."
            )
        if row["Column"] not in spec.header:
            raise VerificationError(
                f"{path.name} line {line_no}: {row['Column']!r} is not a PARTNERS column. "
                "Nothing has been written."
            )
        if row["Guard_Column"] and row["Guard_Column"] not in spec.header:
            raise VerificationError(
                f"{path.name} line {line_no}: Guard_Column {row['Guard_Column']!r} is not a "
                "PARTNERS column. Nothing has been written."
            )
        record = FixupRecord(
            kind=kind,
            row=row_number,
            column=row["Column"],
            old=row["Old"],
            new=row["New"],
            guard_column=row["Guard_Column"],
            guard_value=row["Guard_Value"],
            why=row["Why"],
        )
        if kind == "renumber":
            if record.column != spec.columns["id"]:
                raise VerificationError(
                    f"{path.name} line {line_no}: a renumber must target the "
                    f"{spec.columns['id']!r} column, not {record.column!r}. "
                    "Nothing has been written."
                )
            plan.renumbers.append(record)
        else:
            plan.cells.append(record)
    return plan


# ---------------------------------------------------------------------------
# Guards
# ---------------------------------------------------------------------------


def fixable_columns(cfg: Config) -> set[str]:
    """Columns a corrective release may touch: the tool's own, plus `ID`.

    Never a CRM column, a formula column or a formula precedent -- those belong to
    the human and to the cached results the writer preserves.
    """
    spec = cfg.workbook
    allowed = set(spec.writable_columns) | set(spec.provenance_columns) | {spec.columns["id"]}
    forbidden = set(spec.crm_columns) | set(spec.formula_columns) | set(spec.formula_precedents)
    return allowed - forbidden


def assert_fixups_legal(plan: FixupPlan, cfg: Config, view: WorkbookView) -> None:
    """Refuse a plan that targets the wrong columns, rows or values."""
    spec = cfg.workbook
    allowed = fixable_columns(cfg)
    crm = set(spec.crm_columns) | set(spec.formula_columns) | set(spec.formula_precedents)
    data_rows = {pr.row for pr in view.rows}
    problems: list[str] = []
    seen: set[tuple[int, str]] = set()

    if plan.input_sha256 and view.fingerprint and plan.input_sha256 != view.fingerprint:
        problems.append(
            f"the plan was built against {plan.input_sha256[:12]}... ({plan.input_name}) but "
            f"the workbook is {view.fingerprint[:12]}... - regenerate it with --propose"
        )
    if len(plan.cells) > MAX_CELLS:
        problems.append(
            f"{len(plan.cells)} cell edits exceed the {MAX_CELLS} ceiling; raise --max-cells "
            "deliberately if that is really intended"
        )

    for record in plan.records:
        label = f"line for {record.column}{record.row}"
        if record.column in crm:
            problems.append(
                f"{label} is a human-owned CRM, formula or precedent column - refusing to write"
            )
            continue
        if record.column not in allowed:
            problems.append(f"{label} is not a fixable column ({sorted(allowed)})")
            continue
        if record.row not in data_rows:
            problems.append(f"row {record.row} is not a PARTNERS data row")
            continue
        key = (record.row, record.column)
        if key in seen:
            problems.append(f"{label} appears twice in the plan")
            continue
        seen.add(key)
        if not record.new:
            problems.append(f"{label} has an empty New value; blanking a cell is a Sheets job")
        if record.new.lstrip().startswith("="):
            problems.append(f"{label} value {record.new!r} would be stored as a live formula")
        if record.old == record.new:
            problems.append(f"{label} changes nothing ({record.old!r})")

    if problems:
        joined = "\n  - ".join(problems)
        raise VerificationError("The fixup plan is illegal and was NOT applied:\n  - " + joined)


def assert_ids_unique_after(
    plan: FixupPlan,
    cfg: Config,
    view: WorkbookView,
    *,
    reserved_ids: set[str] | None = None,
    allow_backfill: bool = False,
) -> None:
    """The post-image must hold one distinct ID per row, and take nobody's number."""
    spec = cfg.workbook
    id_col = spec.column_for("id")
    reserved = reserved_ids or set()
    current = {pr.row: pr.get(id_col) for pr in view.rows}
    mapped = {r.row: r.new for r in plan.renumbers}
    problems: list[str] = []

    highest = 0
    for value in current.values():
        if ID_PATTERN.match(value):
            highest = max(highest, int(value[4:]))

    targets: dict[str, int] = {}
    for record in plan.renumbers:
        if not ID_PATTERN.match(record.new):
            problems.append(f"row {record.row}: {record.new!r} is not of the form EFE-dddd")
            continue
        if record.new in targets:
            problems.append(
                f"row {record.row}: {record.new} is also the target of row {targets[record.new]}"
            )
            continue
        targets[record.new] = record.row
        if record.new in reserved:
            problems.append(
                f"row {record.row}: {record.new} is reserved by another list and may not be used"
            )
        if not allow_backfill and int(record.new[4:]) <= highest:
            problems.append(
                f"row {record.row}: {record.new} is not above the sheet's highest ID "
                f"EFE-{highest:04d} (pass --allow-backfill to reuse freed numbers)"
            )

    post = {row: mapped.get(row, value) for row, value in current.items()}
    taken: dict[str, list[int]] = {}
    for row, value in post.items():
        taken.setdefault(value, []).append(row)
    for value, rows in sorted(taken.items()):
        if len(rows) > 1:
            problems.append(f"after the plan, {value} would still be on rows {rows}")

    if problems:
        joined = "\n  - ".join(problems)
        raise VerificationError("The renumbering would not leave every ID unique:\n  - " + joined)


def _assert_old_values_match(ws, spec, plan: FixupPlan, *, skip_stale: bool) -> list[FixupRecord]:
    """Every record must find the value it expects, in the cell and in its witness."""
    stale: list[str] = []
    applied: list[FixupRecord] = []
    for record in plan.records:
        letter = spec.letter_of(record.column)
        cell = ws.cell(record.row, column_index_from_string(letter))
        found = "" if cell.value is None else str(cell.value).strip()
        if not isinstance(cell.value, (str, type(None))):
            stale.append(
                f"{letter}{record.row} holds {type(cell.value).__name__} {cell.value!r}, "
                "not text; a corrective release only rewrites text"
            )
            continue
        if found != record.old:
            stale.append(f"{letter}{record.row}: plan expected {record.old!r}, found {found!r}")
            continue
        if record.guard_column:
            gletter = spec.letter_of(record.guard_column)
            gcell = ws.cell(record.row, column_index_from_string(gletter))
            gfound = "" if gcell.value is None else str(gcell.value).strip()
            if gfound != record.guard_value:
                stale.append(
                    f"{gletter}{record.row}: witness expected {record.guard_value!r}, "
                    f"found {gfound!r} - the row moved"
                )
                continue
        applied.append(record)
    if stale and not skip_stale:
        joined = "\n  - ".join(stale)
        raise VerificationError(
            "The plan no longer matches the workbook and was NOT applied:\n  - "
            + joined
            + "\n\nRegenerate it with `efe fixup --propose`. Nothing has been written."
        )
    if stale:
        log.warning("--skip-stale: %d record(s) skipped: %s", len(stale), stale[:5])
    return applied


# ---------------------------------------------------------------------------
# The proposer
# ---------------------------------------------------------------------------


def _cell(view: WorkbookView, cfg: Config, row: int, column: str) -> str:
    letter = cfg.workbook.letter_of(column)
    for pr in view.rows:
        if pr.row == row:
            return pr.get(letter)
    return ""


def _propose_renumbers(
    view: WorkbookView, cfg: Config, block_start: int | None
) -> list[FixupRecord]:
    """Move the block that was pasted in last off the IDs it collided with.

    The later row of a colliding pair is the newcomer; its whole `Round` block moves
    with it, so the new numbers stay contiguous and in row order.
    """
    spec = cfg.workbook
    duplicates = view.schema.duplicate_ids if view.schema else {}
    if not duplicates:
        return []
    id_col, round_col = spec.column_for("id"), spec.column_for("round")
    by_row = {pr.row: pr for pr in view.rows}
    newcomer_rows = {max(rows) for rows in duplicates.values()}
    blocks = {by_row[r].get(round_col) for r in newcomer_rows if r in by_row}
    rows = sorted(pr.row for pr in view.rows if pr.get(round_col) in blocks and blocks != {""})
    if not rows:
        rows = sorted(newcomer_rows)

    highest = 0
    for pr in view.rows:
        value = pr.get(id_col)
        if ID_PATTERN.match(value):
            highest = max(highest, int(value[4:]))
    start = block_start if block_start is not None else highest + 1

    out: list[FixupRecord] = []
    for offset, row in enumerate(rows):
        pr = by_row[row]
        old = pr.get(id_col)
        new = f"EFE-{start + offset:04d}"
        if old in duplicates:
            others = [r for r in duplicates[old] if r != row]
            why = (
                f"{old} is also on row {others[0]} "
                f"({by_row[others[0]].get(round_col)}); this block moves"
            )
        else:
            why = f"moves with its {pr.get(round_col)} block so the block stays contiguous"
        out.append(
            FixupRecord(
                kind="renumber",
                row=row,
                column=spec.columns["id"],
                old=old,
                new=new,
                guard_column=spec.columns["entity_name"],
                guard_value=pr.get(spec.column_for("entity_name")),
                why=why,
            )
        )
    return out


def _propose_sentinels(view: WorkbookView, cfg: Config) -> list[FixupRecord]:
    """`TBD (sin email trade)` is not `TBD`: the enricher reads it as a real value."""
    spec = cfg.workbook
    tokens = [t.strip().upper() for t in spec.empty_tokens if t.strip()]
    out: list[FixupRecord] = []
    for column in spec.writable_columns:
        letter = spec.letter_of(column)
        for pr in view.rows:
            value = pr.get(letter)
            if not value or spec.is_empty(value):
                continue
            upper = value.upper()
            for token in tokens:
                if upper.startswith(token) and upper != token:
                    out.append(
                        FixupRecord(
                            kind="cell",
                            row=pr.row,
                            column=column,
                            old=value,
                            new=token,
                            guard_column=spec.columns["id"],
                            guard_value=pr.get(spec.column_for("id")),
                            why=(
                                f"looks like the {token} sentinel but is not one, so the "
                                "enricher treats the cell as filled and never fills it"
                            ),
                        )
                    )
                    break
    return out


def _propose_emails(view: WorkbookView, cfg: Config) -> tuple[list[FixupRecord], list[str]]:
    """Strip an annotation from an address. Anything ambiguous goes to the human."""
    spec = cfg.workbook
    columns = [spec.columns["general_email"], spec.columns["sales_b2b_email"]]
    out: list[FixupRecord] = []
    review: list[str] = []
    for column in columns:
        letter = spec.letter_of(column)
        for pr in view.rows:
            value = pr.get(letter)
            if not value or spec.is_empty(value):
                continue
            if _EMAIL_RE.fullmatch(value):
                continue
            if any(
                value.upper().startswith(t.strip().upper()) for t in spec.empty_tokens if t.strip()
            ):
                continue  # a sentinel look-alike; the sentinel rule owns this cell
            found = _EMAIL_RE.findall(value)
            where = f"{letter}{pr.row} ({pr.get(spec.column_for('entity_name'))[:28]})"
            if len(found) == 1:
                out.append(
                    FixupRecord(
                        kind="cell",
                        row=pr.row,
                        column=column,
                        old=value,
                        new=found[0],
                        guard_column=spec.columns["id"],
                        guard_value=pr.get(spec.column_for("id")),
                        why="annotation stripped; the full text stays in CHANGELOG_DETAIL",
                    )
                )
            elif len(found) > 1:
                review.append(
                    f"{where} holds {len(found)} addresses ({', '.join(found)}); choosing "
                    "between them is a GDPR call, not a mechanical one"
                )
            else:
                review.append(f"{where} holds no address at all: {value!r}")
    return out, review


def _propose_phones(view: WorkbookView, cfg: Config) -> tuple[list[FixupRecord], list[str]]:
    """Normalise to E.164 with the row's own region. Never guess a country code."""
    spec = cfg.workbook
    columns = [spec.columns["phone"], spec.columns["whatsapp"]]
    out: list[FixupRecord] = []
    review: list[str] = []
    for column in columns:
        letter = spec.letter_of(column)
        for pr in view.rows:
            value = pr.get(letter)
            if not value or spec.is_empty(value):
                continue
            if re.fullmatch(r"\+[0-9]{7,15}", value):
                continue
            country = pr.get(spec.column_for("country"))
            domain = domain_of(pr.get(spec.column_for("website_url")))
            region = region_for(country, domain, cfg.phone)
            where = f"{letter}{pr.row} ({pr.get(spec.column_for('entity_name'))[:28]})"
            if "wa.me" in value.lower() or value.lower().startswith("http"):
                review.append(f"{where} holds a link, not a number: {value!r}")
                continue
            if _MULTI_PHONE_RE.search(value) or _more_than_one_number(value, region):
                review.append(f"{where} holds more than one number: {value!r}")
                continue
            if not value.strip().startswith("+") and not region:
                review.append(
                    f"{where} is a national number with no country to derive from: {value!r}"
                )
                continue
            parsed = _parse(value, region)
            if parsed is not None and parsed.extension:
                review.append(f"{where} carries an extension, which E.164 cannot hold: {value!r}")
                continue
            if parsed is not None and phonenumbers.number_type(parsed) in _NOT_DIALABLE_ABROAD:
                review.append(
                    f"{where} is a toll-free/premium number, not dialable from abroad: {value!r}"
                )
                continue
            normalised = to_e164(value, region, cfg.phone)
            if not normalised:
                review.append(f"{where} is not a valid dialable number: {value!r}")
                continue
            out.append(
                FixupRecord(
                    kind="cell",
                    row=pr.row,
                    column=column,
                    old=value,
                    new=normalised,
                    guard_column=spec.columns["id"],
                    guard_value=pr.get(spec.column_for("id")),
                    why=f"E.164 (region {region or 'from the number itself'})",
                )
            )
    return out, review


def _parse(value: str, region: str | None):
    """libphonenumber's view of the raw string, or None if it cannot parse it."""
    try:
        return phonenumbers.parse(value, None if value.strip().startswith("+") else region)
    except phonenumbers.NumberParseException:
        return None


def _more_than_one_number(value: str, region: str | None) -> bool:
    try:
        matches = list(phonenumbers.PhoneNumberMatcher(value, region or "ZZ"))
    except Exception:  # pragma: no cover - libphonenumber is defensive already
        return False
    return len(matches) > 1


def propose_fixups(
    view: WorkbookView,
    cfg: Config,
    *,
    reserved_ids: set[str] | None = None,
    block_start: int | None = None,
    include_cells: bool = True,
) -> FixupPlan:
    """Read the workbook and propose everything a corrective release can fix."""
    plan = FixupPlan(
        source=Path("(proposed)"),
        input_name=view.path.name,
        input_sha256=view.fingerprint,
    )
    plan.renumbers = _propose_renumbers(view, cfg, block_start)
    reserved = reserved_ids or set()
    clash = [r for r in plan.renumbers if r.new in reserved]
    if clash:
        highest = max(int(r.new[4:]) for r in plan.renumbers)
        plan.renumbers = _propose_renumbers(view, cfg, highest + 1)
    if include_cells:
        emails, email_review = _propose_emails(view, cfg)
        phones, phone_review = _propose_phones(view, cfg)
        plan.cells = _propose_sentinels(view, cfg) + emails + phones
        plan.review = email_review + phone_review
    return plan


# ---------------------------------------------------------------------------
# The write
# ---------------------------------------------------------------------------


def write_fixups(
    cfg: Config,
    view: WorkbookView,
    plan: FixupPlan,
    *,
    run_id: str,
    output_path: Path | None = None,
    workdir: Path | None = None,
    skip_stale: bool = False,
    handoff: list[str] | None = None,
) -> Path:
    """Apply the plan and emit the next version. Returns the written path."""
    if not plan.records:
        raise VerificationError("nothing to fix: the plan holds no records. Nothing written.")
    spec = cfg.workbook
    source = view.path
    guard_readable(source, spec.min_plausible_bytes)
    if view.fingerprint and file_fingerprint(source) != view.fingerprint:
        raise VerificationError(
            f"The input workbook changed since it was read:\n  {source}\n"
            "Re-run from the current file. Nothing has been written."
        )

    letters = {spec.letter_of(r.column) for r in plan.records}
    assert_touches_no_formula_input(source, cfg, letters)

    destination = output_path or next_version_path(cfg, view.version)
    target_version = parse_version(destination, cfg.output_basename) or view.version + 1
    assert_version_free(cfg, target_version)
    guard_writable(destination)

    staging_dir = workdir or (cfg.state_directory / "staging")
    staging_dir.mkdir(parents=True, exist_ok=True)
    candidate = staging_dir / f"{run_id}_{destination.name}"
    if candidate.exists():
        candidate.unlink()
    shutil.copy2(source, candidate)

    allowed: set[tuple[str, str]] = set()
    sheet_name = spec.sheet
    detail_sheet = spec.changelog_detail_sheet
    run_date = today_iso()
    version_label = destination.stem.rsplit("_", 1)[-1]
    detail_created = False
    now = datetime.now()

    try:
        wb = load_workbook(candidate, data_only=False)
        try:
            ws = wb[sheet_name]
            applied = _assert_old_values_match(ws, spec, plan, skip_stale=skip_stale)
            audit: list[CellChange] = []
            name_letter = spec.column_for("entity_name")
            id_letter = spec.column_for("id")
            for record in applied:
                letter = spec.letter_of(record.column)
                cell = ws.cell(record.row, column_index_from_string(letter))
                cell.value = record.new
                allowed.add((sheet_name, cell.coordinate))
                entity_id = (
                    record.new
                    if record.column == spec.columns["id"]
                    else str(ws.cell(record.row, column_index_from_string(id_letter)).value or "")
                )
                audit.append(
                    CellChange(
                        row=record.row,
                        column=letter,
                        field=record.column,
                        entity_id=entity_id,
                        entity_name=str(
                            ws.cell(record.row, column_index_from_string(name_letter)).value or ""
                        ),
                        old_value=record.old,
                        new_value=record.new,
                        confidence=Confidence.HIGH,
                        data_class=DataClass.CORPORATE_ROLE,
                        source_url=plan.source.name,
                        fetched_at=now,
                        extractor=f"fixup.{record.kind}",
                        note=record.why,
                    )
                )

            parts = []
            if plan.renumbers:
                parts.append(f"renumbered {plan.renumber_summary()}")
            if plan.cells:
                counts = ", ".join(f"{c} x{n}" for c, n in sorted(plan.cell_summary().items()))
                parts.append(f"normalised {len(plan.cells)} cell(s) ({counts})")
            message = (
                f"Corrective release (efe fixup, run {run_id}) from {plan.source.name}: "
                + "; ".join(parts)
                + ". No CRM column, formula or precedent was touched, so the DASHBOARD's "
                "cached totals stay correct for the rows they cover. "
                + (
                    f"{len(plan.review)} item(s) were listed for a human instead. "
                    if plan.review
                    else ""
                )
                + (
                    "Still to do in Sheets: " + "; ".join(h.split(".")[0] for h in handoff) + "."
                    if handoff
                    else ""
                )
            )
            for coordinate in _append_changelog(
                wb[spec.changelog_sheet], version_label, run_date, message, author="efe fixup"
            ):
                allowed.add((spec.changelog_sheet, coordinate))
            coordinates, detail_created = _write_changelog_detail(wb, detail_sheet, audit, run_id)
            for coordinate in coordinates:
                allowed.add((detail_sheet, coordinate))
            wb.save(candidate)
        finally:
            wb.close()

        repaired = reinject_cached_values(candidate, source)
        problems = compare(
            snapshot(source),
            snapshot(candidate),
            allowed_value_changes=allowed,
            allowed_new_sheets={detail_sheet} if detail_created else set(),
            allowed_autofilter_changes={detail_sheet},
            require_cached_values=True,
        )
        if problems:
            raise VerificationError(
                format_report(problems)
                + f"\n\nThe candidate output was DELETED. {source.name} is untouched."
            )
        source_cached = sum(1 for v in read_cached_values(source).values() if v is not None)
        if repaired == 0 and source_cached:
            raise VerificationError(
                "No cached formula results were reinjected although the input carries "
                f"{source_cached}. The candidate output was DELETED."
            )

        assert_version_free(cfg, target_version)
        guard_writable(destination)
        deliver(candidate, destination)
        return destination
    finally:
        if candidate.exists():
            candidate.unlink()
