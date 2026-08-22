"""The fidelity gate.

openpyxl rewrites the entire workbook. This module proves, cell by cell, that the
rewrite changed exactly the cells we intended and nothing else -- 491 formulas, 5
data validations, the autofilter, freeze panes, column widths, number formats and
every untouched value.

Empirically (see `docs/WORKBOOK_NOTES.md`) openpyxl preserves all of that for this
workbook. The one thing it does destroy is the cached `<v>` result of every formula,
which `writer.reinject_cached_values` puts back. This module is what makes that
claim checkable rather than hopeful.
"""

from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from efe.workbook.xmlutil import read_cached_values

#: Parts openpyxl drops on rewrite that provably carry no workbook data.
#:
#: Drawings are deliberately NOT in this list. They are checked by content instead
#: (see `drawing_anchor_counts`): an empty `<xdr:wsDr/>` stub is safe to drop, a
#: drawing holding an actual anchored object is not, and matching on the filename
#: alone would silently accept the second case. This mattered the day the workbook
#: came back re-exported by Google Sheets with nine 775-byte drawing parts instead
#: of three 299-byte ones.
BENIGN_DROPPED_PARTS = re.compile(
    r"^(xl/worksheets/_rels/sheet\d+\.xml\.rels"
    r"|xl/sharedStrings\.xml"          # openpyxl emits inline strings instead
    r"|docProps/custom\.xml"           # empty <Properties/>
    r"|docProps/app\.xml"              # application name and version only
    # Google Sheets' private round-trip blob: locale, timezone and default font.
    # Not the OOXML cell-metadata part (which would be `xl/metadata.xml` and is
    # referenced from sheet cells); this one is inert and Excel ignores it.
    r"|xl/metadata)$"
)

_DRAWING_PART_RE = re.compile(r"^xl/drawings/drawing\d+\.xml$")
_ANCHOR_RE = re.compile(r"<xdr:(?:twoCellAnchor|oneCellAnchor|absoluteAnchor)")



@dataclass
class FidelitySnapshot:
    """Everything worth comparing between two versions of the workbook."""

    sheets: list[str] = field(default_factory=list)
    formulas: dict[tuple[str, str], str] = field(default_factory=dict)
    values: dict[tuple[str, str], Any] = field(default_factory=dict)
    number_formats: dict[tuple[str, str], str] = field(default_factory=dict)
    data_validations: dict[str, list[tuple[str, str, str]]] = field(default_factory=dict)
    autofilter: dict[str, str | None] = field(default_factory=dict)
    freeze_panes: dict[str, str | None] = field(default_factory=dict)
    column_dimensions: dict[str, dict[str, tuple[float | None, bool]]] = field(default_factory=dict)
    defined_names: list[str] = field(default_factory=list)
    cached_values: dict[tuple[str, str], str | None] = field(default_factory=dict)
    parts: set[str] = field(default_factory=set)
    #: drawing part -> number of anchored objects it holds. Zero means an empty
    #: stub, which is safe to drop; anything else is real content.
    drawing_anchor_counts: dict[str, int] = field(default_factory=dict)


def snapshot(path: Path) -> FidelitySnapshot:
    """Read every structural property of a workbook that we promise to preserve."""
    snap = FidelitySnapshot()
    wb = load_workbook(path, data_only=False)
    try:
        snap.sheets = list(wb.sheetnames)
        snap.defined_names = sorted(wb.defined_names.keys())
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    key = (ws.title, cell.coordinate)
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        snap.formulas[key] = cell.value
                    if cell.value is not None:
                        snap.values[key] = cell.value
                        snap.number_formats[key] = cell.number_format
            snap.data_validations[ws.title] = sorted(
                (dv.type or "", str(dv.formula1 or ""), str(dv.sqref))
                for dv in ws.data_validations.dataValidation
            )
            snap.autofilter[ws.title] = ws.auto_filter.ref
            snap.freeze_panes[ws.title] = ws.freeze_panes
            snap.column_dimensions[ws.title] = {
                letter: (dim.width, bool(dim.hidden))
                for letter, dim in ws.column_dimensions.items()
            }
    finally:
        wb.close()

    snap.cached_values = read_cached_values(path)
    with zipfile.ZipFile(path) as zf:
        snap.parts = set(zf.namelist())
        for name in snap.parts:
            if _DRAWING_PART_RE.match(name):
                body = zf.read(name).decode("utf-8", "replace")
                snap.drawing_anchor_counts[name] = len(_ANCHOR_RE.findall(body))
    return snap


def _diff_mapping(
    label: str, a: dict, b: dict, limit: int = 8, ignore=None
) -> list[str]:
    keep = (lambda k: True) if ignore is None else (lambda k: not ignore(k))
    problems: list[str] = []
    missing = sorted(filter(keep, set(a) - set(b)), key=repr)
    added = sorted(filter(keep, set(b) - set(a)), key=repr)
    changed = sorted(
        (k for k in set(a) & set(b) if a[k] != b[k] and keep(k)), key=repr
    )
    if missing:
        problems.append(f"{label}: {len(missing)} lost, e.g. {missing[:limit]}")
    if added:
        problems.append(f"{label}: {len(added)} unexpectedly added, e.g. {added[:limit]}")
    if changed:
        detail = [f"{k}: {a[k]!r} -> {b[k]!r}" for k in changed[:limit]]
        problems.append(f"{label}: {len(changed)} changed, e.g. {detail}")
    return problems


def compare(
    before: FidelitySnapshot,
    after: FidelitySnapshot,
    *,
    allowed_value_changes: set[tuple[str, str]] | None = None,
    allowed_new_sheets: set[str] | None = None,
    require_cached_values: bool = True,
) -> list[str]:
    """Return a list of fidelity problems. An empty list means the output is safe.

    Args:
        before: snapshot of the input workbook.
        after: snapshot of the candidate output.
        allowed_value_changes: (sheet, coordinate) pairs we intended to write. Any
            value change outside this set is a failure.
        allowed_new_sheets: sheets the writer legitimately adds -- in practice just
            CHANGELOG_DETAIL. Everything on them is exempt; every pre-existing sheet
            is still compared in full.
        require_cached_values: assert every formula still carries its cached result.
    """
    allowed = allowed_value_changes or set()
    new_sheets = allowed_new_sheets or set()
    problems: list[str] = []

    def exempt(key) -> bool:
        """True for anything belonging to an intended new sheet or an intended cell."""
        if isinstance(key, tuple):
            return key[0] in new_sheets or key in allowed
        return key in new_sheets

    if before.sheets != after.sheets[: len(before.sheets)]:
        problems.append(
            f"sheet order/name changed: {before.sheets} -> {after.sheets}"
        )

    problems += _diff_mapping("formula", before.formulas, after.formulas, ignore=exempt)
    problems += _diff_mapping(
        "number format", before.number_formats, after.number_formats, ignore=exempt
    )
    problems += _diff_mapping(
        "data validation", before.data_validations, after.data_validations, ignore=exempt
    )
    problems += _diff_mapping("autofilter", before.autofilter, after.autofilter, ignore=exempt)
    problems += _diff_mapping(
        "freeze panes", before.freeze_panes, after.freeze_panes, ignore=exempt
    )
    problems += _diff_mapping(
        "column dimensions", before.column_dimensions, after.column_dimensions, ignore=exempt
    )

    if before.defined_names != after.defined_names:
        problems.append(
            f"defined names changed: {before.defined_names} -> {after.defined_names}"
        )

    # Values: every difference must have been intended.
    unexpected_changed, unexpected_lost = [], []
    for key in set(before.values) | set(after.values):
        old, new = before.values.get(key), after.values.get(key)
        if old == new or exempt(key):
            continue
        (unexpected_lost if new is None else unexpected_changed).append(key)
    if unexpected_lost:
        problems.append(
            f"{len(unexpected_lost)} cell values were lost, e.g. {unexpected_lost[:8]}"
        )
    if unexpected_changed:
        detail = [
            f"{k} {before.values.get(k)!r} -> {after.values.get(k)!r}"
            for k in unexpected_changed[:8]
        ]
        problems.append(
            f"{len(unexpected_changed)} cell values changed without being intended: {detail}"
        )

    if require_cached_values:
        lost = [k for k, v in before.cached_values.items() if v is not None
                and after.cached_values.get(k) is None]
        wrong = [
            (k, v, after.cached_values.get(k))
            for k, v in before.cached_values.items()
            if v is not None and after.cached_values.get(k) not in (None, v)
        ]
        if lost:
            problems.append(
                f"{len(lost)} formulas lost their cached result, e.g. {lost[:8]} "
                "(reinjection did not run or did not cover them)"
            )
        if wrong:
            problems.append(f"{len(wrong)} cached results changed, e.g. {wrong[:5]}")

    dropped = set()
    for part in before.parts - after.parts:
        if BENIGN_DROPPED_PARTS.match(part):
            continue
        if _DRAWING_PART_RE.match(part):
            # Safe only when the drawing held nothing.
            if before.drawing_anchor_counts.get(part, 0) == 0:
                continue
            problems.append(
                f"{part} holds {before.drawing_anchor_counts[part]} anchored object(s) "
                "and was dropped -- that is real content, not a stub"
            )
            continue
        dropped.add(part)
    if dropped:
        problems.append(f"workbook parts dropped: {sorted(dropped)}")

    return problems


def format_report(problems: list[str]) -> str:
    if not problems:
        return "fidelity gate: PASS"
    return "fidelity gate: FAIL\n  - " + "\n  - ".join(problems)
