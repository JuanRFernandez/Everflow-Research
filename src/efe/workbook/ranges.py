"""Where the sheet's range-bound structures end, and what that leaves uncounted.

A workbook is more than its cells: an autofilter covers a rectangle, dropdowns cover
a rectangle, and the DASHBOARD's COUNTIFs read a fixed span of PARTNERS rows. Rows
appended below those rectangles are real data that no filter filters, no dropdown
validates and no total counts.

None of it can be fixed from here. Extending a DASHBOARD range rewrites a formula
whose cached result this tool reinjects from the input, so the emitted file would
claim to count to row 410 while carrying a number that counted to 400; and
`verify.compare` has no vocabulary for a changed data validation, so allowing one
would weaken the gate for every future run. What this module does instead is
measure the drift and say it out loud, in the same words, from every command.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from efe.config import Config
from efe.models import VerificationError

#: `PARTNERS!$C$2:$C$400` and `PARTNERS!A:A`, on any sheet.
_QUALIFIED_RE = re.compile(
    r"(?P<sheet>[A-Za-z_][A-Za-z0-9_]*|'[^']+')!"
    r"\$?(?P<c1>[A-Z]{1,3})\$?(?P<r1>\d*)"
    r"(?::\$?(?P<c2>[A-Z]{1,3})\$?(?P<r2>\d*))?"
)
#: A bare `AA2` / `AA2:AB9` reference, for formulas living on PARTNERS itself.
_BARE_RE = re.compile(
    r"(?<![A-Z0-9_!$])\$?(?P<c1>[A-Z]{1,3})\$?(?P<r1>\d+)"
    r"(?::\$?(?P<c2>[A-Z]{1,3})\$?(?P<r2>\d+))?"
)
#: The criterion cell of a COUNTIF over the category column: `...,A6)`.
_COUNTIF_CRITERION_RE = re.compile(
    r"COUNTIFS?\((?P<args>[^()]*)\)",
    re.I,
)


def _columns_between(first: str, second: str | None) -> set[str]:
    start = _col_index(first)
    end = _col_index(second) if second else start
    if end < start:
        start, end = end, start
    return {get_column_letter(i) for i in range(start, end + 1)}


def _col_index(letter: str) -> int:
    value = 0
    for char in letter:
        value = value * 26 + (ord(char) - ord("A") + 1)
    return value


@dataclass
class SheetRanges:
    """Every range-bound fact about the data sheet, measured from the file."""

    sheet: str
    last_data_row: int
    header_last_col: str
    autofilter_ref: str | None = None
    autofilter_last_row: int | None = None
    autofilter_last_col: str | None = None
    validation_last_row: int | None = None
    #: (sqref as written, last row) for every data validation on the sheet.
    validations: list[tuple[str, int]] = field(default_factory=list)
    #: Highest PARTNERS row any formula on another sheet reads.
    formula_reach_last_row: int | None = None
    #: Category labels the DASHBOARD counts, read from its own COUNTIF criteria.
    dashboard_categories: list[str] = field(default_factory=list)
    #: Category present in PARTNERS that no DASHBOARD row counts -> its row count.
    uncounted_categories: dict[str, int] = field(default_factory=dict)
    #: Formula column -> (dominant formula at the first data row, rows differing).
    formula_shape_gaps: dict[str, tuple[str, list[int]]] = field(default_factory=dict)

    @property
    def rows_beyond_formulas(self) -> int:
        reach = self.formula_reach_last_row
        return 0 if reach is None else max(0, self.last_data_row - reach)

    @property
    def rows_beyond_autofilter(self) -> int:
        last = self.autofilter_last_row
        return 0 if last is None else max(0, self.last_data_row - last)

    @property
    def rows_beyond_validation(self) -> int:
        last = self.validation_last_row
        return 0 if last is None else max(0, self.last_data_row - last)

    def drift(self) -> list[str]:
        """One short line per drifted structure, for `efe check`. Empty == aligned."""
        out: list[str] = []
        if self.autofilter_ref and (
            self.rows_beyond_autofilter
            or (self.autofilter_last_col and self.autofilter_last_col != self.header_last_col)
        ):
            out.append(
                f"autofilter   ends at {self.autofilter_ref}; data reaches row "
                f"{self.last_data_row} and column {self.header_last_col}"
            )
        if self.rows_beyond_validation:
            out.append(
                f"validation   {len(self.validations)} dropdown range(s) end at row "
                f"{self.validation_last_row}; {self.rows_beyond_validation} row(s) have none"
            )
        if self.rows_beyond_formulas:
            out.append(
                f"dashboard    formulas read {self.sheet} to row "
                f"{self.formula_reach_last_row}; {self.rows_beyond_formulas} data row(s) "
                "are counted nowhere"
            )
        for name, count in self.uncounted_categories.items():
            out.append(
                f"categories   {name!r} is on {count} {self.sheet} row(s) and on no "
                "DASHBOARD row; the totals under-count by that many"
            )
        for name, (_, rows) in self.formula_shape_gaps.items():
            out.append(
                f"formula      {len(rows)} row(s) hold a different {name} formula from the "
                f"column's dominant one, e.g. {rows[:6]}"
            )
        return out


def inspect_sheet_ranges(path: Path, cfg: Config) -> SheetRanges:
    """Measure every range-bound structure of the workbook at `path`."""
    spec = cfg.workbook
    wb = load_workbook(path, data_only=False)
    try:
        ws = wb[spec.sheet]
        header = [ws.cell(spec.header_row, c).value for c in range(1, ws.max_column + 1)]
        while header and header[-1] in (None, ""):
            header.pop()
        id_col = _col_index(spec.column_for("id"))
        name_col = _col_index(spec.column_for("entity_name"))
        last_row = spec.header_row
        for row in range(ws.max_row, spec.first_data_row - 1, -1):
            if (
                str(ws.cell(row, id_col).value or "").strip()
                or str(ws.cell(row, name_col).value or "").strip()
            ):
                last_row = row
                break

        ranges = SheetRanges(
            sheet=spec.sheet,
            last_data_row=last_row,
            header_last_col=get_column_letter(max(1, len(header))),
        )

        if ws.auto_filter.ref:
            ranges.autofilter_ref = ws.auto_filter.ref
            _, _, right, bottom = range_boundaries(ws.auto_filter.ref.replace("$", ""))
            ranges.autofilter_last_row = bottom
            ranges.autofilter_last_col = get_column_letter(right)

        for dv in ws.data_validations.dataValidation:
            for rng in str(dv.sqref).split():
                _, _, _, bottom = range_boundaries(rng.replace("$", ""))
                ranges.validations.append((rng, bottom))
                ranges.validation_last_row = (
                    bottom
                    if ranges.validation_last_row is None
                    else max(ranges.validation_last_row, bottom)
                )

        # How far other sheets read into this one, and which categories they count.
        category_letter = spec.column_for("category")
        counted: list[str] = []
        for other in wb.worksheets:
            if other.title == spec.sheet:
                continue
            for row in other.iter_rows():
                for cell in row:
                    if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                        continue
                    for match in _QUALIFIED_RE.finditer(cell.value):
                        if match.group("sheet").strip("'") != spec.sheet:
                            continue
                        for group in ("r1", "r2"):
                            raw = match.group(group)
                            if raw:
                                reach = int(raw)
                                ranges.formula_reach_last_row = (
                                    reach
                                    if ranges.formula_reach_last_row is None
                                    else max(ranges.formula_reach_last_row, reach)
                                )
                    counted += _counted_labels(other, cell.value, spec.sheet, category_letter)
        ranges.dashboard_categories = list(dict.fromkeys(counted))

        listed = set(ranges.dashboard_categories)
        if listed:
            seen: dict[str, int] = {}
            cat_col = _col_index(category_letter)
            for row in range(spec.first_data_row, last_row + 1):
                value = str(ws.cell(row, cat_col).value or "").strip()
                if value and value not in listed:
                    seen[value] = seen.get(value, 0) + 1
            ranges.uncounted_categories = seen

        ranges.formula_shape_gaps = _formula_shape_gaps(ws, spec, last_row)
        return ranges
    finally:
        wb.close()


def _counted_labels(ws, formula: str, sheet: str, category_letter: str) -> list[str]:
    """Category labels a COUNTIF over the category column uses as its criterion.

    Read from the formula's own criterion reference, never a hardcoded list: the
    DASHBOARD is the human's and its rows move.
    """
    out: list[str] = []
    for match in _COUNTIF_CRITERION_RE.finditer(formula):
        args = match.group("args")
        if f"{sheet}!" not in args:
            continue
        target = _QUALIFIED_RE.search(args)
        if target is None or target.group("c1") != category_letter:
            continue
        criterion = args.rsplit(",", 1)[-1].strip()
        if criterion.startswith('"') and criterion.endswith('"'):
            out.append(criterion[1:-1])
            continue
        ref = re.fullmatch(r"\$?([A-Z]{1,3})\$?(\d+)", criterion)
        if ref is None:
            continue
        value = ws.cell(int(ref.group(2)), _col_index(ref.group(1))).value
        if isinstance(value, str) and value.strip():
            out.append(value.strip())
    return out


def _formula_shape_gaps(ws, spec, last_row: int) -> dict[str, tuple[str, list[int]]]:
    """Rows whose formula differs from the column's dominant pattern.

    Cosmetic, never fatal: `=AA9+AB9` and `=IFERROR(AA9+AB9,"")` both satisfy the
    contract, but a column that holds two shapes is a column somebody edited by hand.
    """
    out: dict[str, tuple[str, list[int]]] = {}
    for name in spec.formula_columns:
        letter = spec.letter_of(name)
        col = _col_index(letter)
        origin = f"{letter}{spec.first_data_row}"
        forms: dict[str, list[int]] = {}
        for row in range(spec.first_data_row, last_row + 1):
            value = ws.cell(row, col).value
            if not (isinstance(value, str) and value.startswith("=")):
                continue
            try:
                normal = Translator(value, origin=f"{letter}{row}").translate_formula(origin)
            except (ValueError, TypeError):  # pragma: no cover - malformed formula
                normal = value
            forms.setdefault(normal, []).append(row)
        if len(forms) <= 1:
            continue
        dominant = max(forms, key=lambda k: len(forms[k]))
        others = sorted(r for k, rows in forms.items() if k != dominant for r in rows)
        out[name] = (dominant, others)
    return out


def inspect_ranges(path: Path, cfg: Config) -> dict[str, int | None]:
    """The three numbers `efe promote` reports. Kept as a dict for its callers."""
    ranges = inspect_sheet_ranges(path, cfg)
    return {
        "autofilter_last_row": ranges.autofilter_last_row,
        "validation_last_row": ranges.validation_last_row,
        "formula_reach_last_row": ranges.formula_reach_last_row,
    }


def referenced_columns(path: Path, cfg: Config) -> set[str]:
    """Every column of the data sheet that some formula in the workbook reads.

    Includes bare references from formulas living on the data sheet itself, which is
    where `Next_Follow_Up = Contact_Date + Follow_Up_Days` hides.
    """
    spec = cfg.workbook
    wb = load_workbook(path, data_only=False)
    try:
        columns: set[str] = set()
        for ws in wb.worksheets:
            own = ws.title == spec.sheet
            for row in ws.iter_rows():
                for cell in row:
                    if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                        continue
                    body = cell.value
                    for match in _QUALIFIED_RE.finditer(body):
                        if match.group("sheet").strip("'") == spec.sheet:
                            columns |= _columns_between(match.group("c1"), match.group("c2"))
                    if own:
                        stripped = _QUALIFIED_RE.sub(" ", body)
                        for match in _BARE_RE.finditer(stripped):
                            columns |= _columns_between(match.group("c1"), match.group("c2"))
        return columns
    finally:
        wb.close()


def assert_touches_no_formula_input(path: Path, cfg: Config, letters: set[str]) -> None:
    """Refuse to write a column that some formula reads.

    `writer.assert_no_precedents_touched` asks the same question of `config.yaml`;
    this one asks it of the file, so the day somebody adds `=VLOOKUP(PARTNERS!A:A,...)`
    to the DASHBOARD, the write stops instead of silently preserving a stale result.
    """
    referenced = referenced_columns(path, cfg)
    offending = sorted(letters & referenced)
    if offending:
        names = [cfg.workbook.header_of(letter) for letter in offending]
        raise VerificationError(
            f"Columns {names} ({offending}) are read by a formula in {path.name}. "
            "Writing them would leave the reinjected cached results stale. "
            "Nothing has been written."
        )


def render_sheets_handoff(ranges: SheetRanges, cfg: Config) -> list[str]:
    """The numbered instructions for the human, built from the file's own numbers.

    One function so `check`, `promote` and `fixup` cannot describe the same drift in
    three different ways.
    """
    spec = cfg.workbook
    lines: list[str] = []
    if ranges.autofilter_ref and (
        ranges.rows_beyond_autofilter
        or (ranges.autofilter_last_col and ranges.autofilter_last_col != ranges.header_last_col)
    ):
        lines.append(
            f"Autofilter. Select A1:{ranges.header_last_col}{ranges.last_data_row}, "
            f"Data > Create a filter. (Currently {ranges.autofilter_ref}.)"
        )
    if ranges.rows_beyond_validation:
        spans = ", ".join(
            f"{rng.replace('$', '').split(':')[0]}:"
            f"{re.sub(r'[0-9]+$', str(ranges.last_data_row), rng.replace('$', '').split(':')[-1])}"
            for rng, _ in ranges.validations
        )
        lines.append(
            f"Dropdowns. {len(ranges.validations)} validation(s) end at row "
            f"{ranges.validation_last_row}. Extend each to row {ranges.last_data_row}: {spans}"
        )
    if ranges.rows_beyond_formulas:
        headroom = ((ranges.last_data_row // 100) + 1) * 100
        lines.append(
            f"DASHBOARD ranges. Every formula over {ranges.sheet} reads to row "
            f"{ranges.formula_reach_last_row}; rows "
            f"{ranges.formula_reach_last_row + 1}..{ranges.last_data_row} are counted nowhere. "
            f"On the DASHBOARD: Find and replace, tick 'Search within formulas', "
            f"${ranges.formula_reach_last_row} -> ${headroom} (headroom, so the next "
            "promotion does not reopen this)."
        )
    for name, count in ranges.uncounted_categories.items():
        lines.append(
            f"DASHBOARD categories. {name!r} is on {count} {ranges.sheet} row(s) and on no "
            "DASHBOARD row. Add a row for it and extend the TOTAL sum to include it."
        )
    for name, (dominant, rows) in ranges.formula_shape_gaps.items():
        letter = spec.letter_of(name)
        lines.append(
            f"{name} consistency (cosmetic). {len(rows)} row(s) hold a different formula "
            f"from the column's dominant {dominant!r}. Copy a good cell down over "
            f"{letter}{rows[0]}:{letter}{rows[-1]} to match. No efe command may write this "
            "column - it is a formula column."
        )
    return lines
