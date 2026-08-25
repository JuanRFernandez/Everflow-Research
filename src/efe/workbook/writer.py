"""Guarded writing of the enriched workbook.

The input file is never modified. A copy is built in the repo's gitignored `data/`
directory, verified cell by cell against the input, and only then copied to the
Google Drive folder under a new version number -- one above the version of the file
that was actually read, never a number from config. If verification fails, the
candidate is deleted and the run reports what broke.

Three rules are enforced here rather than trusted to callers:

1. Only columns listed as `writable_columns` (plus the three provenance columns) can
   ever be written. A CRM column or a formula column is a hard error.
2. A cell that already holds a real value is never overwritten.
3. A value with no source URL, no fetch timestamp and no matched text is refused.
"""

from __future__ import annotations

import logging
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string

from efe.config import Config
from efe.models import CellChange, VerificationError, VersionConflictError, today_iso
from efe.workbook.reader import (
    WorkbookView,
    file_fingerprint,
    guard_readable,
    guard_writable,
)
from efe.workbook.resolve import highest_version_present, parse_version
from efe.workbook.verify import compare, format_report, snapshot
from efe.workbook.xmlutil import read_cached_values, reinject_cached_values

log = logging.getLogger(__name__)

CHANGELOG_DETAIL_HEADERS = [
    "Timestamp",
    "Run_ID",
    "Row",
    "Entity_ID",
    "Entity_Name",
    "Column",
    "Field",
    "Old_Value",
    "New_Value",
    "Confidence",
    "Data_Class",
    "Source_URL",
    "Fetched_At",
    "Extractor",
    "Note",
]


# ---------------------------------------------------------------------------
# Output naming
# ---------------------------------------------------------------------------


def next_version_path(cfg: Config, input_version: int, when: str | None = None) -> Path:
    """`YYYY-MM-DD_<basename>_v<input+1>.xlsx` in the output folder.

    The version comes from the file that was actually read -- never from config,
    never from a directory scan. If any file in the workbook or output folder
    already carries that version or a higher one -- a tiny placeholder counts; a
    copy retired as `_SUPERSEDED` does not, it released its number -- writing is
    refused: version numbers only ever go up, and the resolver would read that
    higher file on the next run anyway.
    """
    basename = cfg.output_basename
    if input_version < 1:
        raise VersionConflictError(
            "Cannot derive an output version: the input file name carries no vNN "
            f"(expected <date>_{basename}_vNN.xlsx). Nothing has been written."
        )
    next_version = input_version + 1
    assert_version_free(cfg, next_version)
    stamp = when or today_iso()
    return cfg.output_directory / f"{stamp}_{basename}_v{next_version:02d}.xlsx"


def assert_version_free(cfg: Config, version: int) -> None:
    """Refuse `version` if any file in the workbook/output folders already carries
    it or a higher one. Called when the output name is derived AND again right
    before the copy: a same-version file that lands mid-run must not be shadowed."""
    highest, where = highest_version_present(
        [cfg.workbook_directory, cfg.output_directory], cfg.output_basename
    )
    if where is not None and highest >= version:
        raise VersionConflictError(
            f"Refusing to emit v{version:02d}: v{highest:02d} already exists "
            f"({where}). Version numbers only ever go up; the resolver reads the "
            "highest version on the next run. Nothing has been written."
        )


# ---------------------------------------------------------------------------
# Guards on the change set
# ---------------------------------------------------------------------------


def assert_changes_legal(changes: list[CellChange], cfg: Config, view: WorkbookView) -> None:
    """Refuse a change set that would break any of the three standing rules."""
    spec = cfg.workbook
    writable_only = set(spec.writable_letters)
    writable = writable_only | set(spec.provenance_letters)
    forbidden = set(spec.crm_letters) | set(spec.formula_letters)
    by_row = {pr.row: pr for pr in view.rows}
    problems: list[str] = []

    seen: set[tuple[int, str]] = set()
    for change in changes:
        target = (change.row, change.column)
        label = f"{change.column}{change.row} ({spec.header_of(change.column)})"
        if change.column in forbidden:
            problems.append(
                f"{label} is a human-owned CRM or formula column "
                f"({change.field}) - refusing to write"
            )
            continue
        if change.column not in writable:
            problems.append(f"{label} is not in writable_columns ({change.field})")
            continue
        if target in seen:
            problems.append(f"{label} written twice in one run")
            continue
        seen.add(target)

        row = by_row.get(change.row)
        if row is None:
            problems.append(f"row {change.row} is not a PARTNERS data row")
            continue
        current = row.get(change.column)
        if change.column in writable_only and not spec.is_empty(current):
            problems.append(
                f"{label} already holds {current!r} - refusing to overwrite a non-TBD value"
            )
        if not change.source_url or not change.source_url.startswith(("http://", "https://")):
            problems.append(
                f"{label} has no usable source URL "
                f"({change.source_url!r}) - refusing to write a value without provenance"
            )
        if str(change.new_value).lstrip().startswith("="):
            problems.append(
                f"{label} value {change.new_value!r} starts with '=' and would be stored "
                "as a live formula, not text - refusing"
            )

    if problems:
        joined = "\n  - ".join(problems)
        raise VerificationError("The change set is illegal and was NOT written:\n  - " + joined)


def assert_no_precedents_touched(changes: list[CellChange], cfg: Config) -> None:
    """Confirm no written cell feeds any formula in the workbook.

    This is the premise that makes cached-value reinjection sound. The columns that
    feed formulas (DASHBOARD -> PARTNERS!{Category, Country, Priority_Score,
    Contacted, Status}; Next_Follow_Up -> Contact_Date + Follow_Up_Days) are listed
    as `workbook.formula_precedents`. If a future config makes one of them writable,
    reinjection would preserve a stale result, so the run stops instead.
    """
    spec = cfg.workbook
    precedents = set(spec.precedent_letters)
    offending = sorted({c.column for c in changes} & precedents)
    if offending:
        names = [spec.header_of(letter) for letter in offending]
        raise VerificationError(
            f"Columns {names} ({offending}) feed live formulas. Writing them would "
            "invalidate the cached results this writer preserves. Nothing has been written."
        )


# ---------------------------------------------------------------------------
# Provenance
# ---------------------------------------------------------------------------


def build_provenance_changes(
    changes: list[CellChange], cfg: Config, view: WorkbookView, run_date: str
) -> list[CellChange]:
    """Source_URL/Date_Verified/Round updates for rows that actually changed.

    Untouched rows stay untouched. `Date_Verified` is taken from the evidence, not
    from the clock: it is the date the page carrying the value was actually fetched.
    A long run that crosses midnight, or one replayed from cache days later, would
    otherwise stamp rows with a date on which nothing was verified. `run_date` is
    only the fallback.
    """
    spec = cfg.workbook
    ak = spec.column_for("source_url")
    al = spec.column_for("date_verified")
    am = spec.column_for("round")
    by_row = {pr.row: pr for pr in view.rows}

    urls_by_row: dict[int, list[str]] = defaultdict(list)
    meta_by_row: dict[int, CellChange] = {}
    verified_on: dict[int, str] = {}
    for change in changes:
        if change.source_url not in urls_by_row[change.row]:
            urls_by_row[change.row].append(change.source_url)
        meta_by_row.setdefault(change.row, change)
        fetched = change.fetched_at.date().isoformat()
        # The most recent page that contributed a value to this row.
        if fetched > verified_on.get(change.row, ""):
            verified_on[change.row] = fetched

    out: list[CellChange] = []
    for row_number, urls in sorted(urls_by_row.items()):
        row = by_row[row_number]
        sample = meta_by_row[row_number]

        existing = row.get(ak)
        joined = "; ".join(urls)
        new_source = joined if spec.is_empty(existing) else f"{existing}; {joined}"

        for column, old, new, field in (
            (ak, existing, new_source, "Source_URL"),
            (al, row.get(al), verified_on.get(row_number, run_date), "Date_Verified"),
            (am, row.get(am), cfg.selection.round_tag, "Round"),
        ):
            if str(old) == str(new):
                continue
            out.append(
                CellChange(
                    row=row_number,
                    column=column,
                    field=field,
                    entity_id=sample.entity_id,
                    entity_name=sample.entity_name,
                    old_value=str(old),
                    new_value=str(new),
                    confidence=sample.confidence,
                    data_class=sample.data_class,
                    source_url=sample.source_url,
                    fetched_at=sample.fetched_at,
                    extractor="provenance",
                    note="provenance updated because this row gained at least one value",
                )
            )
    return out


# ---------------------------------------------------------------------------
# The write
# ---------------------------------------------------------------------------


def _last_populated_row(ws, width: int) -> int:
    """Last row holding anything in the first `width` columns.

    `ws.max_row` is the used range: Google Sheets pads every sheet it exports to
    1000 rows, so appending at `max_row + 1` would leave hundreds of blank lines
    between the last entry and the new one.
    """
    for row in range(ws.max_row, 0, -1):
        if any(ws.cell(row, c).value not in (None, "") for c in range(1, width + 1)):
            return row
    return 0


def _append_changelog(ws, version_label: str, run_date: str, message: str) -> list[str]:
    """Append one version-history row after the last real entry; return coordinates."""
    row = _last_populated_row(ws, 4) + 1
    written = []
    for offset, value in enumerate((version_label, run_date, message, "efe enrich (Phase 0)")):
        cell = ws.cell(row, 1 + offset)
        cell.value = value
        cell.alignment = Alignment(vertical="top", wrap_text=offset == 2)
        written.append(cell.coordinate)
    return written


def _write_changelog_detail(
    wb, sheet_name: str, records: list[CellChange], run_id: str
) -> tuple[list[str], bool]:
    """Audit rows: one per cell change and per held-back candidate.

    CHANGELOG_DETAIL is a required sheet, so rows are appended after its last real
    entry with header and layout untouched. Creating it is kept for a workbook that
    predates the audit sheet, should the requirement ever be relaxed. Returns
    (coordinates written, created).
    """
    created = sheet_name not in wb.sheetnames
    written: list[str] = []
    if created:
        ws = wb.create_sheet(sheet_name)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F3864")
        for index, title in enumerate(CHANGELOG_DETAIL_HEADERS, start=1):
            cell = ws.cell(1, index)
            cell.value = title
            cell.font = header_font
            cell.fill = header_fill
            written.append(cell.coordinate)
        start = 2
    else:
        ws = wb[sheet_name]
        found = [ws.cell(1, i).value for i in range(1, len(CHANGELOG_DETAIL_HEADERS) + 1)]
        if found != CHANGELOG_DETAIL_HEADERS:
            raise VerificationError(
                f"{sheet_name} exists but its header is not the audit layout this "
                f"writer produces:\n  expected {CHANGELOG_DETAIL_HEADERS}\n  found    {found}\n"
                "Refusing to append misaligned rows. Nothing has been written."
            )
        start = _last_populated_row(ws, len(CHANGELOG_DETAIL_HEADERS)) + 1

    for offset, change in enumerate(records, start=start):
        values = [
            datetime.now().isoformat(timespec="seconds"),
            run_id,
            change.row,
            change.entity_id,
            change.entity_name,
            change.column,
            change.field,
            change.old_value,
            change.new_value,
            change.confidence.value,
            change.data_class.value,
            change.source_url,
            change.fetched_at.isoformat(timespec="seconds"),
            change.extractor,
            change.note,
        ]
        for index, value in enumerate(values, start=1):
            cell = ws.cell(offset, index)
            cell.value = value
            if isinstance(value, str) and value.startswith("="):
                cell.data_type = "s"  # audit text, never a formula
            written.append(cell.coordinate)

    if created:
        widths = {
            "A": 20, "B": 14, "C": 6, "D": 11, "E": 34, "F": 7, "G": 22, "H": 16,
            "I": 34, "J": 11, "K": 15, "L": 46, "M": 20, "N": 20, "O": 52,
        }  # fmt: skip
        for letter, width in widths.items():
            ws.column_dimensions[letter].width = width
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:O{max(1, len(records) + 1)}"
    elif ws.auto_filter.ref and records:
        # Keep the audit filter covering the rows just appended.
        ws.auto_filter.ref = f"A1:O{start + len(records) - 1}"
    return written, created


def write_enriched(
    cfg: Config,
    view: WorkbookView,
    changes: list[CellChange],
    *,
    run_id: str,
    held_back: list[CellChange] | None = None,
    changelog_message: str = "",
    output_path: Path | None = None,
    workdir: Path | None = None,
) -> Path:
    """Emit a verified, enriched copy of the workbook. Returns the written path.

    The output version is the input's plus one (`view.version`), unless an explicit
    `output_path` is given.

    Raises:
        VerificationError: the candidate failed the fidelity gate and was discarded.
        VersionConflictError: the next version already exists somewhere.
        WorkbookLockedError / DriveSyncError: the environment is not safe to write.
    """
    source = view.path
    guard_readable(source, cfg.workbook.min_plausible_bytes)
    if view.fingerprint and file_fingerprint(source) != view.fingerprint:
        raise VerificationError(
            "The input workbook changed since it was read:"
            + chr(10)
            + f"  {source}"
            + chr(10)
            + "A Drive sync or a new download landed mid-run. The changes computed from "
            "the old content would be applied to rows this run never looked at. "
            "Re-run from the current file. Nothing has been written."
        )

    destination = output_path or next_version_path(cfg, view.version)
    target_version = parse_version(destination, cfg.output_basename) or view.version + 1
    assert_version_free(cfg, target_version)
    guard_writable(destination)

    run_date = today_iso()
    assert_changes_legal(changes, cfg, view)
    assert_no_precedents_touched(changes, cfg)

    provenance = build_provenance_changes(changes, cfg, view, run_date)
    assert_changes_legal(changes + provenance, cfg, view)

    staging_dir = workdir or (cfg.state_directory / "staging")
    staging_dir.mkdir(parents=True, exist_ok=True)
    candidate = staging_dir / f"{run_id}_{destination.name}"

    if candidate.exists():
        candidate.unlink()
    shutil.copy2(source, candidate)

    allowed: set[tuple[str, str]] = set()
    sheet_name = cfg.workbook.sheet
    detail_sheet = cfg.workbook.changelog_detail_sheet
    version_label = destination.stem.rsplit("_", 1)[-1]
    detail_created = False

    try:
        wb = load_workbook(candidate, data_only=False)
        try:
            ws = wb[sheet_name]
            for change in changes + provenance:
                cell = ws.cell(change.row, column_index_from_string(change.column))
                cell.value = change.new_value
                allowed.add((sheet_name, cell.coordinate))

            filled = len(changes)
            rows_touched = len({c.row for c in changes})
            message = changelog_message or (
                f"Contact enrichment (efe enrich, run {run_id}). "
                f"{filled} cells filled across {rows_touched} rows from published "
                f"company pages; every value carries a source URL and fetch date in "
                f"{detail_sheet}. "
                f"{len(held_back or [])} candidates held for review, not written. "
                "No CRM column, pre-existing value or formula was modified."
            )
            for coordinate in _append_changelog(
                wb[cfg.workbook.changelog_sheet], version_label, run_date, message
            ):
                allowed.add((cfg.workbook.changelog_sheet, coordinate))

            detail_records = sorted(
                (changes + provenance + list(held_back or [])),
                key=lambda c: (c.row, c.column, c.field),
            )
            coordinates, detail_created = _write_changelog_detail(
                wb, detail_sheet, detail_records, run_id
            )
            for coordinate in coordinates:
                allowed.add((detail_sheet, coordinate))

            wb.save(candidate)
        finally:
            wb.close()

        repaired = reinject_cached_values(candidate, source)

        problems = compare(
            snapshot(source),
            snapshot(candidate),
            allowed_value_changes=allowed,
            allowed_new_sheets={detail_sheet} if detail_created else set(),
            allowed_autofilter_changes={detail_sheet},
            require_cached_values=True,
        )
        if problems:
            raise VerificationError(
                format_report(problems)
                + f"\n\nThe candidate output was DELETED. {source.name} is untouched."
            )
        # The tripwire is the SOURCE's cached results, not its formula count: a file
        # last written by openpyxl carries none, and then there is nothing to lose.
        source_cached = sum(1 for v in read_cached_values(source).values() if v is not None)
        if repaired == 0 and source_cached:
            raise VerificationError(
                "No cached formula results were reinjected although the input carries "
                f"{source_cached}. The candidate output was DELETED."
            )
        if not source_cached:
            log.warning(
                "%s carries no cached formula results (last written by %s); the output "
                "carries none either. Sheets and Excel recompute them on open.",
                source.name,
                view.last_writer or "unknown",
            )

        assert_version_free(cfg, target_version)
        guard_writable(destination)
        shutil.copy2(candidate, destination)
        return destination
    finally:
        if candidate.exists():
            candidate.unlink()
